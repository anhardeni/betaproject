import openpyxl
import sys

file_path = "/home/acer25/frappe-bench/base erp xls/00002372017220250925000113.xlsx"

try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb["ENTITAS"]
    
    headers = []
    for cell in ws[1]:
        if cell.value:
            h_norm = " ".join(str(cell.value).strip().upper().split())
            headers.append(h_norm)
            
    print(f"Normalized Headers: {headers}")

    nomor_col = "NOMOR IJIN ENTITAS"
    tanggal_col = "TANGGAL IJIN ENTITAS"
    
    nomor_idx = headers.index(nomor_col) if nomor_col in headers else -1
    tanggal_idx = headers.index(tanggal_col) if tanggal_col in headers else -1
    
    print(f"Indices: NOMOR={nomor_idx}, TANGGAL={tanggal_idx}")
    
    if nomor_idx == -1:
        print("CRITICAL: NOMOR IJIN ENTITAS not found in headers!")
        sys.exit(1)

    print("\nInspecting first 10 rows:")
    row_count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row_count >= 10: break
        
        nomor_val = row[nomor_idx] if nomor_idx != -1 else "N/A"
        tanggal_val = row[tanggal_idx] if tanggal_idx != -1 else "N/A"
        
        print(f"Row {row_count+2}:")
        print(f"  NOMOR   : '{nomor_val}' (Type: {type(nomor_val)})")
        print(f"  TANGGAL : '{tanggal_val}' (Type: {type(tanggal_val)})")
        row_count += 1

except Exception as e:
    print(f"Error: {e}")
