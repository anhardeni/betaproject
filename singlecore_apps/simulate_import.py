import frappe
import sys
import os
import base64
import openpyxl
from singlecore_apps.api.ceisa_import import import_ceisa_excel

FILE_PATH = "/home/acer25/frappe-bench/base erp xls/00002372017220250925000113.xlsx"

def run_simulation():
    try:
        print(f"--- Starting Import Simulation for {os.path.basename(FILE_PATH)} ---")
        
        # 1. Read and Encode File
        if not os.path.exists(FILE_PATH):
            print(f"Error: File not found at {FILE_PATH}")
            return
            
        with open(FILE_PATH, "rb") as f:
            content = f.read()
            # Prepare mock input: "data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,....."
            b64_content = base64.b64encode(content).decode('utf-8')
            file_data = f"data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64_content}"

        # 2. Inspect Excel Manually for Validation Baseline
        wb = openpyxl.load_workbook(FILE_PATH, data_only=True)
        
        # Helper to count non-empty rows excluding header
        def count_excel_rows(sheet_name):
            if sheet_name not in wb.sheetnames: return 0
            ws = wb[sheet_name]
            count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(row): count += 1
            return count

        # Get Header info to find the doc
        h_ws = wb["HEADER"]
        headers = [str(c.value).strip().upper() for c in h_ws[1] if c.value]
        # Find NOMOR AJU index
        nomor_aju_idx = -1
        for i, h in enumerate(headers):
             if "NOMOR AJU" in h: 
                 nomor_aju_idx = i
                 break
        
        nomor_aju = None
        for row in h_ws.iter_rows(min_row=2, values_only=True):
            if nomor_aju_idx != -1: nomor_aju = row[nomor_aju_idx]
            break # Just need first row
            
        print(f"Target NOMOR AJU: {nomor_aju}")
        
        excel_counts = {
            "ENTITAS": count_excel_rows("ENTITAS"),
            "KEMASAN": count_excel_rows("KEMASAN"),
            "DOKUMEN": count_excel_rows("DOKUMEN"),
            "PENGANGKUT": count_excel_rows("PENGANGKUT"),
            "BARANG": count_excel_rows("BARANG")
        }
        
        print("\n[Baseline] Rows found in Excel:")
        for k, v in excel_counts.items():
            print(f"  {k}: {v}")

        # 3. Run Import
        print("\n>>> Running import_ceisa_excel()...")
        result = import_ceisa_excel(file_data)
        
        if result.get("status") == "error":
            print(f"\n[FAIL] Import returned error: {result.get('message')}")
            # Even on error, audit might show stats
            if result.get("audit"):
                print("Audit dump:", result.get("audit"))
        else:
            print("\n[SUCCESS] Import Function returned success.")
            print(result.get("message").replace("<br>", "\n"))

            # 4. Verification Check in DB
            print(f"\n>>> Verifying Database Records for {nomor_aju}...")
            
            # Check Header
            hdr = frappe.get_doc("HEADER V21", nomor_aju)
            print(f"  HEADER V21 found: {hdr.name}")
            
            # Check Child Tables
            db_counts = {
                "ENTITAS": len(hdr.entitas),
                "KEMASAN": len(hdr.kemasan),
                "DOKUMEN": len(hdr.dokumen),
                "PENGANGKUT": len(hdr.pengangkut),
                # Barang is separate
                "BARANG": frappe.db.count("BARANG V1", {"nomoraju": nomor_aju})
            }
            
            print("\n[Comparison] Excel vs Database:")
            discrepancies = []
            for k in excel_counts.keys():
                xls_c = excel_counts.get(k, 0)
                db_c = db_counts.get(k, 0)
                status = "MATCH" if xls_c == db_c else "** MISMATCH **"
                print(f"  {k:10}: Excel={xls_c} -> DB={db_c}  [{status}]")
                
                if xls_c != db_c:
                    discrepancies.append(f"{k} (Missing {xls_c - db_c} records)")

            if discrepancies:
                print("\n[WARNING] DATA LOSS DETECTED:")
                for d in discrepancies:
                    print(f"  - {d}")
                print("Possible causes: Empty required fields, duplicate primary keys (seri), or unmapped rows.")
            else:
                print("\n[PERFECT] All data imported successfully.")
                
                # Deep Check for User's specific concern
                print("\n>>> Deep Content Check (ENTITAS):")
                match_entitas = False
                for e in hdr.entitas:
                     print(f"  - Kode: {e.kode_entitas} | Ijin: '{e.nomor_ijin_entitas}' | Tgl Ijin: '{e.tanggal_ijin_entitas}'")
                     if e.nomor_ijin_entitas: match_entitas = True
                
                if match_entitas:
                     print("[CONFIRMED] NOMOR IJIN ENTITAS was saved successfully.")
                else:
                     print("[WARNING] NOMOR IJIN ENTITAS is empty/None in all records!")

    except Exception as e:
        print(f"\n[CRASH] Simulation failed with exception: {e}")
        import traceback
        traceback.print_exc()
    
    finally:
        print("\n>>> Rolling back transaction (Simulation Mode)...")
        frappe.db.rollback()
        frappe.destroy()

if __name__ == "__main__":
    run_simulation()
