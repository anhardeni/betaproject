import frappe
import openpyxl

def main():
    """Debug script to check why BAHANBAKU is not being imported"""
    
    # Use one of the test files
    file_path = "/home/acer25/frappe-bench/sites/dens9.com/private/files/00026272017220251017000063.xlsx"
    
    print(f"\n=== Debugging BAHANBAKU Import ===")
    print(f"File: {file_path}")
    
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        
        # Check if BAHANBAKU sheet exists
        if "BAHANBAKU" not in wb.sheetnames:
            print("ERROR: Sheet 'BAHANBAKU' not found!")
            print(f"Available sheets: {wb.sheetnames}")
            return
        
        ws = wb["BAHANBAKU"]
        
        # Get headers
        headers = [cell.value for cell in ws[1]]
        print(f"\nHeaders in BAHANBAKU sheet: {headers}")
        
        # Count rows with data
        row_count = 0
        sample_rows = []
        for row in ws.iter_rows(min_row=2, max_row=10, values_only=True):
            item = dict(zip(headers, row))
            if any(item.values()):
                row_count += 1
                sample_rows.append(item)
        
        print(f"\nSample rows (first up to 10 with data): {row_count}")
        for i, row in enumerate(sample_rows[:3]):
            print(f"  Row {i+1}: SERI BARANG={row.get('SERI BARANG')}, SERI BAHAN BAKU={row.get('SERI BAHAN BAKU')}, HS={row.get('HS')}")
        
        # Check BARANG sheet for matching SERI BARANG
        if "BARANG" in wb.sheetnames:
            ws_barang = wb["BARANG"]
            barang_headers = [cell.value for cell in ws_barang[1]]
            barang_seri = set()
            for row in ws_barang.iter_rows(min_row=2, values_only=True):
                item = dict(zip(barang_headers, row))
                if item.get("SERI BARANG"):
                    barang_seri.add(item.get("SERI BARANG"))
            print(f"\nSERI BARANG values in BARANG sheet: {sorted(list(barang_seri)[:10])}...")
        
        # Check if BAHAN BAKU DocType exists
        print("\n=== Checking BAHAN BAKU DocType ===")
        try:
            meta = frappe.get_meta("BAHAN BAKU")
            print(f"DocType 'BAHAN BAKU' exists with {len(meta.fields)} fields")
            
            # Check required fields
            required_fields = ["nomoraju", "seri_barang", "seri_bahan_baku", "parent_barang"]
            for field in required_fields:
                if meta.has_field(field):
                    print(f"  ✓ Field '{field}' exists")
                else:
                    print(f"  ✗ Field '{field}' MISSING!")
                    
        except Exception as e:
            print(f"ERROR: DocType 'BAHAN BAKU' does not exist! {e}")
            
    except Exception as e:
        print(f"ERROR: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()
