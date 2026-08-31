import frappe
import openpyxl
import base64
import io
from frappe.utils import getdate, flt, cint
from datetime import datetime, date

@frappe.whitelist(allow_guest=True)
def import_ceisa_excel(file_data, dry_run=False):
    audit_report = {
        "unmapped_columns": {},
        "missing_columns": {},
        "empty_fields": {},  # Track which fields had empty values
        "stats": {}
    }

    def save_doc(d):
        d.flags.ignore_links = True
        d.flags.ignore_permissions = True
        d.save(ignore_permissions=True)

    try:
        # Decode file
        if file_data.startswith("/private/files/") or file_data.startswith("/files/"):
            if file_data.startswith("/private/"):
                file_path = frappe.get_site_path(file_data.strip("/"))
            else:
                # /files/ points to public/files/
                file_path = frappe.get_site_path("public", file_data.strip("/"))
            with open(file_path, "rb") as f:
                decoded_file = f.read()
        else:
            if "," in file_data:
                file_data = file_data.split(",")[1]
            decoded_file = base64.b64decode(file_data)

        wb = openpyxl.load_workbook(io.BytesIO(decoded_file), data_only=True)
        
        # Helper: Get Sheet Data
        def get_sheet_data(sheet_name, optional=False, expected_columns=None):
            if sheet_name not in wb.sheetnames:
                if not optional:
                    frappe.throw(f"Sheet {sheet_name} not found")
                return []
            
            ws = wb[sheet_name]
            rows = list(ws.values)
            if not rows: return []
            
            # Helper to normalize header
            def normalize(h):
                if not h: return ""
                return " ".join(str(h).strip().upper().split())

            headers = [normalize(c) for c in rows[0]]
            
            # Audit
            if expected_columns:
                present = set(headers)
                # Filter out empty headers
                present = {x for x in present if x}

                if isinstance(expected_columns, dict):
                    # Smart Audit with Aliases
                    mapping = expected_columns
                    expected_keys = set(mapping.keys())
                    
                    unmapped = [h for h in present if h not in expected_keys]
                    
                    # Missing by Target Field
                    found_targets = {mapping[h] for h in present if h in mapping}
                    all_targets = set(mapping.values())
                    missing_targets = all_targets - found_targets
                    
                    # Reverse map for display
                    reverse_map = {}
                    for k, v in mapping.items():
                        reverse_map.setdefault(v, []).append(k)
                    
                    missing = [reverse_map[mt][0] for mt in missing_targets]

                else:
                    # Legacy List Audit
                    expected = set(expected_columns)
                    unmapped = [ue for ue in present if ue not in expected]
                    missing = [mex for mex in expected if mex not in present]
                
                if unmapped:
                    audit_report["unmapped_columns"][sheet_name] = list(unmapped)
                if missing:
                    audit_report["missing_columns"][sheet_name] = missing

            data_list = []
            for row in rows[1:]:
                row_dict = {}
                has_data = False
                for idx, val in enumerate(row):
                    if idx < len(headers):
                        row_dict[headers[idx]] = val
                        if val: has_data = True
                if has_data:
                    data_list.append(row_dict)
            return data_list

        DATE_FIELDS = [
            "tanggal_bc11", "tanggal_berangkat", "tanggal_ekspor", "tanggal_masuk",
            "tanggal_muat", "tanggal_tiba", "tanggal_periksa", "tanggal_stuffing",
            "tanggal_pernyataan", "tanggal_bukti_bayar", "tanggal_daftar",
            "tanggal_ijin_entitas", "tanggal_dokumen", "tanggal_jaminan", 
            "tanggal_jatuh_tempo", "tanggal_bpj", "tanggal_daftar_asal",
            "jatuh_tempo_royalti", "tanggal_respon"
        ]

        def parse_excel_date(val):
            if not val: return None
            if isinstance(val, (datetime, date)): return val
            
            # Handle Excel date serial numbers
            if isinstance(val, (int, float)):
                try:
                    from openpyxl.utils.datetime import from_excel
                    return from_excel(val).date()
                except Exception:
                    pass
            
            val_str = str(val).strip()
            # If it's a float/int string representing excel date number (e.g. "45293.0")
            try:
                float_val = float(val_str)
                from openpyxl.utils.datetime import from_excel
                return from_excel(float_val).date()
            except ValueError:
                pass

            formats = [
                "%Y-%m-%d %H:%M:%S", "%Y-%m-%d", 
                "%d-%m-%Y %H:%M:%S", "%d-%m-%Y",
                "%m-%d-%Y %H:%M:%S", "%m-%d-%Y",
                "%d/%m/%Y %H:%M:%S", "%d/%m/%Y",
                "%m/%d/%Y %H:%M:%S", "%m/%d/%Y",
                "%Y/%m/%d %H:%M:%S", "%Y/%m/%d"
            ]
            for fmt in formats:
                try:
                    return datetime.strptime(val_str, fmt).date()
                except ValueError:
                    continue
            return None

        def clean_excel_val(doctype, fieldname, val):
            if val is None or val == "": return None
            
            if fieldname in DATE_FIELDS:
                return parse_excel_date(val)
                
            meta = frappe.get_meta(doctype)
            field = meta.get_field(fieldname)
            if not field: return val
            
            if field.fieldtype in ["Link", "Select", "Data", "Text", "Small Text", "Long Text"]:
                # Clean numeric values from Excel (e.g. 1.0 -> "1")
                if isinstance(val, (int, float)):
                    if isinstance(val, float) and val.is_integer():
                        return str(cint(val))
                    return str(val)
                return str(val).strip()
                
            if field.fieldtype in ["Float", "Int", "Currency", "Percent"]:
                return flt(val)
                
            return val

        # 1. HEADER MAPPING
        HEADER_MAPPING = {
            "NOMOR AJU": "nomoraju",
            "KODE DOKUMEN": "kode_dokumen",
            "KODE KANTOR": "kode_kantor",
            "KOTA PERNYATAAN": "kota_pernyataan",
            "TANGGAL PERNYATAAN": "tanggal_pernyataan",
            "NAMA PERNYATAAN": "nama_pernyataan",
            "JABATAN PERNYATAAN": "jabatan_pernyataan",
            "KODE KANTOR BONGKAR": "kode_kantor_bongkar",
            "KODE KANTOR PERIKSA": "kode_kantor_periksa",
            "KODE KANTOR TUJUAN": "kode_kantor_tujuan",
            "KODE KANTOR EKSPOR": "kode_kantor_ekspor",
            #"KODE JENIS PIB": "kode_jenis_pib", # Sometimes not in Excel?
            "KODE JENIS EKSPOR": "kode_jenis_ekspor",
            "KODE JENIS TPB": "kode_jenis_tpb",
            "KODE JENIS PLB": "kode_jenis_plb",
            "KODE JENIS IMPOR": "kode_jenis_impor",
            "KODE TUJUAN PEMASUKAN": "kode_tujuan_pemasukan",
            "KODE TUJUAN PENGIRIMAN": "kode_tujuan_pengiriman",
            "KODE TUJUAN TPB": "kode_tujuan_tpb",
            "KODE CARA DAGANG": "kode_cara_dagang",
            "KODE CARA BAYAR": "kode_cara_bayar",
            "KODE CARA BAYAR LAINNYA": "kode_cara_bayar_lainnya",
            "KODE GUDANG ASAL": "kode_gudang_asal",
            "KODE GUDANG TUJUAN": "kode_gudang_tujuan",
            "KODE JENIS KIRIM": "kode_jenis_kirim",
            "KODE JENIS PENGIRIMAN": "kode_jenis_pengiriman",
            "KODE KATEGORI EKSPOR": "kode_kategori_ekspor",
            "KODE KATEGORI MASUK FTZ": "kode_kategori_masuk_ftz",
            "KODE KATEGORI KELUAR FTZ": "kode_kategori_keluar_ftz",
            "KODE KATEGORI BARANG FTZ": "kode_kategori_barang_ftz",
            "KODE LOKASI": "kode_lokasi",
            "KODE LOKASI BAYAR": "kode_lokasi_bayar",
            "LOKASI ASAL": "lokasi_asal",
            "LOKASI TUJUAN": "lokasi_tujuan",
            "KODE DAERAH ASAL": "kode_daerah_asal",
            "KODE NEGARA TUJUAN": "kode_negara_tujuan",
            "KODE TUTUP PU": "kode_tutup_pu",
            "NOMOR BC11": "nomor_bc11",
            "NOMOR BC 11": "nomor_bc11",  # Alias with space
            "NO BC11": "nomor_bc11",  # Alias short form
            "TANGGAL BC11": "tanggal_bc11",
            "TANGGAL BC 11": "tanggal_bc11",  # Alias with space
            "TGL BC11": "tanggal_bc11",  # Alias short form
            "NOMOR POS": "nomor_pos",
            "NO POS": "nomor_pos",  # Alias short form
            "NOMOR SUB POS": "nomor_sub_pos",
            "NO SUB POS": "nomor_sub_pos",  # Alias short form
            "NOMOR SUBPOS": "nomor_sub_pos",  # Alias no space
            "KODE PELABUHAN BONGKAR": "kode_pelabuhan_bongkar",
            "KODE PELABUHAN MUAT": "kode_pelabuhan_muat",
            "KODE PELABUHAN MUAT AKHIR": "kode_pelabuhan_muat_akhir",
            "KODE PELABUHAN TRANSIT": "kode_pelabuhan_transit",
            "KODE PELABUHAN TUJUAN": "kode_pelabuhan_tujuan",
            "KODE PELABUHAN EKSPOR": "kode_pelabuhan_ekspor",
            "KODE TPS": "kode_tps",
            "TANGGAL BERANGKAT": "tanggal_berangkat",
            "TANGGAL EKSPOR": "tanggal_ekspor",
            "TANGGAL MASUK": "tanggal_masuk",
            "TANGGAL MUAT": "tanggal_muat",
            "TANGGAL TIBA": "tanggal_tiba",
            "TANGGAL PERIKSA": "tanggal_periksa",
            "TEMPAT STUFFING": "tempat_stuffing",
            "TANGGAL STUFFING": "tanggal_stuffing",
            "KODE TANDA PENGAMAN": "kode_tanda_pengaman",
            "JUMLAH TANDA PENGAMAN": "jumlah_tanda_pengaman",
            "FLAG CURAH": "flag_curah",
            "FLAG SDA": "flag_sda",
            "FLAG VD": "flag_vd",
            "FLAG MIGAS": "flag_migas",
            "KODE ASURANSI": "kode_asuransi",
            "ASURANSI": "asuransi",
            "NILAI BARANG": "nilai_barang",
            "NILAI INCOTERM": "nilai_incoterm",
            "NILAI MAKLON": "nilai_maklon",
            "FREIGHT": "freight",
            "FOB": "fob",
            "BIAYA TAMBAHAN": "biaya_tambahan",
            "BIAYA PENGURANG": "biaya_pengurang",
            "VD": "vd",
            "CIF": "cif",
            "HARGA_PENYERAHAN": "harga_penyerahan",
            "NDPBM": "ndpbm",
            "TOTAL DANA SAWIT": "total_dana_sawit",
            "DASAR PENGENAAN PAJAK": "dasar_pengenaan_pajak",
            "NILAI JASA": "nilai_jasa",
            "UANG MUKA": "uang_muka",
            "BRUTO": "bruto",
            "NETTO": "netto",
            "VOLUME": "volume",
            "KODE VALUTA": "kode_valuta",
            "KODE INCOTERM": "kode_incoterm",
            "KODE JASA KENA PAJAK": "kode_jasa_kena_pajak",
            "NOMOR BUKTI BAYAR": "nomor_bukti_bayar",
            "TANGGAL BUKTI BAYAR": "tanggal_bukti_bayar",
            "KODE JENIS NILAI": "kode_jenis_nilai",
            "KODE KANTOR MUAT": "kode_kantor_muat",
            "NOMOR DAFTAR": "nomor_daftar",
            "TANGGAL DAFTAR": "tanggal_daftar",
            "KODE ASAL BARANG FTZ": "kode_barang_asal_ftz",
            "KODE TUJUAN PENGELUARAN": "kode_tujuan_pengeluaran",
            "PPN PAJAK": "ppn_pajak",
            "PPNBM PAJAK": "ppnbm_pajak",
            "TARIF PPN PAJAK": "tarif_ppn_pajak",
            "TARIF PPNBM PAJAK": "tarif_ppnbm_pajak",
            "BARANG TIDAK BERWUJUD": "barang_tidak_berwujud",
            # Additional Mappings
            "FLAG KONSOL": "flag_konsol",
            "FLAG PROPORSIONAL NETTO": "flag_proporsional_netto",
            "FLAG AP BK": "flag_ap_bk", # If exists
            "KODE JENIS PENGELUARAN": "kode_jenis_pengeluaran",
            "KODE JENIS PROSEDUR": "kode_jenis_prosedur",
            "KODE JENIS PENGANGKUTAN": "kode_jenis_pengangkutan",
            "BARANG KIRIMAN": "barang_kiriman"
        }

        # Fetch HEADER data
        header_rows = get_sheet_data("HEADER", expected_columns=HEADER_MAPPING)
        if not header_rows:
            frappe.throw("No data in HEADER sheet")
        
        header_row = header_rows[0]
        nomor_aju = header_row.get("NOMOR AJU")
        if not nomor_aju:
            frappe.throw("NOMOR AJU is missing in HEADER")
        # Normalize nomor_aju to string for consistent comparison
        nomor_aju = str(nomor_aju).strip()

        # Get/Create Header Doc
        if frappe.db.exists("HEADER V21", {"nomoraju": nomor_aju}):
            doc = frappe.get_doc("HEADER V21", {"nomoraju": nomor_aju})
        else:
            doc = frappe.new_doc("HEADER V21")
            doc.nomoraju = nomor_aju
            doc.name = nomor_aju  # Set document name to NOMOR AJU
            doc.flags.name_set = True  # Prevent autoname from overriding

        # Map Header
        for excel_col, doc_field in HEADER_MAPPING.items():
            if excel_col in header_row:
                val = header_row.get(excel_col)
                doc.set(doc_field, clean_excel_val("HEADER V21", doc_field, val))

        # Helper for Child Tables
        def create_child(doctype, parent_field, sheet_name, mapping, optional=False):
            rows = get_sheet_data(sheet_name, optional, expected_columns=mapping)
            if not rows and optional: return
            
            child_list = []
            for row in rows:
                row_nomor_aju = str(row.get("NOMOR AJU") or "").strip()
                if row_nomor_aju != nomor_aju: continue
                child_item = {}
                for excel_col, doc_field in mapping.items():
                    if excel_col in row:
                        val = row.get(excel_col)
                        child_item[doc_field] = clean_excel_val(doctype, doc_field, val)
                child_list.append(child_item)
            
            doc.set(parent_field, child_list)
            audit_report["stats"][doctype] = len(child_list)

        # CHILD TABLES IMPORT
        
        # ENTITAS
        create_child("ENTITAS", "entitas", "ENTITAS", {
            "NOMOR AJU": "nomoraju",
            "SERI": "seri",
            "KODE ENTITAS": "kode_entitas",
            #"KODE JENIS ENTITAS": "kode_jenis_entitas",
            "NOMOR IDENTITAS": "nomor_identitas",
            "NAMA ENTITAS": "nama_entitas",
            "ALAMAT ENTITAS": "alamat_entitas",
            "NIB ENTITAS": "nib_entitas",
            "KODE JENIS API": "kode_jenis_api",
            "KODE STATUS": "kode_status",
            "KODE NEGARA": "kode_negara",
            "NOMOR IJIN ENTITAS": "nomor_ijin_entitas",
            "TANGGAL IJIN ENTITAS": "tanggal_ijin_entitas",
            "KODE JENIS IDENTITAS": "kode_jenis_identitas",
            "NIPER ENTITAS": "niper_entitas",
            "KODE AFILIASI": "kode_afiliasi",
            "KODE KATEGORI KONSOLIDATOR": "kode_kategori_konsolidator"
        })

        # KEMASAN
        create_child("KEMASAN", "kemasan", "KEMASAN", {
            "NOMOR AJU": "nomoraju",
            "SERI": "seri",
            "KODE KEMASAN": "kode_kemasan",
            "JUMLAH KEMASAN": "jumlah_kemasan",
            "MERK KEMASAN": "merek_kemasan",
            "MEREK": "merek_kemasan", 
            "NOMOR SEGEL": "no_segel_kemasan"
        })

        # DOKUMEN
        create_child("DOKUMEN", "dokumen", "DOKUMEN", {
            "NOMOR AJU": "nomoraju",
            "SERI DOKUMEN": "seri",
            "SERI": "seri", # Alias
            "KODE DOKUMEN": "kode_dokumen",
            "NOMOR DOKUMEN": "nomor_dokumen",
            "TANGGAL DOKUMEN": "tanggal_dokumen",
            "KODE FASILITAS": "kode_fasilitas",
            "KODE IJIN": "kode_ijin"
        })

        # PENGANGKUT
        create_child("PENGANGKUT", "pengangkut", "PENGANGKUT", {
            "NOMOR AJU": "nomoraju",
            "SERI PENGANGKUT": "seri_pengangkut",
            "SERI": "seri_pengangkut", # Alias
            "KODE CARA ANGKUT": "kode_cara_angkut",
            "NAMA PENGANGKUT": "nama_pengangkut",
            "NOMOR PENGANGKUT": "nomor_pengangkut",
            "KODE BENDERA": "kode_bendera",
            "CALL SIGN": "call_sign",
            "FLAG ANGKUT PLB": "flag_angkut_plb",
            "CARA PENGANGKUTAN LAINNYA": "cara_pengangkutan_lainnya"
        })
        
        # KONTAINER
        create_child("KONTAINER", "kontainer", "KONTAINER", {
            "NOMOR AJU": "nomoraju",
            "SERI KONTAINER": "seri",
            "SERI": "seri", # Alias
            "NOMOR KONTAINER": "nomor_kontainer",
            "NOMOR KONTINER": "nomor_kontainer", # Typo handling
            "KODE UKURAN KONTAINER": "kode_ukuran_kontainer",
            "KODE JENIS KONTAINER": "kode_jenis_kontainer",
            "KODE TIPE KONTAINER": "kode_tipe_kontainer",
            "NOMOR SEGEL": "nomor_segel_kontainer"
        })

        # KOMPONEN BIAYA (Optional)
        create_child("komponen_biaya", "komponen_biaya", "KOMPONENBIAYA", {
            "NOMOR AJU": "nomoraju",
            "JENIS NILAI": "jenisnilai",
            "HARGA INVOICE": "hargainvoice",
            "PEMBAYARAN TIDAK LANGSUNG": "pembayarantidaklangsung",
            "DISKON": "diskon",
            "KOMISI PENJUALAN": "komisipenjualan",
            "BIAYA PENGEMASAN": "biayapengemasan",
            "BIAYA PENGEPAKAN": "biayapengepakan",
            "ASSIST": "assist",
            "ROYALTI": "royalti",
            "PROCEEDS": "proceeds",
            "BIAYA TRANSPORTASI": "biayatransportasi",
            "BIAYA PEMUATAN": "biayapemuatan",
            "ASURANSI": "asuransi",
            "GARANSI": "garansi",
            "BIAYA KEPENTINGAN SENDIRI": "biayakepentingansendiri",
            "BIAYA PASCA IMPOR": "biayapascaimpor",
            "BIAYA PAJAK INTERNAL": "biayapajakinternal",
            "BUNGA": "bunga",
            "DEVIDEN": "deviden"
        }, optional=True)

        # PUNGUTAN & JAMINAN & BANK DEVISA (Using new logic if needed, referencing previous conv)
        create_child("pungutan", "pungutan", "PUNGUTAN", {
            "NOMOR AJU": "nomoraju",
            "KODE FASILITAS TARIF": "kode_fasilitas_tarif",
            "KODE JENIS PUNGUTAN": "kode_jenis_pungutan",
            "NILAI PUNGUTAN": "nilai_pungutan",
            "NPWP BILLING": "npwp_billing"
        })

        create_child("jaminan", "jaminan", "JAMINAN", {
            "NOMOR AJU": "nomoraju",
            "KODE JAMINAN": "kode_jaminan",
            "NOMOR BPJ": "nomor_bpj",
            "TANGGAL BPJ": "tanggal_bpj",
            "NILAI JAMINAN": "nilai_jaminan",
            "TANGGAL JATUH TEMPO": "tanggal_jatuh_tempo",
            "PENJAMIN": "penjamin",
            "NOMOR JAMINAN": "nomor_jaminan",
            "TANGGAL JAMINAN": "tanggal_jaminan",
            "KODE KANTOR": "kode_kantor" 
        })

        create_child("bank_devisa", "bank_devisa", "BANKDEVISA", {
            "NOMOR AJU": "nomoraju",
            "KODE": "kode",
            "NAMA": "nama",
            "SERI": "seri"
        })

        # RESPON Sheet (JSON Dump)
        respon_rows = get_sheet_data("RESPON", optional=True)
        if respon_rows:
            import json
            doc.respon_json = json.dumps(respon_rows, default=str)

        # --- BARANG PROCESSING ---
        barang_rows = get_sheet_data("BARANG")
        
        BARANG_MAPPING = {
            "NOMOR AJU": "nomoraju",
            "SERI BARANG": "seri_barang",
            "HS": "hs",
            "KODE BARANG": "kode_barang",
            "URAIAN": "uraian",
            "MEREK": "merek",
            "MERK": "merek", # Alias
            "TIPE": "tipe",
            "UKURAN": "ukuran",
            "SPESIFIKASI LAIN": "spesifikasi_lain",
            "KODE SATUAN": "kode_satuan",
            "JUMLAH SATUAN": "jumlah_satuan",
            "JUMLAH BAHAN BAKU": "jumlah_bahan_baku",
            "KODE KEMASAN": "kode_kemasan",
            "JUMLAH KEMASAN": "jumlah_kemasan",
            "NETTO": "netto",
            "BRUTO": "bruto",
            "VOLUME": "volume",
            "CIF": "cif",
            "CIF RUPIAH": "cif_rupiah",
            "NDPBM": "ndpbm",
            "FOB": "fob",
            "ASURANSI": "asuransi",
            "FREIGHT": "freight",
            "NILAI BARANG": "nilai_barang",
            "NILAI JASA": "nilai_jasa",
            "NILAI DANA SAWIT": "nilai_dana_sawit",
            "NILAI DEVISA": "nilai_devisa",
            "PERSENTASE IMPOR": "persentase_impor",
            "DISKON": "diskon",
            "HARGA PENYERAHAN": "harga_penyerahan",
            "HARGA PEROLEHAN": "harga_perolehan",
            "HARGA SATUAN": "harga_satuan",
            "HARGA EKSPOR": "harga_ekspor",
            "HARGA PATOKAN": "harga_patokan",
            "NILAI TAMBAH": "nilai_tambah",
            "PERNYATAAN LARTAS": "pernyataan_lartas",
            "TAHUN PEMBUATAN": "tahun_pembuatan",
            "KODE JENIS EKSPOR": "kode_jenis_ekspor",
            "KODE KATEGORI BARANG": "kode_kategori_barang",
            "KODE KONDISI BARANG": "kode_kondisi_barang",
            "SERI IZIN": "seri_izin",
            "KODE ASAL BARANG": "kode_asal_barang",
            "KODE NEGARA ASAL": "kode_negara_asal",
            "KODE DAERAH ASAL": "kode_daerah_asal",
            "STATEMENT PERBEDAAN HARGA": "statement_perbedaan_harga",
            "SALDO AWAL": "saldo_awal",
            "KAPASITAS SILINDER": "kapasitas_silinder",
            "JATUH TEMPO ROYALTI": "jatuh_tempo_royalti",
            "FLAG TIS": "flag_tis",
            "KODE KANTOR ASAL": "kode_kantor_asal",
            "ISI PER KEMASAN": "isi_per_kemasan",
            "FLAG 4 TAHUN": "flag_4_tahun",
            "KODE DOKUMEN ASAL": "kode_dokumen_asal",
            "KODE BKC": "kode_bkc",
            "KODE SUB KOMODITI BKC": "kode_sub_komoditi_bkc",
            "NOMOR AJU ASAL": "nomor_aju_asal",
            "KODE GUNA BARANG": "kode_guna_barang",
            "SERI BARANG ASAL": "seri_barang_asal",
            "SALDO AKHIR": "saldo_akhir",
            "KODE PERHITUNGAN": "kode_perhitungan",
            "KODE JENIS NILAI": "kode_jenis_nilai",
            "KODE KOMODITI BKC": "kode_komoditi_bkc",
            "NOMOR DAFTAR ASAL": "nomor_daftar_asal",
            "JUMLAH REALISASI": "jumlah_realisasi",
            "TANGGAL DAFTAR ASAL": "tanggal_daftar_asal",
            "METODE PENENTUAN NILAI": "metode_penentuan_nilai",
            "HJE CUKAI": "hje_cukai",
            "TARIF CUKAI": "tarif_cukai",
            "JUMLAH PITA CUKAI": "jumlah_pita_cukai",
            "JUMLAH DILEKATKAN": "jumlah_dilekatkan"
        }

        # Sheets for Barang Children
        bt_rows = get_sheet_data("BARANGTARIF")
        bd_rows = get_sheet_data("BARANGDOKUMEN")
        be_rows = get_sheet_data("BARANGENTITAS") # Maps to barang_pemilik
        bvd_rows = get_sheet_data("BARANGVD")
        bspe_rows = get_sheet_data("BARANGSPEKKHUSUS")
        
        # BAHAN BAKU Sheets
        bb_rows = get_sheet_data("BAHANBAKU")
        bbt_rows = get_sheet_data("BAHANBAKUTARIF")
        bbd_rows = get_sheet_data("BAHANBAKUDOKUMEN")

        # Close workbook as early as possible since we are done reading Excel sheets
        wb.close()

        # Save Header 
        save_doc(doc)

        # Pre-group child rows for O(1) lookups
        bt_by_seri = {}
        for r in bt_rows:
            sb = cint(r.get("SERI BARANG"))
            bt_by_seri.setdefault(sb, []).append(r)

        bd_by_seri = {}
        for r in bd_rows:
            sb = cint(r.get("SERI BARANG"))
            bd_by_seri.setdefault(sb, []).append(r)

        be_by_seri = {}
        for r in be_rows:
            sb = cint(r.get("SERI BARANG"))
            be_by_seri.setdefault(sb, []).append(r)

        bspe_by_seri = {}
        for r in bspe_rows:
            sb = cint(r.get("SERI BARANG"))
            bspe_by_seri.setdefault(sb, []).append(r)

        bvd_by_seri = {}
        for r in bvd_rows:
            sb = cint(r.get("SERI BARANG"))
            bvd_by_seri.setdefault(sb, []).append(r)

        bb_by_seri = {}
        for r in bb_rows:
            sb = cint(r.get("SERI BARANG"))
            bb_by_seri.setdefault(sb, []).append(r)

        bbt_by_key = {}
        for r in bbt_rows:
            key = (cint(r.get("SERI BARANG")), cint(r.get("SERI BAHAN BAKU")))
            bbt_by_key.setdefault(key, []).append(r)

        bbd_by_key = {}
        for r in bbd_rows:
            key = (cint(r.get("SERI BARANG")), cint(r.get("SERI BAHAN BAKU")))
            bbd_by_key.setdefault(key, []).append(r)

        audit_report["stats"]["BARANG V1"] = 0
        
        for b_row in barang_rows:
            row_nomor_aju = str(b_row.get("NOMOR AJU") or "").strip()
            if row_nomor_aju and row_nomor_aju != nomor_aju: continue

            seri_barang = cint(b_row.get("SERI BARANG"))
            if not seri_barang: continue  # Skip rows with no seri_barang
            
            # Check exist - use doc.name (HEADER V21 document name) for the Link field
            existing_b = frappe.get_all("BARANG V1", filters={
                "nomoraju": doc.name, "seri_barang": seri_barang
            })
            
            if existing_b:
                b_doc = frappe.get_doc("BARANG V1", existing_b[0].name)
            else:
                b_doc = frappe.new_doc("BARANG V1")
            
            # Map Barang
            for excel_col, doc_field in BARANG_MAPPING.items():
                if excel_col in b_row:
                    val = b_row.get(excel_col)
                    b_doc.set(doc_field, clean_excel_val("BARANG V1", doc_field, val))
            
            # Ensure b_doc is linked to correct HEADER V21 (in case Excel has mismatched nomoraju)
            b_doc.nomoraju = doc.name
            
            # Child Tables for BARANG
            
            # Tarif
            child_bt = []
            for r in bt_by_seri.get(seri_barang, []):
                child_bt.append({
                    "seri_barang": seri_barang,
                    "kode_pungutan": clean_excel_val("BARANG TARIF", "kode_pungutan", r.get("KODE PUNGUTAN")),
                    "kode_tarif": clean_excel_val("BARANG TARIF", "kode_tarif", r.get("KODE TARIF")),
                    "tarif": clean_excel_val("BARANG TARIF", "tarif", r.get("TARIF")),
                    "kode_fasilitas": clean_excel_val("BARANG TARIF", "kode_fasilitas", r.get("KODE FASILITAS")),
                    "tarif_fasilitas": clean_excel_val("BARANG TARIF", "tarif_fasilitas", r.get("TARIF FASILITAS")),
                    "nilai_bayar": clean_excel_val("BARANG TARIF", "nilai_bayar", r.get("NILAI BAYAR")),
                    "nilai_fasilitas": clean_excel_val("BARANG TARIF", "nilai_fasilitas", r.get("NILAI FASILITAS")),
                    "nilai_sudah_dilunasi": clean_excel_val("BARANG TARIF", "nilai_sudah_dilunasi", r.get("NILAI SUDAH DILUNASI")),
                    "kode_komoditi_cukai": clean_excel_val("BARANG TARIF", "kode_komoditi_cukai", r.get("KODE KOMODITI CUKAI")),
                    "kode_sub_komoditi_cukai": clean_excel_val("BARANG TARIF", "kode_sub_komoditi_cukai", r.get("KODE SUB KOMODITI CUKAI")),
                    "jumlah_satuan": clean_excel_val("BARANG TARIF", "jumlah_satuan", r.get("JUMLAH SATUAN")),
                    "kode_satuan": clean_excel_val("BARANG TARIF", "kode_satuan", r.get("KODE SATUAN"))
                })
            b_doc.set("barang_tarif", child_bt)
            
            # Dokumen
            child_bd = []
            for r in bd_by_seri.get(seri_barang, []):
                child_bd.append({
                    "seri_dokumen": clean_excel_val("BARANG DOKUMEN", "seri_dokumen", r.get("SERI DOKUMEN")),
                    "seri_izin": clean_excel_val("BARANG DOKUMEN", "seri_izin", r.get("SERI IZIN"))
                })
            b_doc.set("barang_dokumen", child_bd)
            
            # Pemilik (Entitas)
            child_be = []
            for r in be_by_seri.get(seri_barang, []):
                child_be.append({
                    "seri_entitas": clean_excel_val("BARANG ENTITAS", "seri_entitas", r.get("SERI ENTITAS"))
                })
            b_doc.set("barang_pemilik", child_be)
            
            # Spek Khusus
            child_sp = []
            for r in bspe_by_seri.get(seri_barang, []):
                child_sp.append({
                    "kode": clean_excel_val("BARANG SPEK KHUSUS", "kode", r.get("KODE")),
                    "uraian": clean_excel_val("BARANG SPEK KHUSUS", "uraian", r.get("URAIAN"))
                })
            b_doc.set("barang_spek_khusus", child_sp)
 
            # VD
            child_vd = []
            for r in bvd_by_seri.get(seri_barang, []):
                child_vd.append({
                    "kode_jenis_vd": clean_excel_val("BARANG VD", "kode_jenis_vd", r.get("KODE VD")),
                    "nilai_barang": clean_excel_val("BARANG VD", "nilai_barang", r.get("NILAI BARANG")),
                    "biaya_tambahan": clean_excel_val("BARANG VD", "biaya_tambahan", r.get("BIAYA TAMBAHAN")),
                    "biaya_pengurang": clean_excel_val("BARANG VD", "biaya_pengurang", r.get("BIAYA PENGURANG")),
                    "jatuh_tempo": clean_excel_val("BARANG VD", "jatuh_tempo", r.get("JATUH TEMPO"))
                })
            b_doc.set("barang_vd", child_vd)

            save_doc(b_doc)
            audit_report["stats"]["BARANG V1"] += 1
            
            # Track child table stats
            audit_report["stats"]["BARANG TARIF"] = audit_report["stats"].get("BARANG TARIF", 0) + len(child_bt)
            audit_report["stats"]["BARANG DOKUMEN"] = audit_report["stats"].get("BARANG DOKUMEN", 0) + len(child_bd)
            audit_report["stats"]["BARANG PEMILIK"] = audit_report["stats"].get("BARANG PEMILIK", 0) + len(child_be)
            audit_report["stats"]["BARANG SPEK KHUSUS"] = audit_report["stats"].get("BARANG SPEK KHUSUS", 0) + len(child_sp)
            audit_report["stats"]["BARANG VD"] = audit_report["stats"].get("BARANG VD", 0) + len(child_vd)
            
            # --- BAHAN BAKU --- (Linked to BARANG)
            for bb_row in bb_by_seri.get(seri_barang, []):
                seri_bahan_baku = cint(bb_row.get("SERI BAHAN BAKU"))
                if not seri_bahan_baku: continue
                
                filters = {
                    "nomoraju": nomor_aju, 
                    "seri_barang": seri_barang,
                    "seri_bahan_baku": seri_bahan_baku
                }
                existing_bb = frappe.get_all("BAHAN BAKU", filters=filters)
                
                if existing_bb:
                    bb_doc = frappe.get_doc("BAHAN BAKU", existing_bb[0].name)
                else:
                    bb_doc = frappe.new_doc("BAHAN BAKU")
                    bb_doc.update(filters)
                    bb_doc.parent_barang = b_doc.name
                
                # Map Bahan Baku
                bb_doc.hs = clean_excel_val("BAHAN BAKU", "hs", bb_row.get("HS"))
                bb_doc.kode_barang = clean_excel_val("BAHAN BAKU", "kode_barang", bb_row.get("KODE BARANG"))
                bb_doc.uraian = clean_excel_val("BAHAN BAKU", "uraian", bb_row.get("URAIAN"))
                bb_doc.merek = clean_excel_val("BAHAN BAKU", "merek", bb_row.get("MEREK"))
                bb_doc.tipe = clean_excel_val("BAHAN BAKU", "tipe", bb_row.get("TIPE"))
                bb_doc.ukuran = clean_excel_val("BAHAN BAKU", "ukuran", bb_row.get("UKURAN"))
                bb_doc.spesifikasi_lain = clean_excel_val("BAHAN BAKU", "spesifikasi_lain", bb_row.get("SPESIFIKASI LAIN"))
                bb_doc.kode_satuan = clean_excel_val("BAHAN BAKU", "kode_satuan", bb_row.get("KODE SATUAN"))
                bb_doc.jumlah_satuan = clean_excel_val("BAHAN BAKU", "jumlah_satuan", bb_row.get("JUMLAH SATUAN"))
                bb_doc.kode_asal_bahan_baku = clean_excel_val("BAHAN BAKU", "kode_asal_bahan_baku", bb_row.get("KODE ASAL BAHAN BAKU"))
                bb_doc.cif = clean_excel_val("BAHAN BAKU", "cif", bb_row.get("CIF"))
                bb_doc.cif_rupiah = clean_excel_val("BAHAN BAKU", "cif_rupiah", bb_row.get("CIF RUPIAH"))
                bb_doc.harga_penyerahan = clean_excel_val("BAHAN BAKU", "harga_penyerahan", bb_row.get("HARGA PENYERAHAN"))
                bb_doc.harga_perolehan = clean_excel_val("BAHAN BAKU", "harga_perolehan", bb_row.get("HARGA PEROLEHAN"))
                bb_doc.ndpbm = clean_excel_val("BAHAN BAKU", "ndpbm", bb_row.get("NDPBM"))
                bb_doc.netto = clean_excel_val("BAHAN BAKU", "netto", bb_row.get("NETTO"))
                bb_doc.bruto = clean_excel_val("BAHAN BAKU", "bruto", bb_row.get("BRUTO"))
                bb_doc.volume = clean_excel_val("BAHAN BAKU", "volume", bb_row.get("VOLUME"))
                
                # New fields
                bb_doc.kode_bkc = clean_excel_val("BAHAN BAKU", "kode_bkc", bb_row.get("KODE BKC"))
                bb_doc.kode_komoditi_bkc = clean_excel_val("BAHAN BAKU", "kode_komoditi_bkc", bb_row.get("KODE KOMODITI BKC"))
                bb_doc.kode_sub_komoditi_bkc = clean_excel_val("BAHAN BAKU", "kode_sub_komoditi_bkc", bb_row.get("KODE SUB KOMODITI BKC"))
                bb_doc.flag_tis = clean_excel_val("BAHAN BAKU", "flag_tis", bb_row.get("FLAG TIS"))
                bb_doc.isi_per_kemasan = clean_excel_val("BAHAN BAKU", "isi_per_kemasan", bb_row.get("ISI PER KEMASAN"))
                bb_doc.jumlah_dilekatkan = clean_excel_val("BAHAN BAKU", "jumlah_dilekatkan", bb_row.get("JUMLAH DILEKATKAN"))
                bb_doc.jumlah_pita_cukai = clean_excel_val("BAHAN BAKU", "jumlah_pita_cukai", bb_row.get("JUMLAH PITA CUKAI"))
                bb_doc.hje_cukai = clean_excel_val("BAHAN BAKU", "hje_cukai", bb_row.get("HJE CUKAI"))
                bb_doc.tarif_cukai = clean_excel_val("BAHAN BAKU", "tarif_cukai", bb_row.get("TARIF CUKAI"))
                bb_doc.nomor_aju_asal = clean_excel_val("BAHAN BAKU", "nomor_aju_asal", bb_row.get("NOMOR AJU ASAL"))
                bb_doc.nomor_daftar_asal = clean_excel_val("BAHAN BAKU", "nomor_daftar_asal", bb_row.get("NOMOR DAFTAR ASAL"))
                bb_doc.tanggal_daftar_asal = clean_excel_val("BAHAN BAKU", "tanggal_daftar_asal", bb_row.get("TANGGAL DAFTAR ASAL"))
                bb_doc.kode_dokumen_asal = clean_excel_val("BAHAN BAKU", "kode_dokumen_asal", bb_row.get("KODE DOKUMEN ASAL"))
                bb_doc.kode_kantor_asal = clean_excel_val("BAHAN BAKU", "kode_kantor_asal", bb_row.get("KODE KANTOR ASAL"))
                
                # Children of Bahan Baku
                key = (seri_barang, seri_bahan_baku)
                
                # BB Tarif
                child_bbt = []
                for r in bbt_by_key.get(key, []):
                    child_bbt.append({
                        "kode_pungutan": clean_excel_val("BAHAN BAKU TARIF", "kode_pungutan", r.get("KODE PUNGUTAN")),
                        "kode_tarif": clean_excel_val("BAHAN BAKU TARIF", "kode_tarif", r.get("KODE TARIF")),
                        "tarif": clean_excel_val("BAHAN BAKU TARIF", "tarif", r.get("TARIF")),
                        "kode_fasilitas": clean_excel_val("BAHAN BAKU TARIF", "kode_fasilitas", r.get("KODE FASILITAS")),
                        "tarif_fasilitas": clean_excel_val("BAHAN BAKU TARIF", "tarif_fasilitas", r.get("TARIF FASILITAS")),
                        "nilai_bayar": clean_excel_val("BAHAN BAKU TARIF", "nilai_bayar", r.get("NILAI BAYAR")),
                        "nilai_fasilitas": clean_excel_val("BAHAN BAKU TARIF", "nilai_fasilitas", r.get("NILAI FASILITAS")),
                        "kode_asal_bahan_baku": clean_excel_val("BAHAN BAKU TARIF", "kode_asal_bahan_baku", r.get("KODE ASAL BAHAN BAKU")),
                        "jumlah_satuan": clean_excel_val("BAHAN BAKU TARIF", "jumlah_satuan", r.get("JUMLAH SATUAN")),
                        "kode_satuan": clean_excel_val("BAHAN BAKU TARIF", "kode_satuan", r.get("KODE SATUAN"))
                    })
                bb_doc.set("bahan_tarif", child_bbt)
                
                # BB Dokumen
                child_bbd = []
                for r in bbd_by_key.get(key, []):
                    child_bbd.append({
                        "seri_dokumen": clean_excel_val("BAHAN BAKU DOKUMEN", "seri_dokumen", r.get("SERI DOKUMEN")),
                        "seri_izin": clean_excel_val("BAHAN BAKU DOKUMEN", "seri_izin", r.get("SERI IZIN")),
                        "kode_asal_bahan_baku": clean_excel_val("BAHAN BAKU DOKUMEN", "kode_asal_bahan_baku", r.get("KODE ASAL BAHAN BAKU"))
                    })
                bb_doc.set("bahan_baku_dokumen", child_bbd)
                
                save_doc(bb_doc)
                
                # Track BAHAN BAKU stats
                audit_report["stats"]["BAHAN BAKU"] = audit_report["stats"].get("BAHAN BAKU", 0) + 1
                audit_report["stats"]["BAHAN BAKU TARIF"] = audit_report["stats"].get("BAHAN BAKU TARIF", 0) + len(child_bbt)
                audit_report["stats"]["BAHAN BAKU DOKUMEN"] = audit_report["stats"].get("BAHAN BAKU DOKUMEN", 0) + len(child_bbd)
        
        message = f"<b>Successfully processed {nomor_aju}</b>"
        
        # Add statistics to message
        if audit_report["stats"]:
            message += "<br><br><b>📊 Import Statistics:</b><br>"
            for table, count in audit_report["stats"].items():
                message += f"- {table}: {count} records<br>"
        
        # Add unmapped columns warning (data in Excel NOT inserted)
        if audit_report["unmapped_columns"]:
            message += "<br><b>⚠️ Unmapped Columns (data in Excel NOT inserted):</b><br>"
            for sheet, cols in audit_report["unmapped_columns"].items():
                message += f"- {sheet}: {', '.join(cols)}<br>"
        
        # Add missing columns info
        if audit_report["missing_columns"]:
            message += "<br><b>ℹ️ Missing Columns (expected but not in Excel):</b><br>"
            for sheet, cols in audit_report["missing_columns"].items():
                message += f"- {sheet}: {', '.join(cols)}<br>"
        
        if cint(dry_run):
            frappe.db.rollback()
            return {"status": "success", "message": "[DRY RUN] " + message, "audit": audit_report, "nomor_aju": nomor_aju}
        
        frappe.db.commit()
        return {"status": "success", "message": message, "audit": audit_report, "nomor_aju": nomor_aju}

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Import CEISA Excel Refactored Error")
        error_msg = f"<b>Error during import:</b> {str(e)}"
        
        # Append stats to error message
        if audit_report["stats"]:
           error_msg += "<br><br><b>Partial Statistics:</b><br>"
           for k, v in audit_report["stats"].items():
                error_msg += f"- {k}: {v}<br>"
                
        return {"status": "error", "message": error_msg, "audit": audit_report}


@frappe.whitelist(allow_guest=True)
def import_ceisa_excel_v2(file_data, dry_run=False):
    """
    Optimized Version 2 of CEISA Excel Importer.
    Uses bulk pre-fetching and dirty-checking to eliminate N+1 DB queries and unnecessary writes.
    """
    audit_report = {
        "unmapped_columns": {},
        "missing_columns": {},
        "empty_fields": {},
        "stats": {
            "BARANG V1": 0,
            "BARANG TARIF": 0,
            "BARANG DOKUMEN": 0,
            "BARANG PEMILIK": 0,
            "BARANG SPEK KHUSUS": 0,
            "BARANG VD": 0,
            "BAHAN BAKU": 0,
            "BAHAN BAKU TARIF": 0,
            "BAHAN BAKU DOKUMEN": 0,
            "saves_skipped": 0,
            "saves_performed": 0
        }
    }

    def save_doc(d):
        d.flags.ignore_links = True
        d.flags.ignore_permissions = True
        d.save(ignore_permissions=True)

    def are_child_tables_identical(child_list, doc_child_table, fields):
        if len(child_list) != len(doc_child_table):
            return False
        for idx, item in enumerate(child_list):
            db_item = doc_child_table[idx]
            for f in fields:
                val_new = str(item.get(f) if item.get(f) is not None else "")
                val_old = str(db_item.get(f) if db_item.get(f) is not None else "")
                if val_new != val_old:
                    return False
        return True

    try:
        # Decode file
        if file_data.startswith("/private/files/") or file_data.startswith("/files/"):
            if file_data.startswith("/private/"):
                file_path = frappe.get_site_path(file_data.strip("/"))
            else:
                file_path = frappe.get_site_path("public", file_data.strip("/"))
            with open(file_path, "rb") as f:
                decoded_file = f.read()
        else:
            if "," in file_data:
                file_data = file_data.split(",")[1]
            decoded_file = base64.b64decode(file_data)

        wb = openpyxl.load_workbook(io.BytesIO(decoded_file), data_only=True)
        
        # Helper: Get Sheet Data
        def get_sheet_data(sheet_name, optional=False, expected_columns=None):
            if sheet_name not in wb.sheetnames:
                if not optional:
                    frappe.throw(f"Sheet {sheet_name} not found")
                return []
            
            ws = wb[sheet_name]
            rows = list(ws.values)
            if not rows: return []
            
            def normalize(h):
                if not h: return ""
                return " ".join(str(h).strip().upper().split())

            headers = [normalize(c) for c in rows[0]]
            
            if expected_columns:
                present = {x for x in headers if x}
                if isinstance(expected_columns, dict):
                    mapping = expected_columns
                    expected_keys = set(mapping.keys())
                    unmapped = [h for h in present if h not in expected_keys]
                    found_targets = {mapping[h] for h in present if h in mapping}
                    all_targets = set(mapping.values())
                    missing_targets = all_targets - found_targets
                    reverse_map = {}
                    for k, v in mapping.items():
                        reverse_map.setdefault(v, []).append(k)
                    missing = [reverse_map[mt][0] for mt in missing_targets]
                else:
                    expected = set(expected_columns)
                    unmapped = [ue for ue in present if ue not in expected]
                    missing = [mex for mex in expected if mex not in present]
                
                if unmapped:
                    audit_report["unmapped_columns"][sheet_name] = list(unmapped)
                if missing:
                    audit_report["missing_columns"][sheet_name] = missing

            data_list = []
            for row in rows[1:]:
                row_dict = {}
                has_data = False
                for idx, val in enumerate(row):
                    if idx < len(headers):
                        row_dict[headers[idx]] = val
                        if val: has_data = True
                if has_data:
                    data_list.append(row_dict)
            return data_list

        DATE_FIELDS = [
            "tanggal_bc11", "tanggal_berangkat", "tanggal_ekspor", "tanggal_masuk",
            "tanggal_muat", "tanggal_tiba", "tanggal_periksa", "tanggal_stuffing",
            "tanggal_pernyataan", "tanggal_bukti_bayar", "tanggal_daftar",
            "tanggal_ijin_entitas", "tanggal_dokumen", "tanggal_jaminan", 
            "tanggal_jatuh_tempo", "tanggal_bpj", "tanggal_daftar_asal",
            "jatuh_tempo_royalti", "tanggal_respon"
        ]

        def parse_excel_date(val):
            if not val: return None
            if isinstance(val, (datetime, date)): return val
            if isinstance(val, (int, float)):
                try:
                    from openpyxl.utils.datetime import from_excel
                    return from_excel(val).date()
                except Exception:
                    pass
            
            val_str = str(val).strip()
            try:
                float_val = float(val_str)
                from openpyxl.utils.datetime import from_excel
                return from_excel(float_val).date()
            except ValueError:
                pass

            formats = [
                "%Y-%m-%d %H:%M:%S", "%Y-%m-%d", 
                "%d-%m-%Y %H:%M:%S", "%d-%m-%Y",
                "%m-%d-%Y %H:%M:%S", "%m-%d-%Y",
                "%d/%m/%Y %H:%M:%S", "%d/%m/%Y",
                "%m/%d/%Y %H:%M:%S", "%m/%d/%Y",
                "%Y/%m/%d %H:%M:%S", "%Y/%m/%d"
            ]
            for fmt in formats:
                try:
                    return datetime.strptime(val_str, fmt).date()
                except ValueError:
                    continue
            return None

        def clean_excel_val(doctype, fieldname, val):
            if val is None or val == "": return None
            if fieldname in DATE_FIELDS:
                return parse_excel_date(val)
                
            meta = frappe.get_meta(doctype)
            field = meta.get_field(fieldname)
            if not field: return val
            
            if field.fieldtype in ["Link", "Select", "Data", "Text", "Small Text", "Long Text"]:
                if isinstance(val, (int, float)):
                    if isinstance(val, float) and val.is_integer():
                        return str(cint(val))
                    return str(val)
                return str(val).strip()
                
            if field.fieldtype in ["Float", "Int", "Currency", "Percent"]:
                return flt(val)
                
            return val

        # 1. HEADER MAPPING
        HEADER_MAPPING = {
            "NOMOR AJU": "nomoraju",
            "KODE DOKUMEN": "kode_dokumen",
            "KODE KANTOR": "kode_kantor",
            "KOTA PERNYATAAN": "kota_pernyataan",
            "TANGGAL PERNYATAAN": "tanggal_pernyataan",
            "NAMA PERNYATAAN": "nama_pernyataan",
            "JABATAN PERNYATAAN": "jabatan_pernyataan",
            "KODE KANTOR BONGKAR": "kode_kantor_bongkar",
            "KODE KANTOR PERIKSA": "kode_kantor_periksa",
            "KODE KANTOR TUJUAN": "kode_kantor_tujuan",
            "KODE KANTOR EKSPOR": "kode_kantor_ekspor",
            "KODE JENIS EKSPOR": "kode_jenis_ekspor",
            "KODE JENIS TPB": "kode_jenis_tpb",
            "KODE JENIS PLB": "kode_jenis_plb",
            "KODE JENIS IMPOR": "kode_jenis_impor",
            "KODE TUJUAN PEMASUKAN": "kode_tujuan_pemasukan",
            "KODE TUJUAN PENGIRIMAN": "kode_tujuan_pengiriman",
            "KODE TUJUAN TPB": "kode_tujuan_tpb",
            "KODE CARA DAGANG": "kode_cara_dagang",
            "KODE CARA BAYAR": "kode_cara_bayar",
            "KODE CARA BAYAR LAINNYA": "kode_cara_bayar_lainnya",
            "KODE GUDANG ASAL": "kode_gudang_asal",
            "KODE GUDANG TUJUAN": "kode_gudang_tujuan",
            "KODE JENIS KIRIM": "kode_jenis_kirim",
            "KODE JENIS PENGIRIMAN": "kode_jenis_pengiriman",
            "KODE KATEGORI EKSPOR": "kode_kategori_ekspor",
            "KODE KATEGORI MASUK FTZ": "kode_kategori_masuk_ftz",
            "KODE KATEGORI KELUAR FTZ": "kode_kategori_keluar_ftz",
            "KODE KATEGORI BARANG FTZ": "kode_kategori_barang_ftz",
            "KODE LOKASI": "kode_lokasi",
            "KODE LOKASI BAYAR": "kode_lokasi_bayar",
            "LOKASI ASAL": "lokasi_asal",
            "LOKASI TUJUAN": "lokasi_tujuan",
            "KODE DAERAH ASAL": "kode_daerah_asal",
            "KODE NEGARA TUJUAN": "kode_negara_tujuan",
            "KODE TUTUP PU": "kode_tutup_pu",
            "NOMOR BC11": "nomor_bc11",
            "NOMOR BC 11": "nomor_bc11",
            "NO BC11": "nomor_bc11",
            "TANGGAL BC11": "tanggal_bc11",
            "TANGGAL BC 11": "tanggal_bc11",
            "TGL BC11": "tanggal_bc11",
            "NOMOR POS": "nomor_pos",
            "NO POS": "nomor_pos",
            "NOMOR SUB POS": "nomor_sub_pos",
            "NO SUB POS": "nomor_sub_pos",
            "NOMOR SUBPOS": "nomor_sub_pos",
            "KODE PELABUHAN BONGKAR": "kode_pelabuhan_bongkar",
            "KODE PELABUHAN MUAT": "kode_pelabuhan_muat",
            "KODE PELABUHAN MUAT AKHIR": "kode_pelabuhan_muat_akhir",
            "KODE PELABUHAN TRANSIT": "kode_pelabuhan_transit",
            "KODE PELABUHAN TUJUAN": "kode_pelabuhan_tujuan",
            "KODE PELABUHAN EKSPOR": "kode_pelabuhan_ekspor",
            "KODE TPS": "kode_tps",
            "TANGGAL BERANGKAT": "tanggal_berangkat",
            "TANGGAL EKSPOR": "tanggal_ekspor",
            "TANGGAL MASUK": "tanggal_masuk",
            "TANGGAL MUAT": "tanggal_muat",
            "TANGGAL TIBA": "tanggal_tiba",
            "TANGGAL PERIKSA": "tanggal_periksa",
            "TEMPAT STUFFING": "tempat_stuffing",
            "TANGGAL STUFFING": "tanggal_stuffing",
            "KODE TANDA PENGAMAN": "kode_tanda_pengaman",
            "JUMLAH TANDA PENGAMAN": "jumlah_tanda_pengaman",
            "FLAG CURAH": "flag_curah",
            "FLAG SDA": "flag_sda",
            "FLAG VD": "flag_vd",
            "FLAG MIGAS": "flag_migas",
            "KODE ASURANSI": "kode_asuransi",
            "ASURANSI": "asuransi",
            "NILAI BARANG": "nilai_barang",
            "NILAI INCOTERM": "nilai_incoterm",
            "NILAI MAKLON": "nilai_maklon",
            "FREIGHT": "freight",
            "FOB": "fob",
            "BIAYA TAMBAHAN": "biaya_tambahan",
            "BIAYA PENGURANG": "biaya_pengurang",
            "VD": "vd",
            "CIF": "cif",
            "HARGA_PENYERAHAN": "harga_penyerahan",
            "NDPBM": "ndpbm",
            "TOTAL DANA SAWIT": "total_dana_sawit",
            "DASAR PENGENAAN PAJAK": "dasar_pengenaan_pajak",
            "NILAI JASA": "nilai_jasa",
            "UANG MUKA": "uang_muka",
            "BRUTO": "bruto",
            "NETTO": "netto",
            "VOLUME": "volume",
            "KODE VALUTA": "kode_valuta",
            "KODE INCOTERM": "kode_incoterm",
            "KODE JASA KENA PAJAK": "kode_jasa_kena_pajak",
            "NOMOR BUKTI BAYAR": "nomor_bukti_bayar",
            "TANGGAL BUKTI BAYAR": "tanggal_bukti_bayar",
            "KODE JENIS NILAI": "kode_jenis_nilai",
            "KODE KANTOR MUAT": "kode_kantor_muat",
            "NOMOR DAFTAR": "nomor_daftar",
            "TANGGAL DAFTAR": "tanggal_daftar",
            "KODE ASAL BARANG FTZ": "kode_barang_asal_ftz",
            "KODE TUJUAN PENGELUARAN": "kode_tujuan_pengeluaran",
            "PPN PAJAK": "ppn_pajak",
            "PPNBM PAJAK": "ppnbm_pajak",
            "TARIF PPN PAJAK": "tarif_ppn_pajak",
            "TARIF PPNBM PAJAK": "tarif_ppnbm_pajak",
            "BARANG TIDAK BERWUJUD": "barang_tidak_berwujud",
            "FLAG KONSOL": "flag_konsol",
            "FLAG PROPORSIONAL NETTO": "flag_proporsional_netto",
            "FLAG AP BK": "flag_ap_bk",
            "KODE JENIS PENGELUARAN": "kode_jenis_pengeluaran",
            "KODE JENIS PROSEDUR": "kode_jenis_prosedur",
            "KODE JENIS PENGANGKUTAN": "kode_jenis_pengangkutan",
            "BARANG KIRIMAN": "barang_kiriman"
        }

        # Fetch HEADER data
        header_rows = get_sheet_data("HEADER", expected_columns=HEADER_MAPPING)
        if not header_rows:
            frappe.throw("No data in HEADER sheet")
        
        header_row = header_rows[0]
        nomor_aju = header_row.get("NOMOR AJU")
        if not nomor_aju:
            frappe.throw("NOMOR AJU is missing in HEADER")
        nomor_aju = str(nomor_aju).strip()

        # Get/Create Header Doc
        if frappe.db.exists("HEADER V21", {"nomoraju": nomor_aju}):
            doc = frappe.get_doc("HEADER V21", {"nomoraju": nomor_aju})
        else:
            doc = frappe.new_doc("HEADER V21")
            doc.nomoraju = nomor_aju
            doc.name = nomor_aju
            doc.flags.name_set = True

        # Map Header
        for excel_col, doc_field in HEADER_MAPPING.items():
            if excel_col in header_row:
                val = header_row.get(excel_col)
                doc.set(doc_field, clean_excel_val("HEADER V21", doc_field, val))

        # Helper for Child Tables
        def create_child(doctype, parent_field, sheet_name, mapping, optional=False):
            rows = get_sheet_data(sheet_name, optional, expected_columns=mapping)
            if not rows and optional: return
            
            child_list = []
            for row in rows:
                row_nomor_aju = str(row.get("NOMOR AJU") or "").strip()
                if row_nomor_aju != nomor_aju: continue
                child_item = {}
                for excel_col, doc_field in mapping.items():
                    if excel_col in row:
                        val = row.get(excel_col)
                        child_item[doc_field] = clean_excel_val(doctype, doc_field, val)
                child_list.append(child_item)
            
            doc.set(parent_field, child_list)
            audit_report["stats"][doctype] = len(child_list)

        # CHILD TABLES IMPORT
        create_child("ENTITAS", "entitas", "ENTITAS", {
            "NOMOR AJU": "nomoraju", "SERI": "seri", "KODE ENTITAS": "kode_entitas",
            "NOMOR IDENTITAS": "nomor_identitas", "NAMA ENTITAS": "nama_entitas",
            "ALAMAT ENTITAS": "alamat_entitas", "NIB ENTITAS": "nib_entitas",
            "KODE JENIS API": "kode_jenis_api", "KODE STATUS": "kode_status",
            "KODE NEGARA": "kode_negara", "NOMOR IJIN ENTITAS": "nomor_ijin_entitas",
            "TANGGAL IJIN ENTITAS": "tanggal_ijin_entitas", "KODE JENIS IDENTITAS": "kode_jenis_identitas",
            "NIPER ENTITAS": "niper_entitas", "KODE AFILIASI": "kode_afiliasi",
            "KODE KATEGORI KONSOLIDATOR": "kode_kategori_konsolidator"
        })

        create_child("KEMASAN", "kemasan", "KEMASAN", {
            "NOMOR AJU": "nomoraju", "SERI": "seri", "KODE KEMASAN": "kode_kemasan",
            "JUMLAH KEMASAN": "jumlah_kemasan", "MERK KEMASAN": "merek_kemasan",
            "MEREK": "merek_kemasan", "NOMOR SEGEL": "no_segel_kemasan"
        })

        create_child("DOKUMEN", "dokumen", "DOKUMEN", {
            "NOMOR AJU": "nomoraju", "SERI DOKUMEN": "seri", "SERI": "seri",
            "KODE DOKUMEN": "kode_dokumen", "NOMOR DOKUMEN": "nomor_dokumen",
            "TANGGAL DOKUMEN": "tanggal_dokumen", "KODE FASILITAS": "kode_fasilitas",
            "KODE IJIN": "kode_ijin"
        })

        create_child("PENGANGKUT", "pengangkut", "PENGANGKUT", {
            "NOMOR AJU": "nomoraju", "SERI PENGANGKUT": "seri_pengangkut", "SERI": "seri_pengangkut",
            "KODE CARA ANGKUT": "kode_cara_angkut", "NAMA PENGANGKUT": "nama_pengangkut",
            "NOMOR PENGANGKUT": "nomor_pengangkut", "KODE BENDERA": "kode_bendera",
            "CALL SIGN": "call_sign", "FLAG ANGKUT PLB": "flag_angkut_plb",
            "CARA PENGANGKUTAN LAINNYA": "cara_pengangkutan_lainnya"
        })
        
        create_child("KONTAINER", "kontainer", "KONTAINER", {
            "NOMOR AJU": "nomoraju", "SERI KONTAINER": "seri", "SERI": "seri",
            "NOMOR KONTAINER": "nomor_kontainer", "NOMOR KONTINER": "nomor_kontainer",
            "KODE UKURAN KONTAINER": "kode_ukuran_kontainer", "KODE JENIS KONTAINER": "kode_jenis_kontainer",
            "KODE TIPE KONTAINER": "kode_tipe_kontainer", "NOMOR SEGEL": "nomor_segel_kontainer"
        })

        create_child("komponen_biaya", "komponen_biaya", "KOMPONENBIAYA", {
            "NOMOR AJU": "nomoraju", "JENIS NILAI": "jenisnilai", "HARGA INVOICE": "hargainvoice",
            "PEMBAYARAN TIDAK LANGSUNG": "pembayarantidaklangsung", "DISKON": "diskon",
            "KOMISI PENJUALAN": "komisipenjualan", "BIAYA PENGEMASAN": "biayapengemasan",
            "BIAYA PENGEPAKAN": "biayapengepakan", "ASSIST": "assist", "ROYALTI": "royalti",
            "PROCEEDS": "proceeds", "BIAYA TRANSPORTASI": "biayatransportasi",
            "BIAYA PEMUATAN": "biayapemuatan", "ASURANSI": "asuransi", "GARANSI": "garansi",
            "BIAYA KEPENTINGAN SENDIRI": "biayakepentingansendiri", "BIAYA PASCA IMPOR": "biayapascaimpor",
            "BIAYA PAJAK INTERNAL": "biayapajakinternal", "BUNGA": "bunga", "DEVIDEN": "deviden"
        }, optional=True)

        create_child("pungutan", "pungutan", "PUNGUTAN", {
            "NOMOR AJU": "nomoraju", "KODE FASILITAS TARIF": "kode_fasilitas_tarif",
            "KODE JENIS PUNGUTAN": "kode_jenis_pungutan", "NILAI PUNGUTAN": "nilai_pungutan",
            "NPWP BILLING": "npwp_billing"
        })

        create_child("jaminan", "jaminan", "JAMINAN", {
            "NOMOR AJU": "nomoraju", "KODE JAMINAN": "kode_jaminan", "NOMOR BPJ": "nomor_bpj",
            "TANGGAL BPJ": "tanggal_bpj", "NILAI JAMINAN": "nilai_jaminan",
            "TANGGAL JATUH TEMPO": "tanggal_jatuh_tempo", "PENJAMIN": "penjamin",
            "NOMOR JAMINAN": "nomor_jaminan", "TANGGAL JAMINAN": "tanggal_jaminan",
            "KODE KANTOR": "kode_kantor" 
        })

        create_child("bank_devisa", "bank_devisa", "BANKDEVISA", {
            "NOMOR AJU": "nomoraju", "KODE": "kode", "NAMA": "nama", "SERI": "seri"
        })

        respon_rows = get_sheet_data("RESPON", optional=True)
        if respon_rows:
            import json
            doc.respon_json = json.dumps(respon_rows, default=str)

        # --- BARANG PROCESSING ---
        barang_rows = get_sheet_data("BARANG")
        bt_rows = get_sheet_data("BARANGTARIF")
        bd_rows = get_sheet_data("BARANGDOKUMEN")
        be_rows = get_sheet_data("BARANGENTITAS")
        bvd_rows = get_sheet_data("BARANGVD")
        bspe_rows = get_sheet_data("BARANGSPEKKHUSUS")
        
        # BAHAN BAKU Sheets
        bb_rows = get_sheet_data("BAHANBAKU")
        bbt_rows = get_sheet_data("BAHANBAKUTARIF")
        bbd_rows = get_sheet_data("BAHANBAKUDOKUMEN")

        wb.close()

        # Save Header V21 first
        save_doc(doc)

        # Pre-group child rows for O(1) lookups
        bt_by_seri = {}
        for r in bt_rows:
            bt_by_seri.setdefault(cint(r.get("SERI BARANG")), []).append(r)

        bd_by_seri = {}
        for r in bd_rows:
            bd_by_seri.setdefault(cint(r.get("SERI BARANG")), []).append(r)

        be_by_seri = {}
        for r in be_rows:
            be_by_seri.setdefault(cint(r.get("SERI BARANG")), []).append(r)

        bspe_by_seri = {}
        for r in bspe_rows:
            bspe_by_seri.setdefault(cint(r.get("SERI BARANG")), []).append(r)

        bvd_by_seri = {}
        for r in bvd_rows:
            bvd_by_seri.setdefault(cint(r.get("SERI BARANG")), []).append(r)

        bb_by_seri = {}
        for r in bb_rows:
            bb_by_seri.setdefault(cint(r.get("SERI BARANG")), []).append(r)

        bbt_by_key = {}
        for r in bbt_rows:
            key = (cint(r.get("SERI BARANG")), cint(r.get("SERI BAHAN BAKU")))
            bbt_by_key.setdefault(key, []).append(r)

        bbd_by_key = {}
        for r in bbd_rows:
            key = (cint(r.get("SERI BARANG")), cint(r.get("SERI BAHAN BAKU")))
            bbd_by_key.setdefault(key, []).append(r)

        # OPTIMIZATION: Bulk delete existing records and perform fast bulk insert
        
        # Get standard database fields for bulk insert dynamically
        from frappe.model import no_value_fields
        
        # Build field types map to prevent "cannot be null" errors for numeric/check fields
        def get_field_type_map(dt):
            m = frappe.get_meta(dt)
            return {df.fieldname: df.fieldtype for df in m.fields}

        field_types = {
            "BARANG V1": get_field_type_map("BARANG V1"),
            "BARANG TARIF": get_field_type_map("BARANG TARIF"),
            "BARANG DOKUMEN": get_field_type_map("BARANG DOKUMEN"),
            "BARANG ENTITAS": get_field_type_map("BARANG ENTITAS"),
            "BARANG SPEK KHUSUS": get_field_type_map("BARANG SPEK KHUSUS"),
            "BARANG VD": get_field_type_map("BARANG VD"),
            "BAHAN BAKU": get_field_type_map("BAHAN BAKU"),
            "BAHAN BAKU TARIF": get_field_type_map("BAHAN BAKU TARIF"),
            "BAHAN BAKU DOKUMEN": get_field_type_map("BAHAN BAKU DOKUMEN"),
        }

        def get_clean_value(dt, fieldname, val):
            if val is not None:
                return val
            if fieldname in ["docstatus", "idx"]:
                return 0
            t = field_types[dt].get(fieldname)
            if t in ["Int", "Check"]:
                return 0
            elif t in ["Float", "Currency", "Percent"]:
                return 0.0
            return None

        def get_db_fields(dt, is_child=False):
            meta_dt = frappe.get_meta(dt)
            fields_dt = [df.fieldname for df in meta_dt.fields if df.fieldtype not in no_value_fields]
            std_fields = ["name", "owner", "creation", "modified", "modified_by", "docstatus", "idx"]
            if is_child:
                std_fields += ["parent", "parentfield", "parenttype"]
            return std_fields + fields_dt

        barang_fields = get_db_fields("BARANG V1")
        b_tarif_fields = get_db_fields("BARANG TARIF", is_child=True)
        b_dokumen_fields = get_db_fields("BARANG DOKUMEN", is_child=True)
        b_entitas_fields = get_db_fields("BARANG ENTITAS", is_child=True)
        b_spek_fields = get_db_fields("BARANG SPEK KHUSUS", is_child=True)
        b_vd_fields = get_db_fields("BARANG VD", is_child=True)

        bahan_fields = get_db_fields("BAHAN BAKU")
        bb_tarif_fields = get_db_fields("BAHAN BAKU TARIF", is_child=True)
        bb_dokumen_fields = get_db_fields("BAHAN BAKU DOKUMEN", is_child=True)

        # 1. Bulk Delete existing records (Clean Overwrite)
        existing_barang = frappe.get_all("BARANG V1", filters={"nomoraju": doc.name}, fields=["name"])
        barang_names = [b.name for b in existing_barang]

        if barang_names:
            existing_bahan = frappe.get_all("BAHAN BAKU", filters={"parent_barang": ["in", barang_names]}, fields=["name"])
            bahan_names = [b.name for b in existing_bahan]

            frappe.db.delete("BARANG TARIF", {"parent": ["in", barang_names]})
            frappe.db.delete("BARANG DOKUMEN", {"parent": ["in", barang_names]})
            frappe.db.delete("BARANG ENTITAS", {"parent": ["in", barang_names]})
            frappe.db.delete("BARANG SPEK KHUSUS", {"parent": ["in", barang_names]})
            frappe.db.delete("BARANG VD", {"parent": ["in", barang_names]})
            
            if bahan_names:
                frappe.db.delete("BAHAN BAKU TARIF", {"parent": ["in", bahan_names]})
                frappe.db.delete("BAHAN BAKU DOKUMEN", {"parent": ["in", bahan_names]})
                frappe.db.delete("BAHAN BAKU", {"name": ["in", bahan_names]})
                
            frappe.db.delete("BARANG V1", {"name": ["in", barang_names]})

        # Count target Barang V1 records to be inserted
        barang_rows_to_insert = []
        for b_row in barang_rows:
            row_nomor_aju = str(b_row.get("NOMOR AJU") or "").strip()
            if row_nomor_aju and row_nomor_aju != nomor_aju: continue
            
            seri_barang = cint(b_row.get("SERI BARANG"))
            if not seri_barang: continue
            barang_rows_to_insert.append((seri_barang, b_row))

        N_barang = len(barang_rows_to_insert)

        # 2. Generate series names for BARANG V1 in-memory
        barang_names_generated = []
        if N_barang > 0:
            prefix = "BRG-"
            frappe.db.sql(
                "insert into `tabSeries` (name, current) values (%s, %s) "
                "on duplicate key update current = current + %s",
                (prefix, N_barang, N_barang)
            )
            current = frappe.db.sql("select current from `tabSeries` where name = %s", prefix)[0][0]
            start = current - N_barang + 1
            barang_names_generated = [f"{prefix}{str(i).zfill(5)}" for i in range(start, current + 1)]

        # 3. Construct records in memory
        records_barang = []
        records_b_tarif = []
        records_b_dokumen = []
        records_b_entitas = []
        records_b_spek = []
        records_b_vd = []

        records_bahan = []
        records_bb_tarif = []
        records_bb_dokumen = []

        from frappe.model.naming import make_autoname
        now = datetime.now()

        BARANG_MAPPING = {
            "NOMOR AJU": "nomoraju", "SERI BARANG": "seri_barang", "HS": "hs",
            "KODE BARANG": "kode_barang", "URAIAN": "uraian", "MEREK": "merek",
            "MERK": "merek", "TIPE": "tipe", "UKURAN": "ukuran", "SPESIFIKASI LAIN": "spesifikasi_lain",
            "KODE SATUAN": "kode_satuan", "JUMLAH SATUAN": "jumlah_satuan",
            "JUMLAH BAHAN BAKU": "jumlah_bahan_baku", "KODE KEMASAN": "kode_kemasan",
            "JUMLAH KEMASAN": "jumlah_kemasan", "NETTO": "netto", "BRUTO": "bruto",
            "VOLUME": "volume", "CIF": "cif", "CIF RUPIAH": "cif_rupiah", "NDPBM": "ndpbm",
            "FOB": "fob", "ASURANSI": "asuransi", "FREIGHT": "freight", "NILAI BARANG": "nilai_barang",
            "NILAI JASA": "nilai_jasa", "NILAI DANA SAWIT": "nilai_dana_sawit",
            "NILAI DEVISA": "nilai_devisa", "PERSENTASE IMPOR": "persentase_impor", "DISKON": "diskon",
            "HARGA PENYERAHAN": "harga_penyerahan", "HARGA PEROLEHAN": "harga_perolehan",
            "HARGA SATUAN": "harga_satuan", "HARGA EKSPOR": "harga_ekspor", "HARGA PATOKAN": "harga_patokan",
            "NILAI TAMBAH": "nilai_tambah", "PERNYATAAN LARTAS": "pernyataan_lartas",
            "TAHUN PEMBUATAN": "tahun_pembuatan", "KODE JENIS EKSPOR": "kode_jenis_ekspor",
            "KODE KATEGORI BARANG": "kode_kategori_barang", "KODE KONDISI BARANG": "kode_kondisi_barang",
            "SERI IZIN": "seri_izin", "KODE ASAL BARANG": "kode_asal_barang",
            "KODE NEGARA ASAL": "kode_negara_asal", "KODE DAERAH ASAL": "kode_daerah_asal",
            "STATEMENT PERBEDAAN HARGA": "statement_perbedaan_harga", "SALDO AWAL": "saldo_awal",
            "KAPASITAS SILINDER": "kapasitas_silinder", "JATUH TEMPO ROYALTI": "jatuh_tempo_royalti",
            "FLAG TIS": "flag_tis", "KODE KANTOR ASAL": "kode_kantor_asal", "ISI PER KEMASAN": "isi_per_kemasan",
            "FLAG 4 TAHUN": "flag_4_tahun", "KODE DOKUMEN ASAL": "kode_dokumen_asal", "KODE BKC": "kode_bkc",
            "KODE SUB KOMODITI BKC": "kode_sub_komoditi_bkc", "NOMOR AJU ASAL": "nomor_aju_asal",
            "KODE GUNA BARANG": "kode_guna_barang", "SERI BARANG ASAL": "seri_barang_asal",
            "SALDO AKHIR": "saldo_akhir", "KODE PERHITUNGAN": "kode_perhitungan",
            "KODE JENIS NILAI": "kode_jenis_nilai", "KODE KOMODITI BKC": "kode_komoditi_bkc",
            "NOMOR DAFTAR ASAL": "nomor_daftar_asal", "JUMLAH REALISASI": "jumlah_realisasi",
            "TANGGAL DAFTAR ASAL": "tanggal_daftar_asal", "METODE PENENTUAN NILAI": "metode_penentuan_nilai",
            "HJE CUKAI": "hje_cukai", "TARIF CUKAI": "tarif_cukai", "JUMLAH PITA CUKAI": "jumlah_pita_cukai",
            "JUMLAH DILEKATKAN": "jumlah_dilekatkan"
        }

        # Build Barang and Bahan Baku records loop
        for idx_b, (seri_barang, b_row) in enumerate(barang_rows_to_insert):
            b_name = barang_names_generated[idx_b]
            
            # Map parent fields from Excel row
            b_dict = {
                "name": b_name,
                "owner": frappe.session.user or "Administrator",
                "creation": now,
                "modified": now,
                "modified_by": frappe.session.user or "Administrator",
                "docstatus": 0,
                "idx": idx_b + 1,
                "nomoraju": doc.name
            }
            
            for excel_col, doc_field in BARANG_MAPPING.items():
                if excel_col in b_row:
                    b_dict[doc_field] = clean_excel_val("BARANG V1", doc_field, b_row.get(excel_col))
            
            b_dict["nomoraju"] = doc.name
            records_barang.append(b_dict)

            # Child Tables for BARANG V1
            
            # Tarif
            for idx_t, r in enumerate(bt_by_seri.get(seri_barang, [])):
                t_dict = {
                    "name": make_autoname('hash'),
                    "owner": frappe.session.user or "Administrator",
                    "creation": now,
                    "modified": now,
                    "modified_by": frappe.session.user or "Administrator",
                    "docstatus": 0,
                    "idx": idx_t + 1,
                    "parent": b_name,
                    "parentfield": "barang_tarif",
                    "parenttype": "BARANG V1",
                    "seri_barang": seri_barang,
                    "kode_pungutan": clean_excel_val("BARANG TARIF", "kode_pungutan", r.get("KODE PUNGUTAN")),
                    "kode_tarif": clean_excel_val("BARANG TARIF", "kode_tarif", r.get("KODE TARIF")),
                    "tarif": clean_excel_val("BARANG TARIF", "tarif", r.get("TARIF")),
                    "kode_fasilitas": clean_excel_val("BARANG TARIF", "kode_fasilitas", r.get("KODE FASILITAS")),
                    "tarif_fasilitas": clean_excel_val("BARANG TARIF", "tarif_fasilitas", r.get("TARIF FASILITAS")),
                    "nilai_bayar": clean_excel_val("BARANG TARIF", "nilai_bayar", r.get("NILAI BAYAR")),
                    "nilai_fasilitas": clean_excel_val("BARANG TARIF", "nilai_fasilitas", r.get("NILAI FASILITAS")),
                    "nilai_sudah_dilunasi": clean_excel_val("BARANG TARIF", "nilai_sudah_dilunasi", r.get("NILAI SUDAH DILUNASI")),
                    "kode_komoditi_cukai": clean_excel_val("BARANG TARIF", "kode_komoditi_cukai", r.get("KODE KOMODITI CUKAI")),
                    "kode_sub_komoditi_cukai": clean_excel_val("BARANG TARIF", "kode_sub_komoditi_cukai", r.get("KODE SUB KOMODITI CUKAI")),
                    "jumlah_satuan": clean_excel_val("BARANG TARIF", "jumlah_satuan", r.get("JUMLAH SATUAN")),
                    "kode_satuan": clean_excel_val("BARANG TARIF", "kode_satuan", r.get("KODE SATUAN"))
                }
                records_b_tarif.append(t_dict)
                
            # Dokumen
            for idx_d, r in enumerate(bd_by_seri.get(seri_barang, [])):
                d_dict = {
                    "name": make_autoname('hash'),
                    "owner": frappe.session.user or "Administrator",
                    "creation": now,
                    "modified": now,
                    "modified_by": frappe.session.user or "Administrator",
                    "docstatus": 0,
                    "idx": idx_d + 1,
                    "parent": b_name,
                    "parentfield": "barang_dokumen",
                    "parenttype": "BARANG V1",
                    "seri_dokumen": clean_excel_val("BARANG DOKUMEN", "seri_dokumen", r.get("SERI DOKUMEN")),
                    "seri_izin": clean_excel_val("BARANG DOKUMEN", "seri_izin", r.get("SERI IZIN"))
                }
                records_b_dokumen.append(d_dict)
                
            # Pemilik (Entitas)
            for idx_e, r in enumerate(be_by_seri.get(seri_barang, [])):
                e_dict = {
                    "name": make_autoname('hash'),
                    "owner": frappe.session.user or "Administrator",
                    "creation": now,
                    "modified": now,
                    "modified_by": frappe.session.user or "Administrator",
                    "docstatus": 0,
                    "idx": idx_e + 1,
                    "parent": b_name,
                    "parentfield": "barang_pemilik",
                    "parenttype": "BARANG V1",
                    "seri_entitas": clean_excel_val("BARANG ENTITAS", "seri_entitas", r.get("SERI ENTITAS"))
                }
                records_b_entitas.append(e_dict)
                
            # Spek Khusus
            for idx_sp, r in enumerate(bspe_by_seri.get(seri_barang, [])):
                sp_dict = {
                    "name": make_autoname('hash'),
                    "owner": frappe.session.user or "Administrator",
                    "creation": now,
                    "modified": now,
                    "modified_by": frappe.session.user or "Administrator",
                    "docstatus": 0,
                    "idx": idx_sp + 1,
                    "parent": b_name,
                    "parentfield": "barang_spek_khusus",
                    "parenttype": "BARANG V1",
                    "kode": clean_excel_val("BARANG SPEK KHUSUS", "kode", r.get("KODE")),
                    "uraian": clean_excel_val("BARANG SPEK KHUSUS", "uraian", r.get("URAIAN"))
                }
                records_b_spek.append(sp_dict)
                
            # VD
            for idx_vd, r in enumerate(bvd_by_seri.get(seri_barang, [])):
                vd_dict = {
                    "name": make_autoname('hash'),
                    "owner": frappe.session.user or "Administrator",
                    "creation": now,
                    "modified": now,
                    "modified_by": frappe.session.user or "Administrator",
                    "docstatus": 0,
                    "idx": idx_vd + 1,
                    "parent": b_name,
                    "parentfield": "barang_vd",
                    "parenttype": "BARANG V1",
                    "kode_jenis_vd": clean_excel_val("BARANG VD", "kode_jenis_vd", r.get("KODE VD")),
                    "nilai_barang": clean_excel_val("BARANG VD", "nilai_barang", r.get("NILAI BARANG")),
                    "biaya_tambahan": clean_excel_val("BARANG VD", "biaya_tambahan", r.get("BIAYA TAMBAHAN")),
                    "biaya_pengurang": clean_excel_val("BARANG VD", "biaya_pengurang", r.get("BIAYA PENGURANG")),
                    "jatuh_tempo": clean_excel_val("BARANG VD", "jatuh_tempo", r.get("JATUH TEMPO"))
                }
                records_b_vd.append(vd_dict)

            # BAHAN BAKU under this barang
            for idx_bb, bb_row in enumerate(bb_by_seri.get(seri_barang, [])):
                seri_bahan_baku = cint(bb_row.get("SERI BAHAN BAKU"))
                if not seri_bahan_baku: continue
                
                bb_name = make_autoname('hash')
                bb_dict = {
                    "name": bb_name,
                    "owner": frappe.session.user or "Administrator",
                    "creation": now,
                    "modified": now,
                    "modified_by": frappe.session.user or "Administrator",
                    "docstatus": 0,
                    "idx": idx_bb + 1,
                    "nomoraju": nomor_aju,
                    "seri_barang": seri_barang,
                    "seri_bahan_baku": seri_bahan_baku,
                    "parent_barang": b_name,
                    "hs": clean_excel_val("BAHAN BAKU", "hs", bb_row.get("HS")),
                    "kode_barang": clean_excel_val("BAHAN BAKU", "kode_barang", bb_row.get("KODE BARANG")),
                    "uraian": clean_excel_val("BAHAN BAKU", "uraian", bb_row.get("URAIAN")),
                    "merek": clean_excel_val("BAHAN BAKU", "merek", bb_row.get("MEREK")),
                    "tipe": clean_excel_val("BAHAN BAKU", "tipe", bb_row.get("TIPE")),
                    "ukuran": clean_excel_val("BAHAN BAKU", "ukuran", bb_row.get("UKURAN")),
                    "spesifikasi_lain": clean_excel_val("BAHAN BAKU", "spesifikasi_lain", bb_row.get("SPESIFIKASI LAIN")),
                    "kode_satuan": clean_excel_val("BAHAN BAKU", "kode_satuan", bb_row.get("KODE SATUAN")),
                    "jumlah_satuan": clean_excel_val("BAHAN BAKU", "jumlah_satuan", bb_row.get("JUMLAH SATUAN")),
                    "kode_asal_bahan_baku": clean_excel_val("BAHAN BAKU", "kode_asal_bahan_baku", bb_row.get("KODE ASAL BAHAN BAKU")),
                    "cif": clean_excel_val("BAHAN BAKU", "cif", bb_row.get("CIF")),
                    "cif_rupiah": clean_excel_val("BAHAN BAKU", "cif_rupiah", bb_row.get("CIF RUPIAH")),
                    "harga_penyerahan": clean_excel_val("BAHAN BAKU", "harga_penyerahan", bb_row.get("HARGA PENYERAHAN")),
                    "harga_perolehan": clean_excel_val("BAHAN BAKU", "harga_perolehan", bb_row.get("HARGA PEROLEHAN")),
                    "ndpbm": clean_excel_val("BAHAN BAKU", "ndpbm", bb_row.get("NDPBM")),
                    "netto": clean_excel_val("BAHAN BAKU", "netto", bb_row.get("NETTO")),
                    "bruto": clean_excel_val("BAHAN BAKU", "bruto", bb_row.get("BRUTO")),
                    "volume": clean_excel_val("BAHAN BAKU", "volume", bb_row.get("VOLUME")),
                    "kode_bkc": clean_excel_val("BAHAN BAKU", "kode_bkc", bb_row.get("KODE BKC")),
                    "kode_komoditi_bkc": clean_excel_val("BAHAN BAKU", "kode_komoditi_bkc", bb_row.get("KODE KOMODITI BKC")),
                    "kode_sub_komoditi_bkc": clean_excel_val("BAHAN BAKU", "kode_sub_komoditi_bkc", bb_row.get("KODE SUB KOMODITI BKC")),
                    "flag_tis": clean_excel_val("BAHAN BAKU", "flag_tis", bb_row.get("FLAG TIS")),
                    "isi_per_kemasan": clean_excel_val("BAHAN BAKU", "isi_per_kemasan", bb_row.get("ISI PER KEMASAN")),
                    "jumlah_dilekatkan": clean_excel_val("BAHAN BAKU", "jumlah_dilekatkan", bb_row.get("JUMLAH DILEKATKAN")),
                    "jumlah_pita_cukai": clean_excel_val("BAHAN BAKU", "jumlah_pita_cukai", bb_row.get("JUMLAH PITA CUKAI")),
                    "hje_cukai": clean_excel_val("BAHAN BAKU", "hje_cukai", bb_row.get("HJE CUKAI")),
                    "tarif_cukai": clean_excel_val("BAHAN BAKU", "tarif_cukai", bb_row.get("TARIF CUKAI")),
                    "nomor_aju_asal": clean_excel_val("BAHAN BAKU", "nomor_aju_asal", bb_row.get("NOMOR AJU ASAL")),
                    "nomor_daftar_asal": clean_excel_val("BAHAN BAKU", "nomor_daftar_asal", bb_row.get("NOMOR DAFTAR ASAL")),
                    "tanggal_daftar_asal": clean_excel_val("BAHAN BAKU", "tanggal_daftar_asal", bb_row.get("TANGGAL DAFTAR ASAL")),
                    "kode_dokumen_asal": clean_excel_val("BAHAN BAKU", "kode_dokumen_asal", bb_row.get("KODE DOKUMEN ASAL")),
                    "kode_kantor_asal": clean_excel_val("BAHAN BAKU", "kode_kantor_asal", bb_row.get("KODE KANTOR ASAL"))
                }
                records_bahan.append(bb_dict)

                # Children of Bahan Baku
                key = (seri_barang, seri_bahan_baku)
                
                # BB Tarif
                for idx_bbt, r_bbt in enumerate(bbt_by_key.get(key, [])):
                    bbt_dict = {
                        "name": make_autoname('hash'),
                        "owner": frappe.session.user or "Administrator",
                        "creation": now,
                        "modified": now,
                        "modified_by": frappe.session.user or "Administrator",
                        "docstatus": 0,
                        "idx": idx_bbt + 1,
                        "parent": bb_name,
                        "parentfield": "bahan_tarif",
                        "parenttype": "BAHAN BAKU",
                        "kode_pungutan": clean_excel_val("BAHAN BAKU TARIF", "kode_pungutan", r_bbt.get("KODE PUNGUTAN")),
                        "kode_tarif": clean_excel_val("BAHAN BAKU TARIF", "kode_tarif", r_bbt.get("KODE TARIF")),
                        "tarif": clean_excel_val("BAHAN BAKU TARIF", "tarif", r_bbt.get("TARIF")),
                        "kode_fasilitas": clean_excel_val("BAHAN BAKU TARIF", "kode_fasilitas", r_bbt.get("KODE FASILITAS")),
                        "tarif_fasilitas": clean_excel_val("BAHAN BAKU TARIF", "tarif_fasilitas", r_bbt.get("TARIF FASILITAS")),
                        "nilai_bayar": clean_excel_val("BAHAN BAKU TARIF", "nilai_bayar", r_bbt.get("NILAI BAYAR")),
                        "nilai_fasilitas": clean_excel_val("BAHAN BAKU TARIF", "nilai_fasilitas", r_bbt.get("NILAI FASILITAS")),
                        "kode_asal_bahan_baku": clean_excel_val("BAHAN BAKU TARIF", "kode_asal_bahan_baku", r_bbt.get("KODE ASAL BAHAN BAKU")),
                        "jumlah_satuan": clean_excel_val("BAHAN BAKU TARIF", "jumlah_satuan", r_bbt.get("JUMLAH SATUAN")),
                        "kode_satuan": clean_excel_val("BAHAN BAKU TARIF", "kode_satuan", r_bbt.get("KODE SATUAN"))
                    }
                    records_bb_tarif.append(bbt_dict)

                # BB Dokumen
                for idx_bbd, r_bbd in enumerate(bbd_by_key.get(key, [])):
                    bbd_dict = {
                        "name": make_autoname('hash'),
                        "owner": frappe.session.user or "Administrator",
                        "creation": now,
                        "modified": now,
                        "modified_by": frappe.session.user or "Administrator",
                        "docstatus": 0,
                        "idx": idx_bbd + 1,
                        "parent": bb_name,
                        "parentfield": "bahan_baku_dokumen",
                        "parenttype": "BAHAN BAKU",
                        "seri_dokumen": clean_excel_val("BAHAN BAKU DOKUMEN", "seri_dokumen", r_bbd.get("SERI DOKUMEN")),
                        "seri_izin": clean_excel_val("BAHAN BAKU DOKUMEN", "seri_izin", r_bbd.get("SERI IZIN")),
                        "kode_asal_bahan_baku": clean_excel_val("BAHAN BAKU DOKUMEN", "kode_asal_bahan_baku", r_bbd.get("KODE ASAL BAHAN BAKU"))
                    }
                    records_bb_dokumen.append(bbd_dict)

        # 4. Perform bulk insert
        def get_values_list(dt, doc_dict, fields_list):
            return [get_clean_value(dt, f, doc_dict.get(f)) for f in fields_list]

        def execute_bulk(dt, records, fields_list):
            if not records: return
            values = [get_values_list(dt, r, fields_list) for r in records]
            frappe.db.bulk_insert(dt, fields_list, values)

        execute_bulk("BARANG V1", records_barang, barang_fields)
        execute_bulk("BARANG TARIF", records_b_tarif, b_tarif_fields)
        execute_bulk("BARANG DOKUMEN", records_b_dokumen, b_dokumen_fields)
        execute_bulk("BARANG ENTITAS", records_b_entitas, b_entitas_fields)
        execute_bulk("BARANG SPEK KHUSUS", records_b_spek, b_spek_fields)
        execute_bulk("BARANG VD", records_b_vd, b_vd_fields)

        execute_bulk("BAHAN BAKU", records_bahan, bahan_fields)
        execute_bulk("BAHAN BAKU TARIF", records_bb_tarif, bb_tarif_fields)
        execute_bulk("BAHAN BAKU DOKUMEN", records_bb_dokumen, bb_dokumen_fields)

        audit_report["stats"]["BARANG V1"] = len(records_barang)
        audit_report["stats"]["BARANG TARIF"] = len(records_b_tarif)
        audit_report["stats"]["BARANG DOKUMEN"] = len(records_b_dokumen)
        audit_report["stats"]["BARANG PEMILIK"] = len(records_b_entitas)
        audit_report["stats"]["BARANG SPEK KHUSUS"] = len(records_b_spek)
        audit_report["stats"]["BARANG VD"] = len(records_b_vd)
        
        audit_report["stats"]["BAHAN BAKU"] = len(records_bahan)
        audit_report["stats"]["BAHAN BAKU TARIF"] = len(records_bb_tarif)
        audit_report["stats"]["BAHAN BAKU DOKUMEN"] = len(records_bb_dokumen)
        audit_report["stats"]["saves_performed"] = 1
        audit_report["stats"]["saves_skipped"] = 0

        message = f"<b>Successfully processed {nomor_aju} (V2 Optimized)</b>"
        if audit_report["stats"]:
            message += "<br><br><b>📊 Import Statistics:</b><br>"
            for table, count in audit_report["stats"].items():
                message += f"- {table}: {count} records<br>"
        
        if audit_report["unmapped_columns"]:
            message += "<br><b>⚠️ Unmapped Columns:</b><br>"
            for sheet, cols in audit_report["unmapped_columns"].items():
                message += f"- {sheet}: {', '.join(cols)}<br>"
        
        if audit_report["missing_columns"]:
            message += "<br><b>ℹ️ Missing Columns:</b><br>"
            for sheet, cols in audit_report["missing_columns"].items():
                message += f"- {sheet}: {', '.join(cols)}<br>"
        
        if cint(dry_run):
            frappe.db.rollback()
            return {"status": "success", "message": "[DRY RUN] " + message, "audit": audit_report, "nomor_aju": nomor_aju}
        
        frappe.db.commit()
        return {"status": "success", "message": message, "audit": audit_report, "nomor_aju": nomor_aju}

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Import CEISA Excel V2 Error")
        error_msg = f"<b>Error during import (V2):</b> {str(e)}"
        if audit_report["stats"]:
           error_msg += "<br><br><b>Partial Statistics:</b><br>"
           for k, v in audit_report["stats"].items():
                 error_msg += f"- {k}: {v}<br>"
        return {"status": "error", "message": error_msg, "audit": audit_report}
