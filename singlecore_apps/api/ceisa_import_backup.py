import frappe
import openpyxl
import base64
import io
from frappe.utils import getdate, flt, cint

@frappe.whitelist(allow_guest=True)
def import_ceisa_excel(file_data, dry_run=False):
    # Convert dry_run to bool if passed as string/int from JS
    dry_run = frappe.parse_json(dry_run) if isinstance(dry_run, str) else bool(dry_run)
    
    audit_report = {
        "unmapped_columns": {},
        "missing_columns": {},
        "stats": {}
    }

    try:
        # Check if file_data is a path (e.g., /private/files/...)
        if file_data.startswith("/private/files/") or file_data.startswith("/files/"):
            file_path = frappe.get_site_path(file_data.strip("/"))
            with open(file_path, "rb") as f:
                decoded_file = f.read()
        else:
            # Assume base64
            if "," in file_data:
                file_data = file_data.split(",")[1]
            decoded_file = base64.b64decode(file_data)
        
        # Load workbook
        wb = openpyxl.load_workbook(io.BytesIO(decoded_file), data_only=True)
        
        # Helper to get value from sheet with Audit
        def get_sheet_data(sheet_name, optional=False, expected_columns=None):
            if sheet_name not in wb.sheetnames:
                if not optional:
                    frappe.log_error(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}", "Excel Import Warning")
                return []
            
            ws = wb[sheet_name]
            data = []
            
            # Read unique headers
            headers = []
            for cell in ws[1]:
                if cell.value:
                    # Normalize: Strip, Uppercase, and reduce multiple spaces
                    header_val = str(cell.value).strip().upper()
                    header_val = " ".join(header_val.split())
                    headers.append(header_val)
            
            # Debugging: Log headers found vs expected
            if expected_columns:
                frappe.log_error(f"Sheet: {sheet_name}\nFound: {headers}\nExpected: {list(expected_columns)}", "Import Headers Debug")

            # Audit: Check for unmapped (extra) and missing (required) columns
            if expected_columns:
                # 1. Unmapped: Present in File but Not in Mapping
                unmapped = [h for h in headers if h not in expected_columns]
                if unmapped:
                    audit_report["unmapped_columns"][sheet_name] = unmapped
                
                # 2. Missing: Present in Mapping but Not in File
                missing = [c for c in expected_columns if c not in headers]
                if missing:
                    audit_report["missing_columns"][sheet_name] = missing

            for row in ws.iter_rows(min_row=2, values_only=True):
                # Zip only up to length of row or headers
                row_dict = {}
                for idx, header in enumerate(headers):
                    if idx < len(row):
                         row_dict[header] = row[idx]
                
                # Filter out empty rows
                if any(row_dict.values()):
                    data.append(row_dict)
            return data

        # 1. Create Header
        HEADER_MAPPING = {
            "NOMOR AJU": "nomoraju",
            "KODE DOKUMEN": "kode_dokumen",
            "KODE KANTOR": "kode_kantor",
            "KODE KANTOR BONGKAR": "kode_kantor_bongkar",
            "KODE KANTOR PERIKSA": "kode_kantor_periksa",
            "KODE KANTOR TUJUAN": "kode_kantor_tujuan",
            "KODE KANTOR EKSPOR": "kode_kantor_ekspor",
            "KODE JENIS IMPOR": "kode_jenis_impor",
            "KODE JENIS EKSPOR": "kode_jenis_ekspor",
            "KODE JENIS TPB": "kode_jenis_tpb",
            "KODE JENIS PLB": "kode_jenis_plb", 
            "KODE JENIS PROSEDUR": "kode_jenis_pib",
            "KODE TUJUAN PEMASUKAN": "kode_tujuan_pemasukan",
            "KODE TUJUAN PENGIRIMAN": "kode_tujuan_pengiriman",
            "KODE TUJUAN TPB": "kode_tujuan_tpb",
            "KODE CARA DAGANG": "kode_cara_dagang",
            "KODE CARA BAYAR": "kode_cara_bayar",
            "KODE CARA BAYAR LAINNYA": "kode_cara_bayar_lainnya",
            "KODE GUDANG ASAL": "kode_gudang_asal",
            "KODE GUDANG TUJUAN": "kode_gudang_tujuan",
            "KODE JENIS KIRIM": "kode_jenis_kirim",
            "KODE JENIS PENGIRIMAN": "kode_jenis_pengiriman",
            "KODE KATEGORI EKSPOR": "kode_kategori_ekspor",
            "KODE KATEGORI MASUK FTZ": "kode_kategori_masuk_ftz",
            "KODE KATEGORI KELUAR FTZ": "kode_kategori_keluar_ftz",
            "KODE KATEGORI BARANG FTZ": "kode_kategori_barang_ftz",
            "KODE LOKASI": "kode_lokasi",
            "KODE LOKASI BAYAR": "kode_lokasi_bayar",
            "LOKASI ASAL": "lokasi_asal",
            "LOKASI TUJUAN": "lokasi_tujuan",
            "KODE DAERAH ASAL": "kode_daerah_asal",
            "KODE NEGARA TUJUAN": "kode_negara_tujuan",
            "KODE TUTUP PU": "kode_tutup_pu",
            "NOMOR BC11": "nomor_bc11",
            "TANGGAL BC11": "tanggal_bc11",
            "NOMOR POS": "nomor_pos",
            "NOMOR SUB POS": "nomor_sub_pos",
            "KODE PELABUHAN BONGKAR": "kode_pelabuhan_bongkar",
            "KODE PELABUHAN MUAT": "kode_pelabuhan_muat",
            "KODE PELABUHAN MUAT AKHIR": "kode_pelabuhan_muat_akhir",
            "KODE PELABUHAN TRANSIT": "kode_pelabuhan_transit",
            "KODE PELABUHAN TUJUAN": "kode_pelabuhan_tujuan",
            "KODE PELABUHAN EKSPOR": "kode_pelabuhan_ekspor",
            "KODE TPS": "kode_tps",
            "TANGGAL BERANGKAT": "tanggal_berangkat",
            "TANGGAL EKSPOR": "tanggal_ekspor",
            "TANGGAL MASUK": "tanggal_masuk",
            "TANGGAL MUAT": "tanggal_muat",
            "TANGGAL TIBA": "tanggal_tiba",
            "TANGGAL PERIKSA": "tanggal_periksa",
            "TEMPAT STUFFING": "tempat_stuffing",
            "TANGGAL STUFFING": "tanggal_stuffing",
            "KODE TANDA PENGAMAN": "kode_tanda_pengaman",
            "JUMLAH TANDA PENGAMAN": "jumlah_tanda_pengaman",
            "FLAG CURAH": "flag_curah",
            "FLAG SDA": "flag_sda",
            "FLAG AP BK": "flag_ap_bk",
            "FLAG VD": "flag_vd",
            "FLAG MIGAS": "flag_migas",
            "KODE ASURANSI": "kode_asuransi",
            "ASURANSI": "asuransi",
            "NILAI BARANG": "nilai_barang",
            "NILAI INCOTERM": "nilai_incoterm",
            "NILAI MAKLON": "nilai_maklon",
            "FREIGHT": "freight",
            "FOB": "fob",
            "BIAYA TAMBAHAN": "biaya_tambahan",
            "BIAYA PENGURANG": "biaya_pengurang",
            "VD": "vd",
            "CIF": "cif",
            "HARGA_PENYERAHAN": "harga_penyerahan",
            "NDPBM": "ndpbm",
            "TOTAL DANA SAWIT": "total_dana_sawit",
            "DASAR PENGENAAN PAJAK": "dasar_pengenaan_pajak",
            "NILAI JASA": "nilai_jasa",
            "UANG MUKA": "uang_muka",
            "BRUTO": "bruto",
            "NETTO": "netto",
            "VOLUME": "volume",
            "KOTA PERNYATAAN": "kota_pernyataan",
            "TANGGAL PERNYATAAN": "tanggal_pernyataan",
            "NAMA PERNYATAAN": "nama_pernyataan",
            "JABATAN PERNYATAAN": "jabatan_pernyataan",
            "KODE VALUTA": "kode_valuta",
            "KODE INCOTERM": "kode_incoterm",
            "KODE JASA KENA PAJAK": "kode_jasa_kena_pajak",
            "NOMOR BUKTI BAYAR": "nomor_bukti_bayar",
            "TANGGAL BUKTI BAYAR": "tanggal_bukti_bayar",
            "KODE JENIS NILAI": "kode_jenis_nilai",
            "KODE KANTOR MUAT": "kode_kantor_muat",
            "NOMOR DAFTAR": "nomor_daftar",
            "TANGGAL DAFTAR": "tanggal_daftar",
            "KODE ASAL BARANG FTZ": "kode_barang_asal_ftz",
            "KODE TUJUAN PENGELUARAN": "kode_tujuan_pengeluaran",
            "PPN PAJAK": "ppn_pajak",
            "PPNBM PAJAK": "ppnbm_pajak",
            "TARIF PPN PAJAK": "tarif_ppn_pajak",
            "TARIF PPNBM PAJAK": "tarif_ppnbm_pajak",
            "BARANG TIDAK BERWUJUD": "barang_tidak_berwujud",
            "KODE JENIS PENGELUARAN" :	"kode_jenis_pengeluaran",
            "BARANG KIRIMAN" :	"barang_kiriman",
            "KODE JENIS PENGANGKUTAN" : "kode_jenis_pengangkutan"
            
        }
        
        header_data = get_sheet_data("HEADER", expected_columns=HEADER_MAPPING.keys())
        if not header_data:
            return {"status": "error", "message": f"Sheet 'HEADER' is empty or missing. Available sheets: {', '.join(wb.sheetnames)}"}
        
        h_row = header_data[0]
        
        DATE_FIELDS = [
            "tanggal_bc11", "tanggal_berangkat", "tanggal_ekspor", "tanggal_masuk",
            "tanggal_muat", "tanggal_tiba", "tanggal_periksa", "tanggal_stuffing",
            "tanggal_pernyataan", "tanggal_bukti_bayar", "tanggal_daftar",
            "tanggal_ijin_entitas", "tanggal_dokumen", "tanggal_jaminan", 
            "tanggal_jatuh_tempo", "tanggal_bpj", "tanggal_daftar_asal"
        ]
        
        # Get nomorAju
        nomor_aju = h_row.get("NOMOR AJU")
        if not nomor_aju:
            return {"status": "error", "message": "NOMOR AJU is required in HEADER sheet"}

        # Check if exists
        existing_header = frappe.get_all("HEADER V21", filters={"nomoraju": nomor_aju}, limit=1)
        if existing_header:
             doc = frappe.get_doc("HEADER V21", existing_header[0].name)
        else:
            doc = frappe.new_doc("HEADER V21")
            doc.name = nomor_aju
            doc.nomoraju = nomor_aju
            
        # Map all Header Fields dynamically
        for excel_col, doc_field in HEADER_MAPPING.items():
            value = h_row.get(excel_col)
            if value is not None and value != "":
                if doc_field in DATE_FIELDS:
                    try:
                        doc.set(doc_field, getdate(value))
                    except:
                        pass
                elif isinstance(value, (int, float)):
                    doc.set(doc_field, flt(value))
                else:
                    doc.set(doc_field, str(value))
        
        doc.flags.ignore_links = True
        doc.save(ignore_permissions=True)
        
        if doc.name != nomor_aju:
            try:
                frappe.rename_doc("HEADER V21", doc.name, nomor_aju, force=True, merge=False)
                doc = frappe.get_doc("HEADER V21", nomor_aju)
            except Exception as rename_err:
                frappe.log_error(f"Could not rename HEADER V21 to {nomor_aju}: {rename_err}", "Rename Error")
        
        from datetime import datetime
       
        def parse_excel_date(val):
            if not val: return None
            val_str = str(val).strip()
            # Try specific format seen in file MM-DD-YYYY HH:mm:ss
            formats = [
                "%m-%d-%Y %H:%M:%S", 
                "%m-%d-%Y", 
                "%d-%m-%Y", 
                "%Y-%m-%d"
            ]
            for fmt in formats:
                try:
                    return datetime.strptime(val_str, fmt).date()
                except ValueError:
                    continue
            # Fallback to getdate
            return getdate(val)

        # Helper to create child table with stats
        def create_child(doctype, parent_field, sheet_name, mapping, optional=False):
            rows = get_sheet_data(sheet_name, optional, expected_columns=mapping.keys())
            child_list = []
            for row in rows:
                child_item = {}
                for excel_col, doc_field in mapping.items():
                    val = row.get(excel_col)
                    if val is not None and val != "":
                        if doc_field in DATE_FIELDS:
                            try:
                                child_item[doc_field] = parse_excel_date(val)
                            except Exception as e:
                                # Start DEBUG logging for specific fields
                                if doc_field == "tanggal_ijin_entitas":
                                     frappe.log_error(f"Date Parse Error {doc_field}: {val} -> {e}", "Import Debug")
                                pass
                        else:
                            child_item[doc_field] = val
                child_list.append(child_item)
            
            doc.set(doctype, child_list)
            audit_report["stats"][doctype] = len(child_list)

        # 2. Child Tables
        create_child("entitas", "parent", "ENTITAS", {
            "NOMOR AJU": "nomoraju",                    # ← TAMBAH INI
            "SERI": "seri",                             # ← UBAH DARI SERI ENTITAS
            "KODE ENTITAS": "kode_entitas",
            "NAMA ENTITAS": "nama_entitas",
            "ALAMAT ENTITAS": "alamat_entitas",
            "NOMOR IDENTITAS": "nomor_identitas",
            "KODE JENIS IDENTITAS": "kode_jenis_identitas",
            "NIB ENTITAS": "nib_entitas",
            "KODE JENIS API": "kode_jenis_api",
            "KODE NEGARA": "kode_negara",
            "KODE STATUS": "kode_status",
            "NOMOR IJIN ENTITAS": "nomor_ijin_entitas",
            "TANGGAL IJIN ENTITAS": "tanggal_ijin_entitas",
            "NIPER ENTITAS": "niper_entitas",
            "KODE KATEGORI KONSOLIDATOR": "kode_kategori_konsolidator"
        })
        
        create_child("kemasan", "parent", "KEMASAN", {
            "JUMLAH KEMASAN": "jumlah_kemasan",
            "KODE JENIS KEMASAN": "kode_jenis_kemasan",
            "MERK KEMASAN": "merek_kemasan",
            "SERI KEMASAN": "seri_kemasan"
        })
        
        create_child("dokumen", "parent", "DOKUMEN", {
            "KODE DOKUMEN": "kode_dokumen",
            "NOMOR DOKUMEN": "nomor_dokumen",
            "TANGGAL DOKUMEN": "tanggal_dokumen",
            "SERI DOKUMEN": "seri_dokumen"
        })
        
        create_child("pengangkut", "parent", "PENGANGKUT", {
            "NAMA PENGANGKUT": "nama_pengangkut",
            "NOMOR PENGANGKUT": "nomor_pengangkut",
            "KODE CARA ANGKUT": "kode_cara_angkut",
            "KODE BENDERA": "kode_bendera",
            "SERI PENGANGKUT": "seri_pengangkut"
        })
        
        create_child("kontainer", "parent", "KONTAINER", {
            "KODE TIPE KONTAINER": "kode_tipe_kontainer",
            "KODE UKURAN KONTAINER": "kode_ukuran_kontainer",
            "NOMOR KONTAINER": "nomor_kontainer",
            "SERI KONTAINER": "seri_kontainer",
            "KODE JENIS KONTAINER": "kode_jenis_kontainer"
        })

        # Additional HEADER child tables
        create_child("pungutan", "parent", "PUNGUTAN", {
            "KODE FASILITAS TARIF": "kode_fasilitas_tarif",
            "KODE JENIS PUNGUTAN": "kode_jenis_pungutan",
            "NILAI PUNGUTAN": "nilai_pungutan",
            "NPWP BILLING": "npwp_billing"
        })
        
        create_child("jaminan", "parent", "JAMINAN", {
            "KODE KANTOR": "kode_kantor",
            "KODE JAMINAN": "kode_jaminan",
            "NOMOR JAMINAN": "nomor_jaminan",
            "TANGGAL JAMINAN": "tanggal_jaminan",
            "NILAI JAMINAN": "nilai_jaminan",
            "PENJAMIN": "penjamin",
            "TANGGAL JATUH TEMPO": "tanggal_jatuh_tempo",
            "NOMOR BPJ": "nomor_bpj",
            "TANGGAL BPJ": "tanggal_bpj"
        })
        
        create_child("bank_devisa", "parent", "BANKDEVISA", {
            "SERI": "seri",
            "KODE": "kode",
            "NAMA": "nama"
        })
        
        create_child("komponen_biaya", "parent", "KOMPONENBIAYA", {
           "JENIS NILAI" : "jenisNilai",
            "HARGA INVOICE": "hargaInvoice",
            "PEMBAYARAN TIDAK LANGSUNG": "pembayaranTidakLangsung",
            "DISKON": "diskon",
            "KOMISI PENJUALAN": "komisiPenjualan",
            "BIAYA PENGEMASAN": "biayaPengemasan",
            "BIAYA PENGEPAKAN": "biayaPengepakan",
            "ASSIST": "assist",
            "ROYALTI": "royalti",
            "PROCEEDS": "proceeds",
            "BIAYA TRANSPORTASI": "biayaTransportasi",
            "BIAYA PEMUATAN": "biayaPemuatan",
            "ASURANSI": "asuransi",
            "GARANSI": "garansi",
            "BIAYA KEPENTINGAN SENDIRI": "biayaKepentinganSendiri",
            "BIAYA PASCA IMPOR": "biayaPascaImpor",
            "BIAYA PAJAK INTERNAL": "biayaPajakInternal",
            "BUNGA": "bunga",
            "DEVIDEN": "deviden"
        }, optional=True)

        doc.save(ignore_permissions=True)
        
        # 3. Barang V1
        BARANG_MAPPING = {
            "SERI BARANG": "seri_barang",
            "HS": "hs",
            "KODE BARANG": "kode_barang",
            "URAIAN": "uraian",
            "MERK": "merek",
            "TIPE": "tipe",
            "UKURAN": "ukuran",
            "SPESIFIKASI LAIN": "spesifikasi_lain",
            "KODE SATUAN": "kode_satuan",
            "JUMLAH SATUAN": "jumlah_satuan",
            "KODE KEMASAN": "kode_kemasan",
            "JUMLAH KEMASAN": "jumlah_kemasan",
            "NETTO": "netto",
            "BRUTO": "bruto",
            "VOLUME": "volume",
            "CIF": "cif",
            "CIF RUPIAH": "cif_rupiah",
            "NDPBM": "ndpbm",
            "FOB": "fob",
            "ASURANSI": "asuransi",
            "FREIGHT": "freight",
            "NILAI TAMBAH": "nilai_tambah",
            "DISKON": "diskon",
            "HARGA PENYERAHAN": "harga_penyerahan",
            "HARGA PEROLEHAN": "harga_perolehan",
            "HARGA SATUAN": "harga_satuan",
            "HARGA EKSPOR": "harga_ekspor",
            "HARGA PATOKAN": "harga_patokan",
            "NILAI BARANG": "nilai_barang",
            "NILAI DANA SAWIT": "nilai_dana_sawit",
            "NILAI DEVISA": "nilai_devisa",
            "KODE NEGARA ASAL": "kode_negara_asal",
            "KODE KONDISI BARANG": "kode_kondisi_barang",
            "PERNYATAAN LARTAS": "pernyataan_lartas"
        }
        
        barang_rows = get_sheet_data("BARANG", expected_columns=BARANG_MAPPING.keys())
        audit_report["stats"]["BARANG V1"] = len(barang_rows)
        
        # Child Data Pre-fetch
        tarif_rows = get_sheet_data("BARANGTARIF", expected_columns=["SERI BARANG", "KODE PUNGUTAN", "KODE TARIF", "TARIF", "NILAI BAYAR", "NILAI FASILITAS", "TARIF FASILITAS", "KODE FASILITAS", "JUMLAH SATUAN", "KODE SATUAN"])
        bd_rows = get_sheet_data("BARANGDOKUMEN", expected_columns=["SERI BARANG", "SERI DOKUMEN", "SERI IZIN"])
        be_rows = get_sheet_data("BARANGENTITAS", expected_columns=["SERI BARANG", "SERI ENTITAS"])
        bsk_rows = get_sheet_data("BARANGSPEKKHUSUS", expected_columns=["SERI BARANG", "KODE", "URAIAN"])
        bvd_rows = get_sheet_data("BARANGVD", expected_columns=["SERI BARANG", "KODE VD", "NILAI BARANG"])
        bb_rows = get_sheet_data("BAHANBAKU", optional=True, expected_columns=["SERI BARANG", "SERI BAHAN BAKU", "KODE ASAL BAHAN BAKU", "HS", "KODE BARANG", "URAIAN", "MEREK", "TIPE", "UKURAN", "SPESIFIKASI LAIN", "KODE SATUAN", "JUMLAH SATUAN", "NETTO"])
        bbt_rows = get_sheet_data("BAHANBAKUTARIF", optional=True, expected_columns=["SERI BARANG", "SERI BAHAN BAKU", "KODE PUNGUTAN", "KODE TARIF", "TARIF", "KODE FASILITAS", "TARIF FASILITAS", "NILAI BAYAR", "NILAI FASILITAS"])
        bbd_rows = get_sheet_data("BAHANBAKUDOKUMEN", optional=True, expected_columns=["SERI BARANG", "SERI BAHAN BAKU", "SERI DOKUMEN", "SERI IZIN"])

        for b_row in barang_rows:
            seri_barang = cint(b_row.get("SERI BARANG"))
            if not seri_barang: continue
            
            # Check/Create Barang
            existing_barang = frappe.get_all("BARANG V1", filters={"nomoraju": nomor_aju, "seri_barang": seri_barang})
            if existing_barang:
                b_doc = frappe.get_doc("BARANG V1", existing_barang[0].name)
            else:
                b_doc = frappe.new_doc("BARANG V1")
                b_doc.nomoraju = nomor_aju
                b_doc.seri_barang = seri_barang
            
            for excel_col, doc_field in BARANG_MAPPING.items():
                value = b_row.get(excel_col)
                if value is not None and value != "":
                    if isinstance(value, (int, float)):
                        b_doc.set(doc_field, flt(value))
                    else:
                        b_doc.set(doc_field, str(value))
            
            b_doc.flags.ignore_links = True
            b_doc.save(ignore_permissions=True)
            
            # Filter and Set Child Tables
            # In-memory filtering for speed
            
            # BARANG TARIF
            b_tarifs = [
                {
                    "seri_barang": seri_barang,
                    "kode_pungutan": r.get("KODE PUNGUTAN"),
                    "kode_tarif": r.get("KODE TARIF"),
                    "tarif": flt(r.get("TARIF")),
                    "nilai_bayar": flt(r.get("NILAI BAYAR")),
                    "nilai_fasilitas": flt(r.get("NILAI FASILITAS")),
                    "tarif_fasilitas": flt(r.get("TARIF FASILITAS")),
                    "kode_fasilitas": r.get("KODE FASILITAS"),
                    "jumlah_satuan": flt(r.get("JUMLAH SATUAN")),
                    "kode_satuan": r.get("KODE SATUAN")
                }
                for r in tarif_rows if cint(r.get("SERI BARANG")) == seri_barang
            ]
            if b_tarifs: b_doc.set("barang_tarif", b_tarifs)

            # BARANG DOKUMEN
            b_docs = [
                {
                    "seri_dokumen": r.get("SERI DOKUMEN"),
                    "seri_izin": r.get("SERI IZIN")
                }
                for r in bd_rows if cint(r.get("SERI BARANG")) == seri_barang
            ]
            if b_docs: b_doc.set("barang_dokumen", b_docs)

            # BARANG ENTITAS (pemilik)
            b_entitas = [
                { "seri_entitas": cint(r.get("SERI ENTITAS")) }
                for r in be_rows if cint(r.get("SERI BARANG")) == seri_barang
            ]
            if b_entitas: b_doc.set("barang_pemilik", b_entitas)

            # BARANG SPEK KHUSUS
            b_spek = [
                {
                    "kode_spek_khusus": r.get("KODE"),
                    "uraian": r.get("URAIAN")
                }
                for r in bsk_rows if cint(r.get("SERI BARANG")) == seri_barang
            ]
            if b_spek: b_doc.set("barang_spek_khusus", b_spek)

            # BARANG VD
            b_vd = [
                {
                    "kode_jenis_vd": r.get("KODE VD"),
                    "nilai_barang": flt(r.get("NILAI BARANG"))
                }
                for r in bvd_rows if cint(r.get("SERI BARANG")) == seri_barang
            ]
            if b_vd: b_doc.set("barang_vd", b_vd)

            b_doc.save(ignore_permissions=True)
            
            # BAHAN BAKU
            for bb_row in bb_rows:
                if cint(bb_row.get("SERI BARANG")) == seri_barang:
                    seri_bahan_baku = cint(bb_row.get("SERI BAHAN BAKU"))
                    
                    existing_bb = frappe.get_all("BAHAN BAKU", filters={
                        "nomoraju": nomor_aju, 
                        "seri_barang": seri_barang,
                        "seri_bahan_baku": seri_bahan_baku
                    })
                    
                    if existing_bb:
                        bb_doc = frappe.get_doc("BAHAN BAKU", existing_bb[0].name)
                    else:
                        bb_doc = frappe.new_doc("BAHAN BAKU")
                        bb_doc.nomoraju = nomor_aju
                        bb_doc.seri_barang = seri_barang
                        bb_doc.seri_bahan_baku = seri_bahan_baku

                    bb_doc.parent_barang = b_doc.name
                    bb_doc.kode_asal_bahan_baku = bb_row.get("KODE ASAL BAHAN BAKU")
                    bb_doc.hs = bb_row.get("HS")
                    bb_doc.kode_barang = bb_row.get("KODE BARANG")
                    bb_doc.uraian = bb_row.get("URAIAN")
                    bb_doc.merek = bb_row.get("MEREK")
                    bb_doc.tipe = bb_row.get("TIPE")
                    bb_doc.ukuran = bb_row.get("UKURAN")
                    bb_doc.spesifikasi_lain = bb_row.get("SPESIFIKASI LAIN")
                    bb_doc.kode_satuan = bb_row.get("KODE SATUAN")
                    bb_doc.jumlah_satuan = flt(bb_row.get("JUMLAH SATUAN"))
                    bb_doc.netto = flt(bb_row.get("NETTO"))
                    
                    # BB TARIF
                    bb_tarifs = [
                        {
                            "kode_pungutan": r.get("KODE PUNGUTAN"),
                            "kode_tarif": r.get("KODE TARIF"),
                            "tarif": flt(r.get("TARIF")),
                            "kode_fasilitas": r.get("KODE FASILITAS"),
                            "tarif_fasilitas": flt(r.get("TARIF FASILITAS")),
                            "nilai_bayar": flt(r.get("NILAI BAYAR")),
                            "nilai_fasilitas": flt(r.get("NILAI FASILITAS"))
                        }
                        for r in bbt_rows 
                        if (cint(r.get("SERI BARANG")) == seri_barang and cint(r.get("SERI BAHAN BAKU")) == seri_bahan_baku)
                    ]
                    if bb_tarifs: bb_doc.set("bahan_tarif", bb_tarifs)
                    
                    # BB DOKUMEN
                    bb_docs = [
                        {
                            "seri_dokumen": r.get("SERI DOKUMEN"),
                            "seri_izin": r.get("SERI IZIN")
                        }
                        for r in bbd_rows
                        if (cint(r.get("SERI BARANG")) == seri_barang and cint(r.get("SERI BAHAN BAKU")) == seri_bahan_baku)
                    ]
                    if bb_docs: bb_doc.set("bahan_baku_dokumen", bb_docs)
                    
                    bb_doc.flags.ignore_links = True
                    bb_doc.save(ignore_permissions=True)

        # Helper to generate audit message
        def generate_audit_msg(base_msg):
            parts = [base_msg]
            
            if audit_report["stats"]:
                parts.append("<br><b>Import Statistics:</b>")
                for table, count in audit_report["stats"].items():
                    parts.append(f"- {table}: {count} records")
            
            if audit_report["unmapped_columns"]:
                parts.append("<br><b>Warning: Extra Columns (Unmapped) in Excel:</b>")
                for sheet, cols in audit_report["unmapped_columns"].items():
                    parts.append(f"- {sheet}: {', '.join(cols)}")

            if audit_report["missing_columns"]:
                parts.append("<br><b>Warning: Missing Columns in Excel (Expected by System):</b>")
                for sheet, cols in audit_report["missing_columns"].items():
                    parts.append(f"- {sheet}: {', '.join(cols)}")
            
            return "<br>".join(parts)

        success_msg = generate_audit_msg(f"Successfully processed Header {nomor_aju}.")
        
        if dry_run:
            frappe.db.rollback()
            return {
                "status": "success", 
                "message": "<b>[SIMULATION MODE]</b><br>" + success_msg + "<br><br><i>No changes were saved to the database.</i>", 
                "audit": audit_report,
                "dry_run": True
            }

        return {"status": "success", "message": success_msg, "audit": audit_report}

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Import CEISA Excel Error")
        # Re-construct audit msg for error too
        error_msg = f"<b>Error during import:</b> {str(e)}"
        
        # We need to recreate the generate function or just do it inline since scope issues
        parts = [error_msg]
        if audit_report["stats"]:
            parts.append("<br><b>Partial Statistics:</b>")
            for table, count in audit_report["stats"].items():
                parts.append(f"- {table}: {count} records")
        
        if audit_report["missing_columns"]:
            parts.append("<br><b>Missing Columns (Possible Cause):</b>")
            for sheet, cols in audit_report["missing_columns"].items():
                parts.append(f"- {sheet}: {', '.join(cols)}")
                
        if audit_report["unmapped_columns"]:
             parts.append("<br><b>Unmapped Columns:</b>")
             for sheet, cols in audit_report["unmapped_columns"].items():
                parts.append(f"- {sheet}: {', '.join(cols)}")

        return {"status": "error", "message": "<br>".join(parts), "audit": audit_report}
