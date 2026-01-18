from os import stat
import frappe
from frappe.query_builder import DocType
import requests
import json


SUCCESS = 200
NOT_FOUND = 400

@frappe.whitelist(allow_guest=True)
def get_all_headerq():
    #headerq = frappe.qb.Table("tabHEADER")
    headerq = frappe.qb.from_('HEADER').select('name' , 'owner', 'nomor_aju' , 'kode_dokumen' ).run(as_dict=True)
    #headerq = frappe.qb.from_('HEADER').select('name').as_('nomorAju').run(as_dict=True)
    return headerq
    #str(query)
    # SELECT "id","fname","lname","phone" FROM "tabCustomer"

    #query.get_sql()
    # SELECT "id","fname","lname","phone" FROM "tabCustomer"

    #str(query) == query.get_sql()



    #headerq = frappe.qb.DocType('HEADER')
        #filtered_data = (
            #frappe.qb.from_(headerq)
                #.select(headerq.kode_kantor_bongkar, headerq.nomoraju, headerq.name, headerq.referensi_kantor)
                #.where((customer.fname == 'Max') | customer.id.like('RA%'))
                #.where(customer.lname == 'Mustermann')
        #).run(as_dict=True)

    #return headerq

frappe.whitelist(allow_guest=True)
def authenticate_and_get_token(username, password):
    auth_url = "https://example.com/api/authenticate"
    auth_data = {
        "username": username,
        "password": password
    }
    try:
        response = requests.post(auth_url, json=auth_data)
        response.raise_for_status()  # Raise an exception for 4xx or 5xx status codes
        token = response.json().get("access_token")
        if token:
            return token
        else:
            raise ValueError("No access token found in response")
    except requests.exceptions.RequestException as e:
        print("Authentication error:", e)
        return None

def get_parent_data():
    parent_data = {}

    # Query to fetch parent data and its related children and grandchildren
    parents = frappe.get_all("Parent", filters={"name": "Parent1"}, fields=["name"])
    for parent in parents:
        parent_name = parent.name
        children = frappe.get_all("Child", filters={"parent": parent_name}, fields=["name"])
        parent_data[parent_name] = {"children": {}}
        for child in children:
            child_name = child.name
            grandchildren = frappe.get_all("Grandchild", filters={"parent": child_name}, fields=["name"])
            parent_data[parent_name]["children"][child_name] = {"grandchildren": []}
            for grandchild in grandchildren:
                parent_data[parent_name]["children"][child_name]["grandchildren"].append(grandchild.name)

    return parent_data


def send_data_to_other_app(data, token):
    url = "https://example.com/api/data"
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json"
    }
    try:
        response = requests.post(url, json=data, headers=headers)
        response.raise_for_status()  # Raise an exception for 4xx or 5xx status codes
        print("Data sent successfully:", response.json())
    except requests.exceptions.RequestException as e:
        print("Error sending data:", e)

# Example usage:
    username = "your_username"
    password = "your_password"

    token = authenticate_and_get_token(username, password)
    if token:
        parent_data = get_parent_data()
        send_data_to_other_app(parent_data, token)



@frappe.whitelist(allow_guest=True)
def get_nested_data_all():    # Example: Retrieve data from Frappe doctypes
    #parent_data = frappe.get_doc("HEADER V2", "nomoraju")
    parent_data = frappe.get_doc("HEADER V2", "000020BT000120241128505790")
    entitas_data = frappe.get_list("ENTITAS",  filters={"parent": parent_data.name}, fields=[ 'seri', 'kode_entitas', 'parent', 'alamat_entitas','kode_jenis_identitas','nama_entitas','nib_entitas','nomor_identitas','nomor_ijin_entitas','tanggal_ijin_entitas'],)
    #entitas_data = parent_data.as_dict().get('ENTITAS', [])
    #entitas_data = [entitas.as_dict() for entitas in parent_data.get("nomoraju")]
    # barang_data = frappe.get_doc("BARANG V1", "000023BT000120240110000050" )
    # barang_data.as_dict()
    
    barang_data = frappe.get_list("BARANG V1", filters={"nomoraju": parent_data.name}, fields=['name','nomoraju', 'seri_barang','asuransi','cif','diskon','fob','freight','harga_ekspor','harga_penyerahan','harga_satuan','isi_per_kemasan','jumlah_kemasan','jumlah_satuan','kode_barang','kode_dokumen_asal','kode_kategori_barang','kode_negara_asal','kode_perhitungan','kode_satuan','merek','netto','nilai_barang','nilai_tambah','hs','spesifikasi_lain','tipe','ukuran','uraian','ndpbm','cif_rupiah','harga_perolehan','kode_asal_barang'],)
    barang_tarif_data = frappe.get_list("BARANG TARIF", fields=['nomoraju', 'seri_barang', 'parent','name','kode_tarif','jumlah_satuan','kode_fasilitas','kode_satuan','kode_pungutan','nilai_bayar','nilai_fasilitas','nilai_sudah_dilunasi','tarif','tarif_fasilitas'],)
    barang_dokumen_data = frappe.get_list("BARANG DOKUMEN",  fields=['seri_izin', 'seri_barang', 'parent','name'],)

    #barang_data1 = frappe.db.get_value("BARANG V1", 'TASK00002', 'subject')
    #barang_data1 = frappe.get_doc({doctype”: "BARANG V1", })
    #barang_tarif_data = frappe.get_list("BARANG TARIF", filters={"parent": barang_data1.name}, fields=['nomoraju', 'seri_barang', 'parent','name'],)
    #barang_dokumen_data = frappe.get_list("BARANG DOKUMEN", filters={"parent": barang_data1.name}, fields=['seri_izin', 'seri_barang', 'parent','name'],)
    #barangTarifItem = frappe.get_list("KEMASAN", filters={"parent": "000023BT000120240110000050"}, fields=['nomoraju', 'name'],)
    
    kemasan_data = frappe.get_list("KEMASAN", filters={"parent": parent_data.name}, fields=[ 'jumlah_kemasan','merek_kemasan', 'kode_kemasan','seri','name'],)
    kontainer_data = frappe.get_list("KONTAINER", filters={"parent": parent_data.name}, fields=['name','kode_jenis_kontainer','kode_tipe_kontainer','kode_ukuran_kontainer','nomor_kontainer','seri'],)
    dokumen_data = frappe.get_list("DOKUMEN", filters={"parent": parent_data.name}, fields=['name', 'kode_dokumen','nomor_dokumen','seri','tanggal_dokumen'],)
    pengangkut_data = frappe.get_list("PENGANGKUT", filters={"parent": parent_data.name}, fields=['name','call_sign','kode_bendera','nama_pengangkut','nomor_pengangkut','kode_cara_angkut','seri_pengangkut'],)

    # Format the data into the nested JSON structure with aliases
    nested_data_all = {
       # "data": {
            "asalData": parent_data.asaldata,  # Alias for parent field 1
            "nomorAju": parent_data.name,  # Alias for parent field 1
            "kodeDokumen": parent_data.kode_dokumen,  # Alias for parent field 2
            "asuransi":float(parent_data.asuransi),
            "bruto":float(parent_data.bruto),
            "cif":float(parent_data.cif),
            "fob":float(parent_data.fob),
            "freight":float(parent_data.freight),
            "hargaPenyerahan":float(parent_data.harga_penyerahan),
            "nik": parent_data.id_pengguna,
            "jabatanTtd":parent_data.jabatan_pernyataan,
            "jumlahKontainer":0,
            "kodeAsuransi":parent_data.kode_asuransi,
            "kodeIncoterm":parent_data.kode_incoterm,
            "kodeKantor": parent_data.kode_kantor,
            "kodeKantorBongkar":parent_data.kode_kantor_bongkar,
            "kodePelBongkar":parent_data.kode_pelabuhan_bongkar,
            "kodePelMuat":parent_data.kode_pelabuhan_muat,
            "kodePelTransit":parent_data.kode_pelabuhan_transit,
            "kodeTps":parent_data.kode_tps,
            "kodeTujuanTpb":parent_data.kode_tujuan_tpb,
            "kodeTutupPu":parent_data.kode_tutup_pu,
            "kodeValuta":parent_data.kode_valuta,
            "kotaTtd":parent_data.kota_pernyataan,
            "namaTtd":parent_data.nama_pernyataan,
            "ndpbm":float(parent_data.ndpbm),
            "netto":float(parent_data.netto),
            "nilaiBarang":float(parent_data.nilai_barang),	
            "nomorBc11":parent_data.nomor_bc11, 
            "posBc11":parent_data.nomor_pos,
            "seri":int(parent_data.seri_dokumen),
            "subposBc11":parent_data.nomor_sub_pos,
            "tanggalBc11":parent_data.tanggal_bc11,
            "tanggalTiba":parent_data.tanggal_tiba,
            "tanggalTtd":parent_data.tanggal_pernyataan,
            "biayaTambahan":float(parent_data.biaya_tambahan),
            "biayaPengurang":float(parent_data.biaya_pengurang),	
            "kodeKenaPajak":parent_data.kode_jasa_kena_pajak,
            "entitas": [
                {
                    "seriEntitas": int(entitas.seri),
                    "kodeEntitas": entitas.kode_entitas,
                    "alamatEntitas": entitas.alamat_entitas,
                    "kodeJenisIdentitas": entitas.kode_jenis_identitas,
                    "namaEntitas": entitas.nama_entitas,
                    "nibEntitas": entitas.nib_entitas,
                    "nomorIdentitas": entitas.nomor_identitas,
                    "nomorIjinEntitas": entitas.nomor_ijin_entitas,
                    "tanggalIjinEntitas": entitas.tanggal_ijin_entitas
                    

                 
                }
                for entitas in entitas_data
            ],
            "barang": [
                {
                    "Nomor_aju_brg": child.nomoraju,  # Alias for child field 1
                    "kode_brg": child.name,  # Alias for child field 1
                    "Seri_barang_brg": child.seri_barang,  # Alias for child field 2
                    "asuransi": float(child.asuransi),
                    "cif": float(child.cif),
                    "diskon": float(child.diskon),
                    "fob": float(child.fob),
                    "freight": float(child.freight),
                    "hargaEkspor": float(child.harga_ekspor),
                    "hargaPenyerahan": float(child.harga_penyerahan),
                    "hargaSatuan": float(child.harga_satuan),
                    "isiPerKemasan": float(child.isi_per_kemasan),
                    "jumlahKemasan": float(child.jumlah_kemasan),
                    "jumlahSatuan": float(child.jumlah_satuan),
                    "kodeBarang": child.kode_barang,
                    "kodeDokumen": child.kode_dokumen_asal,
                    "kodeKategoriBarang": child.kode_kategori_barang,
                    "kodeJenisKemasan": child.kode_jenis_kemasan,
                    "kodeNegaraAsal": child.kode_negara_asal,
                    "kodePerhitungan": child.kode_perhitungan,
                    "kodeSatuanBarang": child.kode_satuan,
                    "merk": child.merek,
                    "netto": float(child.netto),
                    "nilaiBarang": float(child.nilai_barang),
                    "nilaiTambah": float(child.nilai_tambah),
                    "posTarif": child.hs,
                    "seriBarang": child.seri_barang,
                    "spesifikasiLain": child.spesifikasi_lain,
                    "tipe": child.tipe,
                    "ukuran": child.ukuran,
                    "uraian": child.uraian,
                    "ndpbm": float(child.ndpbm),
                    "cifRupiah": float(child.cif_rupiah),
                    "hargaPerolehan": float(child.harga_perolehan),
                    "kodeAsalBahanBaku": child.kode_asal_barang,
                    "barangTarif": [
                       {
                           #"kode_brg": item.parent,  # Alias for grandchild field 1
                           #"grandchild_field1": item.nomoraju,  # Alias for grandchild field 1
                           #"grandchild_field2": item.seri_barang,  # Alias for grandchild field 2
                           "kodeJenisTarif": item.kode_tarif,
                           "jumlahSatuan": float(item.jumlah_satuan),
                           "kodeFasilitasTarif": item.kode_fasilitas,
                           "kodeSatuanBarang": item.kode_satuan,
                           "kodeJenisPungutan": item.kode_pungutan,
                           "nilaiBayar": float(item.nilai_bayar),
                           "nilaiFasilitas": float(item.nilai_fasilitas),
                           "nilaiSudahDilunasi": float(item.nilai_sudah_dilunasi),
                           "seriBarang": int(item.seri_barang),
                           "tarif": float(item.tarif),
                           "tarifFasilitas": float(item.tarif_fasilitas)
                       }
                       for item in barang_tarif_data if item.parent == child.name
                    ],
                    "barangDokumen": [
                       {
                        #    "kode_brg": item.parent,  # Alias for grandchild field 1
                        #    "seriIzin": item.seri_izin,  # Alias for grandchild field 1
                        #    "seriBarang": item.seri_barang,  # Alias for grandchild field 2
                           "seriDokumen": item.seri_dokumen
                       }
                       for item in barang_dokumen_data if item.parent == child.name
                    ]
                }
                for child in barang_data
            ],
            "kemasan": [
                {
                    "jumlahKemasan": kemasan.jumlah_kemasan,
                    "kodeJenisKemasan": kemasan.kode_kemasan,
                    "merkKemasan": kemasan.merek_kemasan,
                    "seriKemasan": int(kemasan.seri)
                
                }
                for kemasan in kemasan_data
            ],
            "kontainer": [
                {
                    "kodeJenisKontainer": kontainer.kode_jenis_kontainer,
                    "kodeTipeKontainer": kontainer.kode_tipe_kontainer,
                    "kodeUkuranKontainer": kontainer.kode_ukuran_kontainer,
                    "nomorKontainer": kontainer.nomor_kontainer,
                    "seriKontainer": int(kontainer.seri)
                                    
                }
                for kontainer in kontainer_data
            ],
             "dokumen": [
                {
                    "kodeDokumen": dokumen.kode_dokumen,
                    "nomorDokumen": dokumen.nomor_dokumen,
                    "seriDokumen": int(dokumen.seri),
                    "tanggalDokumen": dokumen.tanggal_dokumen
                   
                    #"grandchildren": [
                    #    {
                    #        "grandchild_field1": grandchild.grandchild_field1,  # Alias for grandchild field 1
                    #        "grandchild_field2": grandchild.grandchild_field2  # Alias for grandchild field 2
                    #    }
                    #    for grandchild in grandchildren_data if grandchild.parent == child.name
                    #]
                }
                for dokumen in dokumen_data
            ],
             "pengangkut": [
                {
                    "callSign": pengangkut.call_sign, # Alias for child field 1
                    "kodeBendera": pengangkut.kode_bendera,
                    "namaPengangkut": pengangkut.nama_pengangkut,
                    "nomorPengangkut": pengangkut.nomor_pengangkut,
                    "kodeCaraAngkut": pengangkut.kode_cara_angkut,
                    "seriPengangkut": int(pengangkut.seri_pengangkut)
               
                    #"grandchildren": [
                    #    {
                    #        "grandchild_field1": grandchild.grandchild_field1,  # Alias for grandchild field 1
                    #        "grandchild_field2": grandchild.grandchild_field2  # Alias for grandchild field 2
                    #    }
                    #    for grandchild in grandchildren_data if grandchild.parent == child.name
                    #]
                }
                for pengangkut in pengangkut_data
            ],
       # }
    }

    return nested_data_all


@frappe.whitelist(allow_guest=True)
def get_nested_data_bc20():    # Example: Retrieve data from Frappe doctypes
    parent_data = frappe.get_doc("HEADER V2", "000020BT000120240522000060")
    entitas_data = frappe.get_list("ENTITAS",  filters={"parent": parent_data.name}, fields=[ 'seri', 'kode_entitas', 'parent', 'alamat_entitas','kode_jenis_identitas','nama_entitas','nib_entitas','nomor_identitas','nomor_ijin_entitas','tanggal_ijin_entitas'],)
       
    barang_data = frappe.get_list("BARANG V1", filters={"nomoraju": parent_data.name}, fields=['name','nomoraju', 'seri_barang','asuransi','cif','diskon','fob','freight','harga_ekspor','harga_penyerahan','harga_satuan','isi_per_kemasan','jumlah_kemasan','jumlah_satuan','kode_barang','kode_dokumen_asal','kode_kategori_barang','kode_negara_asal','kode_perhitungan','kode_satuan','merek','netto','nilai_barang','nilai_tambah','hs','spesifikasi_lain','tipe','ukuran','uraian','ndpbm','cif_rupiah','harga_perolehan','kode_asal_barang'],)
    barang_tarif_data = frappe.get_list("BARANG TARIF", fields=['nomoraju', 'seri_barang', 'parent','name','kode_tarif','jumlah_satuan','kode_fasilitas','kode_satuan','kode_pungutan','nilai_bayar','nilai_fasilitas','nilai_sudah_dilunasi','tarif','tarif_fasilitas'],)
    barang_dokumen_data = frappe.get_list("BARANG DOKUMEN",  fields=['seri_izin', 'seri_barang', 'parent','name'],)
    barang_spekkhusus_data = frappe.get_list("BARANG SPEK KHUSUS",  fields=['seri_barang', 'parent','name'],)
    barang_vd_data = frappe.get_list("BARANG VD",  fields=['seri_barang', 'parent','name'],)
    barang_pemilik_data = frappe.get_list("BARANG ENTITAS", fields=['seri_barang', 'parent','name'],)
     
    kemasan_data = frappe.get_list("KEMASAN", filters={"parent": parent_data.name}, fields=[ 'jumlah_kemasan','merek_kemasan', 'kode_kemasan','seri','name'],)
    kontainer_data = frappe.get_list("KONTAINER", filters={"parent": parent_data.name}, fields=['name','kode_jenis_kontainer','kode_tipe_kontainer','kode_ukuran_kontainer','nomor_kontainer','seri'],)
    dokumen_data = frappe.get_list("DOKUMEN", filters={"parent": parent_data.name}, fields=['name', 'kode_dokumen','nomor_dokumen','seri','tanggal_dokumen'],)
    pengangkut_data = frappe.get_list("PENGANGKUT", filters={"parent": parent_data.name}, fields=['name','call_sign','kode_bendera','nama_pengangkut','nomor_pengangkut','kode_cara_angkut','seri_pengangkut'],)

    # Format the data into the nested JSON structure with aliases
    nested_data_bc20 = {
       # "data": {
            "asalData": parent_data.asaldata,  # Alias for parent field 1
            "nomorAju": parent_data.name,  # Alias for parent field 1
            "kodeDokumen": parent_data.kode_dokumen,  # Alias for parent field 2
            "asuransi":float(parent_data.asuransi),
            "bruto":float(parent_data.bruto),
            "cif":float(parent_data.cif),
            "fob":float(parent_data.fob),
            "freight":float(parent_data.freight),
            "hargaPenyerahan":float(parent_data.harga_penyerahan),
            "nik": parent_data.id_pengguna,
            "jabatanTtd":parent_data.jabatan_pernyataan,
            "kodeAsuransi":parent_data.kode_asuransi,
            "kodeIncoterm":parent_data.kode_incoterm,
            "kodeKantor": parent_data.kode_kantor,
            "kodeKantorBongkar":parent_data.kode_kantor_bongkar,
            "kodePelBongkar":parent_data.kode_pelabuhan_bongkar,
            "kodePelMuat":parent_data.kode_pelabuhan_muat,
            "kodePelTransit":parent_data.kode_pelabuhan_transit,
            "kodeTps":parent_data.kode_tps,
            "kodeTutupPu":parent_data.kode_tutup_pu,
            "kodeValuta":parent_data.kode_valuta,
            "kotaTtd":parent_data.kota_pernyataan,
            "namaTtd":parent_data.nama_pernyataan,
            "ndpbm":float(parent_data.ndpbm),
            "netto":float(parent_data.netto),
            "nilaiBarang":float(parent_data.nilai_barang),	
            "nomorBc11":parent_data.nomor_bc11, 
            "posBc11":parent_data.nomor_pos,
            "seri":int(parent_data.seri_dokumen),
            "subposBc11":parent_data.nomor_sub_pos,
            "tanggalBc11":parent_data.tanggal_bc11,
            "tanggalTiba":parent_data.tanggal_tiba,
            "tanggalTtd":parent_data.tanggal_pernyataan,
            "biayaTambahan":float(parent_data.biaya_tambahan),
            "biayaPengurang":float(parent_data.biaya_pengurang),
            "disclaimer":parent_data.tanggal_pernyataan,
            "flagVd":parent_data.tanggal_pernyataan,
            "idPengguna":parent_data.tanggal_pernyataan,
            "jumlahTandaPengaman":int(parent_data.biaya_pengurang),
            "kodeCaraBayar":parent_data.tanggal_pernyataan,
            "kodeJenisImpor":parent_data.tanggal_pernyataan,
            "kodeJenisNilai":parent_data.tanggal_pernyataan,
            "kodeJenisProsedur":parent_data.tanggal_pernyataan,
            "kodePelTujuan":parent_data.tanggal_pernyataan,
            "nilaiIncoterm":float(parent_data.biaya_pengurang),
            "nilaiMaklon":float(parent_data.biaya_pengurang),
            "tanggalAju":parent_data.tanggal_pernyataan,
            "totalDanaSawit":float(parent_data.biaya_pengurang),
            "vd":float(parent_data.vd),
            "volume":float(parent_data.biaya_pengurang),	
            "entitas": [
                {
                    "seriEntitas": int(entitas.seri),
                    "kodeEntitas": entitas.kode_entitas,
                    "alamatEntitas": entitas.alamat_entitas,
                    "kodeJenisIdentitas": entitas.kode_jenis_identitas,
                    "namaEntitas": entitas.nama_entitas,
                    "nibEntitas": entitas.nib_entitas,
                    "nomorIdentitas": entitas.nomor_identitas,
                    "nomorIjinEntitas": entitas.nomor_ijin_entitas,
                    "tanggalIjinEntitas": entitas.tanggal_ijin_entitas
                                    
                }
                for entitas in entitas_data
            ],
            "barang": [
                {
                    "Nomor_aju_brg": child.nomoraju,  # Alias for child field 1
                    "kode_brg": child.name,  # Alias for child field 1
                    "Seri_barang_brg": child.seri_barang,  # Alias for child field 2
                    "asuransi": float(child.asuransi),
                    "cif": float(child.cif),
                    "diskon": float(child.diskon),
                    "fob": float(child.fob),
                    "freight": float(child.freight),
                    "hargaEkspor": float(child.harga_ekspor),
                    "hargaPenyerahan": float(child.harga_penyerahan),
                    "hargaSatuan": float(child.harga_satuan),
                    "isiPerKemasan": float(child.isi_per_kemasan),
                    "jumlahKemasan": float(child.jumlah_kemasan),
                    "jumlahSatuan": float(child.jumlah_satuan),
                    "kodeBarang": child.kode_barang,
                    "kodeDokumen": child.kode_dokumen_asal,
                    "kodeKategoriBarang": child.kode_kategori_barang,
                    "kodeJenisKemasan": child.kode_jenis_kemasan,
                    "kodeNegaraAsal": child.kode_negara_asal,
                    "kodePerhitungan": child.kode_perhitungan,
                    "kodeSatuanBarang": child.kode_satuan,
                    "merk": child.merek,
                    "netto": float(child.netto),
                    "nilaiBarang": float(child.nilai_barang),
                    "nilaiTambah": float(child.nilai_tambah),
                    "posTarif": child.hs,
                    "seriBarang": child.seri_barang,
                    "spesifikasiLain": child.spesifikasi_lain,
                    "tipe": child.tipe,
                    "ukuran": child.ukuran,
                    "uraian": child.uraian,
                    "ndpbm": float(child.ndpbm),
                    "cifRupiah": float(child.cif_rupiah),
                    "hargaPerolehan": float(child.harga_perolehan),
                    "brutto": float(child.harga_perolehan),
                    "hargaPatokan": float(child.harga_perolehan),
                    "hjeCukai": float(child.harga_perolehan),
                    "jumlahBahanBaku": float(child.harga_perolehan),
                    "jumlahDilekatkan": float(child.harga_perolehan),
                    "jumlahPitaCukai": float(child.harga_perolehan),
                    "jumlahRealisasi": float(child.harga_perolehan),
                    "kapasitasSilinder": float(child.harga_perolehan),
                    "kodeKondisiBarang"
                    "nilaiDanaSawit": float(child.harga_perolehan),
                    "nilaiDevisa": float(child.harga_perolehan),
                    "pernyataanLartas"
                    "persentaseImpor": float(child.harga_perolehan),
                    "saldoAkhir": float(child.harga_perolehan),
                    "saldoAwal": float(child.harga_perolehan),
                    "seriBarangDokAsal"
                    "seriIjin"
                    "tahunPembuatan"
                    "tarifCukai": float(child.harga_perolehan),
                    "volume": float(child.harga_perolehan),
                    "barangTarif": [
                       {
                           "kodeJenisTarif": item.kode_tarif,
                           "jumlahSatuan": float(item.jumlah_satuan),
                           "kodeFasilitasTarif": item.kode_fasilitas,
                           "kodeSatuanBarang": item.kode_satuan,
                           "kodeJenisPungutan": item.kode_pungutan,
                           "nilaiBayar": float(item.nilai_bayar),
                           "nilaiFasilitas": float(item.nilai_fasilitas),
                           "seriBarang": int(item.seri_barang),
                           "tarif": float(item.tarif),
                           "tarifFasilitas": float(item.tarif_fasilitas)
                       }
                       for item in barang_tarif_data if item.parent == child.name
                    ],
                    "barangDokumen": [
                       {
                           "seriDokumen": item.seri_dokumen
                       }
                       for item in barang_dokumen_data if item.parent == child.name
                    ],
                    "barangSpekKhusus": [
                       {
                           "seriDokumen": item.seri_barang
                       }
                       for item in barang_spekkhusus_data if item.parent == child.name
                    ],
                    "barangPemilik": [
                       {
                           "seriDokumen": item.seri_barang
                       }
                       for item in barang_pemilik_data if item.parent == child.name
                    ],
                    "barangVd": [
                       {
                           "seriDokumen": item.seri_barang
                       }
                       for item in barang_vd_data if item.parent == child.name
                    ]
                }
                for child in barang_data
            ],
            "kemasan": [
                {
                    "jumlahKemasan": kemasan.jumlah_kemasan,
                    "kodeJenisKemasan": kemasan.kode_kemasan,
                    "merkKemasan": kemasan.merek_kemasan,
                    "seriKemasan": int(kemasan.seri)
                
                }
                for kemasan in kemasan_data
            ],
            "kontainer": [
                {
                    "kodeJenisKontainer": kontainer.kode_jenis_kontainer,
                    "kodeTipeKontainer": kontainer.kode_tipe_kontainer,
                    "kodeUkuranKontainer": kontainer.kode_ukuran_kontainer,
                    "nomorKontainer": kontainer.nomor_kontainer,
                    "seriKontainer": int(kontainer.seri)
                                    
                }
                for kontainer in kontainer_data
            ],
             "dokumen": [
                {
                    "kodeDokumen": dokumen.kode_dokumen,
                    "nomorDokumen": dokumen.nomor_dokumen,
                    "seriDokumen": int(dokumen.seri),
                    "tanggalDokumen": dokumen.tanggal_dokumen
                   
                    #"grandchildren": [
                    #    {
                    #        "grandchild_field1": grandchild.grandchild_field1,  # Alias for grandchild field 1
                    #        "grandchild_field2": grandchild.grandchild_field2  # Alias for grandchild field 2
                    #    }
                    #    for grandchild in grandchildren_data if grandchild.parent == child.name
                    #]
                }
                for dokumen in dokumen_data
            ],
             "pengangkut": [
                {
                    "callSign": pengangkut.call_sign, # Alias for child field 1
                    "kodeBendera": pengangkut.kode_bendera,
                    "namaPengangkut": pengangkut.nama_pengangkut,
                    "nomorPengangkut": pengangkut.nomor_pengangkut,
                    "kodeCaraAngkut": pengangkut.kode_cara_angkut,
                    "seriPengangkut": int(pengangkut.seri_pengangkut)
               
                    #"grandchildren": [
                    #    {
                    #        "grandchild_field1": grandchild.grandchild_field1,  # Alias for grandchild field 1
                    #        "grandchild_field2": grandchild.grandchild_field2  # Alias for grandchild field 2
                    #    }
                    #    for grandchild in grandchildren_data if grandchild.parent == child.name
                    #]
                }
                for pengangkut in pengangkut_data
            ],
       # }
    }

    return nested_data_bc20



@frappe.whitelist(allow_guest=True)
def get_ceisa_bc27_json(nomor_aju):
    try:
        doc = frappe.get_doc("HEADER V21", nomor_aju)
        
        # Helper to format date
        def fmt_date(date_obj):
            if not date_obj: return ""
            return str(date_obj)

        # Helper to get child table data
        def get_child_data(child_table_name, fields_map):
            data = []
            for child in doc.get(child_table_name, []):
                item = {}
                for json_field, doc_field in fields_map.items():
                    val = child.get(doc_field)
                    item[json_field] = val if val is not None else ""
                data.append(item)
            return data

        # 1. Map Header Fields
        payload = {
            "idPengguna": "", # Not in DocType
            "nomorAju": doc.nomoraju or doc.name,
            "tanggalAju": fmt_date(doc.tanggal_pernyataan), # Mapping to tanggal_pernyataan
            "asalData": doc.asaldata,
            "asuransi": doc.asuransi,
            "bruto": doc.bruto,
            "cif": doc.cif,
            "disclaimer": doc.disclaimer,
            "fob": doc.fob,
            "freight": doc.freight,
            "jabatanTtd": doc.jabatan_pernyataan,
            "jumlahKontainer": len(doc.get("kontainer", [])), # Calculated
            "kodeAsuransi": doc.kode_asuransi,
            "kodeCaraBayar": doc.kode_cara_bayar,
            "kodeDokumen": doc.kode_dokumen,
            "kodeIncoterm": doc.kode_incoterm,
            "kodeJenisNilai": doc.kode_jenis_nilai,
            "kodeJenisProsedur": doc.kode_jenis_pib, # Assuming mapping
            "kodeKantor": doc.kode_kantor,
            "kodePelMuat": doc.kode_pelabuhan_muat,
            "kodePelTujuan": doc.kode_pelabuhan_tujuan,
            "kodeTps": doc.kode_tps,
            "kodeValuta": doc.kode_valuta,
            "kotaTtd": doc.kota_pernyataan,
            "namaTtd": doc.nama_pernyataan,
            "ndpbm": doc.ndpbm,
            "netto": doc.netto,
            "nilaiMaklon": doc.nilai_maklon,
            "seri": 0, # Default or calculated?
            "tanggalTtd": fmt_date(doc.tanggal_pernyataan),
            "totalDanaSawit": doc.total_dana_sawit,
            "biayaPengurang": doc.biaya_pengurang,
            "biayaTambahan": doc.biaya_tambahan,
            "flagVd": doc.flag_vd,
            "hargaPenyerahan": doc.harga_penyerahan,
            "jumlahTandaPengaman": doc.jumlah_tanda_pengaman,
            "kodeJenisImpor": doc.kode_jenis_impor,
            "kodePelTransit": doc.kode_pelabuhan_transit,
            "kodeTutupPu": doc.kode_tutup_pu,
            "nilaiBarang": doc.nilai_barang,
            "nilaiIncoterm": doc.nilai_incoterm,
            "nomorBc11": doc.nomor_bc11,
            "posBc11": doc.nomor_pos,
            "subPosBc11": doc.nomor_sub_pos,
            "tanggalBc11": fmt_date(doc.tanggal_bc11),
            "tanggalTiba": fmt_date(doc.tanggal_tiba),
            "volume": doc.volume,
            "vd": doc.vd,
        }

        # 2. Map Child Tables (Entitas, Kemasan, Dokumen, Pengangkut)
        payload["entitas"] = get_child_data("entitas", {
            "alamatEntitas": "alamat_entitas",
            "kodeEntitas": "kode_entitas",
            "kodeJenisIdentitas": "kode_jenis_identitas",
            "namaEntitas": "nama_entitas",
            "nibEntitas": "nib_entitas",
            "nomorIdentitas": "nomor_identitas",
            "kodeStatus": "kode_status",
            "seriEntitas": "seri_entitas",
            "kodeJenisApi": "kode_jenis_api",
            "kodeNegara": "kode_negara",
            "kodeAfiliasi": "kode_afiliasi"
        })

        payload["kemasan"] = get_child_data("kemasan", {
            "jumlahKemasan": "jumlah_kemasan",
            "kodeJenisKemasan": "kode_jenis_kemasan",
            "merkKemasan": "merek_kemasan",
            "seriKemasan": "seri_kemasan"
        })

        payload["dokumen"] = get_child_data("dokumen", {
            "kodeDokumen": "kode_dokumen",
            "nomorDokumen": "nomor_dokumen",
            "seriDokumen": "seri_dokumen",
            "tanggalDokumen": "tanggal_dokumen"
        })

        payload["pengangkut"] = get_child_data("pengangkut", {
            "kodeBendera": "kode_bendera",
            "namaPengangkut": "nama_pengangkut",
            "nomorPengangkut": "nomor_pengangkut",
            "kodeCaraAngkut": "kode_cara_angkut",
            "seriPengangkut": "seri_pengangkut"
        })
        
        payload["kontainer"] = get_child_data("kontainer", {
            "kodeTipeKontainer": "kode_tipe_kontainer",
            "kodeUkuranKontainer": "kode_ukuran_kontainer",
            "nomorKontainer": "nomor_kontainer",
            "seriKontainer": "seri_kontainer",
            "kodeJenisKontainer": "kode_jenis_kontainer"
        })

        # 3. Map Barang V1
        barang_list = []
        # Fetch linked BARANG V1 items
        barangs = frappe.get_all("BARANG V1", filters={"nomoraju": doc.name}, fields=["*"], order_by="seri_barang asc")
        
        for brg in barangs:
            brg_item = {
                "cif": brg.cif,
                "cifRupiah": brg.cif_rupiah,
                "fob": brg.fob,
                "hargaEkspor": brg.harga_ekspor,
                "hargaPatokan": brg.harga_patokan,
                "hargaPerolehan": brg.harga_perolehan,
                "hargaSatuan": brg.harga_satuan,
                "jumlahKemasan": brg.jumlah_kemasan,
                "jumlahSatuan": brg.jumlah_satuan,
                "kodeJenisKemasan": brg.kode_kemasan,
                "kodeNegaraAsal": brg.kode_negara_asal,
                "kodeSatuanBarang": brg.kode_satuan,
                "merk": brg.merek,
                "ndpbm": brg.ndpbm,
                "netto": brg.netto,
                "nilaiBarang": brg.nilai_barang,
                "nilaiDanaSawit": brg.nilai_dana_sawit,
                "posTarif": brg.hs,
                "seriBarang": brg.seri_barang,
                "tipe": brg.tipe,
                "uraian": brg.uraian,
                "volume": brg.volume,
                "asuransi": brg.asuransi,
                "bruto": brg.bruto,
                "diskon": brg.diskon,
                "freight": brg.freight,
                "hargaPenyerahan": brg.harga_penyerahan,
                "hjeCukai": brg.hje_cukai,
                "isiPerKemasan": brg.isi_per_kemasan,
                "jumlahBahanBaku": brg.jumlah_bahan_baku,
                "jumlahDilekatkan": brg.jumlah_dilekatkan,
                "jumlahPitaCukai": brg.jumlah_pita_cukai,
                "jumlahRealisasi": brg.jumlah_realisasi,
                "kapasitasSilinder": brg.kapasitas_silinder,
                "kodeKondisiBarang": brg.kode_kondisi_barang,
                "nilaiDevisa": brg.nilai_devisa,
                "nilaiTambah": brg.nilai_tambah,
                "pernyataanLartas": brg.pernyataan_lartas,
                "persentaseImpor": brg.persentase_impor,
                "saldoAkhir": brg.saldo_akhir,
                "saldoAwal": brg.saldo_awal,
                "seriBarangDokAsal": brg.seri_barang_asal,
                "seriIjin": brg.seri_izin,
                "tahunPembuatan": brg.tahun_pembuatan,
                "tarifCukai": brg.tarif_cukai,
            }
            
            # Fetch Child Tables for this Barang
            # BARANG TARIF
            brg_doc = frappe.get_doc("BARANG V1", brg.name) # Need doc to get child tables easily
            
            brg_item["barangTarif"] = []
            for trf in brg_doc.get("barang_tarif", []):
                brg_item["barangTarif"].append({
                    "tarif": trf.tarif,
                    "nilaiBayar": trf.nilai_bayar,
                    "seriBarang": trf.seri_barang,
                    "kodeKemasan": "", # Not in BARANG TARIF
                    "jumlahSatuan": trf.jumlah_satuan,
                    "jumlahKemasan": 0, # Not in BARANG TARIF
                    "kodeJenisTarif": trf.kode_tarif,
                    "nilaiFasilitas": trf.nilai_fasilitas,
                    "tarifFasilitas": trf.tarif_fasilitas,
                    "kodeSatuanBarang": trf.kode_satuan,
                    "kodeJenisPungutan": trf.kode_pungutan,
                    "kodeKomoditiCukai": "", # Not in BARANG TARIF
                    "kodeFasilitasTarif": trf.kode_fasilitas,
                    "nilaiSudahDilunasi": trf.nilai_sudah_dilunasi,
                    "kodeSubKomoditiCukai": "" # Not in BARANG TARIF
                })

            # BARANG DOKUMEN
            brg_item["barangDokumen"] = []
            for dok in brg_doc.get("barang_dokumen", []):
                brg_item["barangDokumen"].append({
                    "seriDokumen": dok.seri_dokumen,
                    "seriIzin": dok.seri_izin
                })
                
            # BARANG SPEK KHUSUS
            brg_item["barangSpekKhusus"] = []
            for spek in brg_doc.get("barang_spek_khusus", []):
                brg_item["barangSpekKhusus"].append({
                    "kodeSpekKhusus": spek.kode_spek_khusus,
                    "uraian": spek.uraian
                })
                
            # BARANG VD
            brg_item["barangVd"] = []
            for vd in brg_doc.get("barang_vd", []):
                brg_item["barangVd"].append({
                    "kodeJenisVd": vd.kode_jenis_vd,
                    "nilaiBarang": vd.nilai_barang
                })
                
            # BARANG PEMILIK
            brg_item["barangPemilik"] = []
             # Assuming barang_pemilik exists in BARANG V1 based on previous file views
            for pem in brg_doc.get("barang_pemilik", []):
                 brg_item["barangPemilik"].append({
                     "seriEntitas": pem.seri_entitas
                 })

            # BAHAN BAKU (Linked Standard DocType, not child table)
            brg_item["bahanBaku"] = []
            bahan_bakus = frappe.get_all("BAHAN BAKU", 
                filters={"parent_barang": brg.name}, 
                fields=["*"],
                order_by="seri_bahan_baku asc"
            )
            for bb in bahan_bakus:
                bb_doc = frappe.get_doc("BAHAN BAKU", bb.name)
                bb_item = {
                    "seriBahanBaku": bb.seri_bahan_baku,
                    "kodeAsalBahanBaku": bb.kode_asal_bahan_baku,
                    "hs": bb.hs,
                    "kodeBarang": bb.kode_barang,
                    "uraian": bb.uraian,
                    "merek": bb.merek,
                    "tipe": bb.tipe,
                    "ukuran": bb.ukuran,
                    "spesifikasiLain": bb.spesifikasi_lain,
                    "kodeSatuan": bb.kode_satuan,
                    "jumlahSatuan": bb.jumlah_satuan,
                    "netto": bb.netto,
                    "bruto": bb.bruto,
                    "volume": bb.volume,
                    "cif": bb.cif,
                    "cifRupiah": bb.cif_rupiah,
                    "ndpbm": bb.ndpbm,
                    "hargaPenyerahan": bb.harga_penyerahan,
                    "hargaPerolehan": bb.harga_perolehan,
                    "kodeDokumenAsal": bb.kode_dokumen_asal,
                    "kodeKantorAsal": bb.kode_kantor_asal,
                    "nomorDaftarAsal": bb.nomor_daftar_asal,
                    "tanggalDaftarAsal": bb.tanggal_daftar_asal,
                    "nomorAjuAsal": bb.nomor_aju_asal,
                    "seriBarangAsal": bb.seri_barang_asal,
                }
                
                # BAHAN BAKU TARIF (child of BAHAN BAKU)
                bb_item["bahanBakuTarif"] = []
                for bbt in bb_doc.get("bahan_tarif", []):
                    bb_item["bahanBakuTarif"].append({
                        "kodePungutan": bbt.kode_pungutan,
                        "kodeTarif": bbt.kode_tarif,
                        "tarif": bbt.tarif,
                        "kodeFasilitas": bbt.kode_fasilitas,
                        "tarifFasilitas": bbt.tarif_fasilitas,
                        "nilaiBayar": bbt.nilai_bayar,
                        "nilaiFasilitas": bbt.nilai_fasilitas
                    })
                
                # BAHAN BAKU DOKUMEN (child of BAHAN BAKU)
                bb_item["bahanBakuDokumen"] = []
                for bbd in bb_doc.get("bahan_baku_dokumen", []):
                    bb_item["bahanBakuDokumen"].append({
                        "seriDokumen": bbd.seri_dokumen,
                        "seriIzin": bbd.seri_izin
                    })
                
                brg_item["bahanBaku"].append(bb_item)

            barang_list.append(brg_item)

        payload["barang"] = barang_list
        
        return payload

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Get CEISA BC20 JSON Error")
        return {"status": "error", "message": str(e)}

import openpyxl
import base64
import io
from frappe.utils import getdate, flt, cint

 
@frappe.whitelist(allow_guest=True)
def get_ceisa_bc20_json(nomor_aju):
    try:
        doc = frappe.get_doc("HEADER V21", nomor_aju)
        
        # Helper to format date
        def fmt_date(date_obj):
            if not date_obj: return ""
            return str(date_obj)

        # Helper to get child table data
        def get_child_data(child_table_name, fields_map):
            data = []
            for child in doc.get(child_table_name, []):
                item = {}
                for json_field, doc_field in fields_map.items():
                    val = child.get(doc_field)
                    item[json_field] = val if val is not None else ""
                data.append(item)
            return data

        # 1. Map Header Fields
        payload = {
            "idPengguna": "", # Not in DocType
            "nomorAju": doc.nomoraju or doc.name,
            "tanggalAju": fmt_date(doc.tanggal_pernyataan), # Mapping to tanggal_pernyataan
            "asalData": doc.asaldata,
            "asuransi": doc.asuransi,
            "bruto": doc.bruto,
            "cif": doc.cif,
            "disclaimer": doc.disclaimer,
            "fob": doc.fob,
            "freight": doc.freight,
            "jabatanTtd": doc.jabatan_pernyataan,
            "jumlahKontainer": len(doc.get("kontainer", [])), # Calculated
            "kodeAsuransi": doc.kode_asuransi,
            "kodeCaraBayar": doc.kode_cara_bayar,
            "kodeDokumen": doc.kode_dokumen,
            "kodeIncoterm": doc.kode_incoterm,
            "kodeJenisNilai": doc.kode_jenis_nilai,
            "kodeJenisProsedur": doc.kode_jenis_pib, # Assuming mapping
            "kodeKantor": doc.kode_kantor,
            "kodePelMuat": doc.kode_pelabuhan_muat,
            "kodePelTujuan": doc.kode_pelabuhan_tujuan,
            "kodeTps": doc.kode_tps,
            "kodeValuta": doc.kode_valuta,
            "kotaTtd": doc.kota_pernyataan,
            "namaTtd": doc.nama_pernyataan,
            "ndpbm": doc.ndpbm,
            "netto": doc.netto,
            "nilaiMaklon": doc.nilai_maklon,
            "seri": 0, # Default or calculated?
            "tanggalTtd": fmt_date(doc.tanggal_pernyataan),
            "totalDanaSawit": doc.total_dana_sawit,
            "biayaPengurang": doc.biaya_pengurang,
            "biayaTambahan": doc.biaya_tambahan,
            "flagVd": doc.flag_vd,
            "hargaPenyerahan": doc.harga_penyerahan,
            "jumlahTandaPengaman": doc.jumlah_tanda_pengaman,
            "kodeJenisImpor": doc.kode_jenis_impor,
            "kodePelTransit": doc.kode_pelabuhan_transit,
            "kodeTutupPu": doc.kode_tutup_pu,
            "nilaiBarang": doc.nilai_barang,
            "nilaiIncoterm": doc.nilai_incoterm,
            "nomorBc11": doc.nomor_bc11,
            "posBc11": doc.nomor_pos,
            "subPosBc11": doc.nomor_sub_pos,
            "tanggalBc11": fmt_date(doc.tanggal_bc11),
            "tanggalTiba": fmt_date(doc.tanggal_tiba),
            "volume": doc.volume,
            "vd": doc.vd,
        }

        # 2. Map Child Tables (Entitas, Kemasan, Dokumen, Pengangkut)
        payload["entitas"] = get_child_data("entitas", {
            "alamatEntitas": "alamat_entitas",
            "kodeEntitas": "kode_entitas",
            "kodeJenisIdentitas": "kode_jenis_identitas",
            "namaEntitas": "nama_entitas",
            "nibEntitas": "nib_entitas",
            "nomorIdentitas": "nomor_identitas",
            "kodeStatus": "kode_status",
            "seriEntitas": "seri_entitas",
            "kodeJenisApi": "kode_jenis_api",
            "kodeNegara": "kode_negara",
            "kodeAfiliasi": "kode_afiliasi"
        })

        payload["kemasan"] = get_child_data("kemasan", {
            "jumlahKemasan": "jumlah_kemasan",
            "kodeJenisKemasan": "kode_jenis_kemasan",
            "merkKemasan": "merek_kemasan",
            "seriKemasan": "seri_kemasan"
        })

        payload["dokumen"] = get_child_data("dokumen", {
            "kodeDokumen": "kode_dokumen",
            "nomorDokumen": "nomor_dokumen",
            "seriDokumen": "seri_dokumen",
            "tanggalDokumen": "tanggal_dokumen"
        })

        payload["pengangkut"] = get_child_data("pengangkut", {
            "kodeBendera": "kode_bendera",
            "namaPengangkut": "nama_pengangkut",
            "nomorPengangkut": "nomor_pengangkut",
            "kodeCaraAngkut": "kode_cara_angkut",
            "seriPengangkut": "seri_pengangkut"
        })
        
        payload["kontainer"] = get_child_data("kontainer", {
            "kodeTipeKontainer": "kode_tipe_kontainer",
            "kodeUkuranKontainer": "kode_ukuran_kontainer",
            "nomorKontainer": "nomor_kontainer",
            "seriKontainer": "seri_kontainer",
            "kodeJenisKontainer": "kode_jenis_kontainer"
        })

        # 3. Map Barang V1
        barang_list = []
        # Fetch linked BARANG V1 items
        barangs = frappe.get_all("BARANG V1", filters={"nomoraju": doc.name}, fields=["*"], order_by="seri_barang asc")
        
        for brg in barangs:
            brg_item = {
                "cif": brg.cif,
                "cifRupiah": brg.cif_rupiah,
                "fob": brg.fob,
                "hargaEkspor": brg.harga_ekspor,
                "hargaPatokan": brg.harga_patokan,
                "hargaPerolehan": brg.harga_perolehan,
                "hargaSatuan": brg.harga_satuan,
                "jumlahKemasan": brg.jumlah_kemasan,
                "jumlahSatuan": brg.jumlah_satuan,
                "kodeJenisKemasan": brg.kode_kemasan,
                "kodeNegaraAsal": brg.kode_negara_asal,
                "kodeSatuanBarang": brg.kode_satuan,
                "merk": brg.merek,
                "ndpbm": brg.ndpbm,
                "netto": brg.netto,
                "nilaiBarang": brg.nilai_barang,
                "nilaiDanaSawit": brg.nilai_dana_sawit,
                "posTarif": brg.hs,
                "seriBarang": brg.seri_barang,
                "tipe": brg.tipe,
                "uraian": brg.uraian,
                "volume": brg.volume,
                "asuransi": brg.asuransi,
                "bruto": brg.bruto,
                "diskon": brg.diskon,
                "freight": brg.freight,
                "hargaPenyerahan": brg.harga_penyerahan,
                "hjeCukai": brg.hje_cukai,
                "isiPerKemasan": brg.isi_per_kemasan,
                "jumlahBahanBaku": brg.jumlah_bahan_baku,
                "jumlahDilekatkan": brg.jumlah_dilekatkan,
                "jumlahPitaCukai": brg.jumlah_pita_cukai,
                "jumlahRealisasi": brg.jumlah_realisasi,
                "kapasitasSilinder": brg.kapasitas_silinder,
                "kodeKondisiBarang": brg.kode_kondisi_barang,
                "nilaiDevisa": brg.nilai_devisa,
                "nilaiTambah": brg.nilai_tambah,
                "pernyataanLartas": brg.pernyataan_lartas,
                "persentaseImpor": brg.persentase_impor,
                "saldoAkhir": brg.saldo_akhir,
                "saldoAwal": brg.saldo_awal,
                "seriBarangDokAsal": brg.seri_barang_asal,
                "seriIjin": brg.seri_izin,
                "tahunPembuatan": brg.tahun_pembuatan,
                "tarifCukai": brg.tarif_cukai,
            }
            
            # Fetch Child Tables for this Barang
            # BARANG TARIF
            brg_doc = frappe.get_doc("BARANG V1", brg.name) # Need doc to get child tables easily
            
            brg_item["barangTarif"] = []
            for trf in brg_doc.get("barang_tarif", []):
                brg_item["barangTarif"].append({
                    "tarif": trf.tarif,
                    "nilaiBayar": trf.nilai_bayar,
                    "seriBarang": trf.seri_barang,
                    "kodeKemasan": "", # Not in BARANG TARIF
                    "jumlahSatuan": trf.jumlah_satuan,
                    "jumlahKemasan": 0, # Not in BARANG TARIF
                    "kodeJenisTarif": trf.kode_tarif,
                    "nilaiFasilitas": trf.nilai_fasilitas,
                    "tarifFasilitas": trf.tarif_fasilitas,
                    "kodeSatuanBarang": trf.kode_satuan,
                    "kodeJenisPungutan": trf.kode_pungutan,
                    "kodeKomoditiCukai": "", # Not in BARANG TARIF
                    "kodeFasilitasTarif": trf.kode_fasilitas,
                    "nilaiSudahDilunasi": trf.nilai_sudah_dilunasi,
                    "kodeSubKomoditiCukai": "" # Not in BARANG TARIF
                })

            # BARANG DOKUMEN
            brg_item["barangDokumen"] = []
            for dok in brg_doc.get("barang_dokumen", []):
                brg_item["barangDokumen"].append({
                    "seriDokumen": dok.seri_dokumen,
                    "seriIzin": dok.seri_izin
                })
                
            # BARANG SPEK KHUSUS
            brg_item["barangSpekKhusus"] = []
            for spek in brg_doc.get("barang_spek_khusus", []):
                brg_item["barangSpekKhusus"].append({
                    "kodeSpekKhusus": spek.kode_spek_khusus,
                    "uraian": spek.uraian
                })
                
            # BARANG VD
            brg_item["barangVd"] = []
            for vd in brg_doc.get("barang_vd", []):
                brg_item["barangVd"].append({
                    "kodeJenisVd": vd.kode_jenis_vd,
                    "nilaiBarang": vd.nilai_barang
                })
                
            # BARANG PEMILIK
            brg_item["barangPemilik"] = []
             # Assuming barang_pemilik exists in BARANG V1 based on previous file views
            for pem in brg_doc.get("barang_pemilik", []):
                 brg_item["barangPemilik"].append({
                     "seriEntitas": pem.seri_entitas
                 })

            barang_list.append(brg_item)

        payload["barang"] = barang_list
        
        return payload

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Get CEISA BC20 JSON Error")
        return {"status": "error", "message": str(e)}

import openpyxl
import base64
import io
from frappe.utils import getdate, flt, cint

@frappe.whitelist(allow_guest=True)
def import_ceisa_excel(file_data):
    try:
        # Check if file_data is a path (e.g., /private/files/...)
        if file_data.startswith("/private/files/") or file_data.startswith("/files/"):
            file_path = frappe.get_site_path(file_data.strip("/"))
            with open(file_path, "rb") as f:
                decoded_file = f.read()
        else:
            # Assume base64
            if "," in file_data:
                file_data = file_data.split(",")[1]
            decoded_file = base64.b64decode(file_data)
        
        # Load workbook
        wb = openpyxl.load_workbook(io.BytesIO(decoded_file), data_only=True)
        
        # Helper to get value from sheet
        def get_sheet_data(sheet_name, optional=False):
            if sheet_name not in wb.sheetnames:
                if not optional:
                    frappe.log_error(f"Sheet '{sheet_name}' not found in workbook. Available: {wb.sheetnames}", "Excel Import Warning")
                return []
            ws = wb[sheet_name]
            data = []
            headers = [cell.value for cell in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                item = dict(zip(headers, row))
                # Filter out empty rows
                if any(item.values()):
                    data.append(item)
            return data

        # 1. Create Header
        header_data = get_sheet_data("HEADER")
        if not header_data:
            return {"status": "error", "message": f"Sheet 'HEADER' is empty or missing. Available sheets: {', '.join(wb.sheetnames)}"}
        
        h_row = header_data[0]
        
        # Column mapping: Excel Column -> DocType Field
        HEADER_MAPPING = {
            "NOMOR AJU": "nomoraju",
            "KODE DOKUMEN": "kode_dokumen",
            "KODE KANTOR": "kode_kantor",
            "KODE KANTOR BONGKAR": "kode_kantor_bongkar",
            "KODE KANTOR PERIKSA": "kode_kantor_periksa",
            "KODE KANTOR TUJUAN": "kode_kantor_tujuan",
            "KODE KANTOR EKSPOR": "kode_kantor_ekspor",
            "KODE JENIS IMPOR": "kode_jenis_impor",
            "KODE JENIS EKSPOR": "kode_jenis_ekspor",
            "KODE JENIS TPB": "kode_jenis_tpb",
            "KODE JENIS PLB": "kode_jenis_plb",
            "KODE JENIS PROSEDUR": "kode_jenis_pib",
            "KODE TUJUAN PEMASUKAN": "kode_tujuan_pemasukan",
            "KODE TUJUAN PENGIRIMAN": "kode_tujuan_pengiriman_value",
            "KODE TUJUAN TPB": "kode_tujuan_tpb",
            "KODE CARA DAGANG": "kode_cara_dagang",
            "KODE CARA BAYAR": "kode_cara_bayar",
            "KODE CARA BAYAR LAINNYA": "kode_cara_bayar_lainnya",
            "KODE GUDANG ASAL": "kode_gudang_asal",
            "KODE GUDANG TUJUAN": "kode_gudang_tujuan",
            "KODE JENIS KIRIM": "kode_jenis_kirim",
            "KODE JENIS PENGIRIMAN": "kode_jenis_pengiriman",
            "KODE KATEGORI EKSPOR": "kode_kategori_ekspor",
            "KODE KATEGORI MASUK FTZ": "kode_kategori_masuk_ftz",
            "KODE KATEGORI KELUAR FTZ": "kode_kategori_keluar_ftz",
            "KODE KATEGORI BARANG FTZ": "kode_kategori_barang_ftz",
            "KODE LOKASI": "kode_lokasi",
            "KODE LOKASI BAYAR": "kode_lokasi_bayar",
            "LOKASI ASAL": "lokasi_asal",
            "LOKASI TUJUAN": "lokasi_tujuan",
            "KODE DAERAH ASAL": "kode_daerah_asal",
            "KODE NEGARA TUJUAN": "kode_negara_tujuan",
            "KODE TUTUP PU": "kode_tutup_pu",
            "NOMOR BC11": "nomor_bc11",
            "TANGGAL BC11": "tanggal_bc11",
            "NOMOR POS": "nomor_pos",
            "NOMOR SUB POS": "nomor_sub_pos",
            "KODE PELABUHAN BONGKAR": "kode_pelabuhan_bongkar",
            "KODE PELABUHAN MUAT": "kode_pelabuhan_muat",
            "KODE PELABUHAN MUAT AKHIR": "kode_pelabuhan_muat_akhir",
            "KODE PELABUHAN TRANSIT": "kode_pelabuhan_transit",
            "KODE PELABUHAN TUJUAN": "kode_pelabuhan_tujuan",
            "KODE PELABUHAN EKSPOR": "kode_pelabuhan_ekspor",
            "KODE TPS": "kode_tps",
            "TANGGAL BERANGKAT": "tanggal_berangkat",
            "TANGGAL EKSPOR": "tanggal_ekspor",
            "TANGGAL MASUK": "tanggal_masuk",
            "TANGGAL MUAT": "tanggal_muat",
            "TANGGAL TIBA": "tanggal_tiba",
            "TANGGAL PERIKSA": "tanggal_periksa",
            "TEMPAT STUFFING": "tempat_stuffing",
            "TANGGAL STUFFING": "tanggal_stuffing",
            "KODE TANDA PENGAMAN": "kode_tanda_pengaman",
            "JUMLAH TANDA PENGAMAN": "jumlah_tanda_pengaman",
            "FLAG CURAH": "flag_curah",
            "FLAG SDA": "flag_sda",
            "FLAG VD": "flag_vd",
            "FLAG MIGAS": "flag_migas",
            "KODE ASURANSI": "kode_asuransi",
            "ASURANSI": "asuransi",
            "NILAI BARANG": "nilai_barang",
            "NILAI INCOTERM": "nilai_incoterm",
            "NILAI MAKLON": "nilai_maklon",
            "FREIGHT": "freight",
            "FOB": "fob",
            "BIAYA TAMBAHAN": "biaya_tambahan",
            "BIAYA PENGURANG": "biaya_pengurang",
            "VD": "vd",
            "CIF": "cif",
            "HARGA_PENYERAHAN": "harga_penyerahan",
            "NDPBM": "ndpbm",
            "TOTAL DANA SAWIT": "total_dana_sawit",
            "DASAR PENGENAAN PAJAK": "dasar_pengenaan_pajak",
            "NILAI JASA": "nilai_jasa",
            "UANG MUKA": "uang_muka",
            "BRUTO": "bruto",
            "NETTO": "netto",
            "VOLUME": "volume",
            "KOTA PERNYATAAN": "kota_pernyataan",
            "TANGGAL PERNYATAAN": "tanggal_pernyataan",
            "NAMA PERNYATAAN": "nama_pernyataan",
            "JABATAN PERNYATAAN": "jabatan_pernyataan",
            "KODE VALUTA": "kode_valuta",
            "KODE INCOTERM": "kode_incoterm",
            "KODE JASA KENA PAJAK": "kode_jasa_kena_pajak",
            "NOMOR BUKTI BAYAR": "nomor_bukti_bayar",
            "TANGGAL BUKTI BAYAR": "tanggal_bukti_bayar",
            "KODE JENIS NILAI": "kode_jenis_nilai",
            "KODE KANTOR MUAT": "kode_kantor_muat",
            "NOMOR DAFTAR": "nomor_daftar",
            "TANGGAL DAFTAR": "tanggal_daftar",
            "KODE ASAL BARANG FTZ": "kode_barang_asal_ftz",
            "KODE TUJUAN PENGELUARAN": "kode_tujuan_pengeluaran",
            "PPN PAJAK": "ppn_pajak",
            "PPNBM PAJAK": "ppnbm_pajak",
            "TARIF PPN PAJAK": "tarif_ppn_pajak",
            "TARIF PPNBM PAJAK": "tarif_ppnbm_pajak",
            "BARANG TIDAK BERWUJUD": "barang_tidak_berwujud",
        }
        
        # Date fields that need special handling
        DATE_FIELDS = [
            "tanggal_bc11", "tanggal_berangkat", "tanggal_ekspor", "tanggal_masuk",
            "tanggal_muat", "tanggal_tiba", "tanggal_periksa", "tanggal_stuffing",
            "tanggal_pernyataan", "tanggal_bukti_bayar", "tanggal_daftar"
        ]
        
        # Get nomorAju
        nomor_aju = h_row.get("NOMOR AJU")
        if not nomor_aju:
            return {"status": "error", "message": "NOMOR AJU is required in HEADER sheet"}

        # Check if exists
        existing_header = frappe.get_all("HEADER V21", filters={"nomoraju": nomor_aju}, limit=1)
        if existing_header:
             doc = frappe.get_doc("HEADER V21", existing_header[0].name)
        else:
            doc = frappe.new_doc("HEADER V21")
            doc.name = nomor_aju  # Set document name to nomoraju
            doc.nomoraju = nomor_aju
            
        # Map all Header Fields dynamically
        for excel_col, doc_field in HEADER_MAPPING.items():
            value = h_row.get(excel_col)
            if value is not None and value != "":  # Skip null/empty values
                if doc_field in DATE_FIELDS:
                    try:
                        doc.set(doc_field, getdate(value))
                    except:
                        pass  # Skip invalid dates
                elif isinstance(value, (int, float)):
                    doc.set(doc_field, flt(value))
                else:
                    doc.set(doc_field, str(value))
        
        doc.flags.ignore_links = True
        doc.save(ignore_permissions=True)
        
        # Rename document to use nomoraju as name if it's different
        if doc.name != nomor_aju:
            try:
                frappe.rename_doc("HEADER V21", doc.name, nomor_aju, force=True, merge=False)
                doc = frappe.get_doc("HEADER V21", nomor_aju)  # Reload after rename
            except Exception as rename_err:
                frappe.log_error(f"Could not rename HEADER V21 to {nomor_aju}: {rename_err}", "Rename Error")
        
        # Helper to create child table
        def create_child(doctype, parent_field, sheet_name, mapping, optional=False):
            # Clear existing?
            # frappe.db.delete(doctype, {parent_field: doc.name}) 
            # Better to use doc.set(field, []) if it was a child table field, but these are separate tables linked by parent?
            # Based on previous code, they seem to be child tables in DocType definition (Table field)
            # BUT the API `get_ceisa_bc20_json` treated them as child tables `doc.get("entitas")`.
            # So we should populate them via the parent doc.
            
            rows = get_sheet_data(sheet_name, optional)
            child_list = []
            for row in rows:
                child_item = {}
                for excel_col, doc_field in mapping.items():
                    val = row.get(excel_col)
                    if val is not None:
                        child_item[doc_field] = val
                child_list.append(child_item)
            
            doc.set(doctype, child_list)

        # 2. Child Tables - Update mappings to match Excel uppercase format
        create_child("entitas", "parent", "ENTITAS", {
            "KODE ENTITAS": "kode_entitas",
            "NAMA ENTITAS": "nama_entitas",
            "ALAMAT ENTITAS": "alamat_entitas",
            "NOMOR IDENTITAS": "nomor_identitas",
            "SERI ENTITAS": "seri_entitas",
            "KODE JENIS IDENTITAS": "kode_jenis_identitas",
            "NIB ENTITAS": "nib_entitas",
            "KODE JENIS API": "kode_jenis_api",
            "KODE NEGARA": "kode_negara"
        })
        
        create_child("kemasan", "parent", "KEMASAN", {
            "JUMLAH KEMASAN": "jumlah_kemasan",
            "KODE JENIS KEMASAN": "kode_jenis_kemasan",
            "MERK KEMASAN": "merek_kemasan",
            "SERI KEMASAN": "seri_kemasan"
        })
        
        create_child("dokumen", "parent", "DOKUMEN", {
            "KODE DOKUMEN": "kode_dokumen",
            "NOMOR DOKUMEN": "nomor_dokumen",
            "TANGGAL DOKUMEN": "tanggal_dokumen",
            "SERI DOKUMEN": "seri_dokumen"
        })
        
        create_child("pengangkut", "parent", "PENGANGKUT", {
            "NAMA PENGANGKUT": "nama_pengangkut",
            "NOMOR PENGANGKUT": "nomor_pengangkut",
            "KODE CARA ANGKUT": "kode_cara_angkut",
            "KODE BENDERA": "kode_bendera",
            "SERI PENGANGKUT": "seri_pengangkut"
        })
        
        create_child("kontainer", "parent", "KONTAINER", {
            "KODE TIPE KONTAINER": "kode_tipe_kontainer",
            "KODE UKURAN KONTAINER": "kode_ukuran_kontainer",
            "NOMOR KONTAINER": "nomor_kontainer",
            "SERI KONTAINER": "seri_kontainer",
            "KODE JENIS KONTAINER": "kode_jenis_kontainer"
        })

        # Additional HEADER child tables
        # PUNGUTAN
        create_child("pungutan", "parent", "PUNGUTAN", {
            "KODE FASILITAS TARIF": "kode_fasilitas_tarif",
            "KODE JENIS PUNGUTAN": "kode_jenis_pungutan",
            "NILAI PUNGUTAN": "nilai_pungutan",
            "NPWP BILLING": "npwp_billing"
        })
        
        # JAMINAN
        create_child("jaminan", "parent", "JAMINAN", {
            "KODE KANTOR": "kode_kantor",
            "KODE JAMINAN": "kode_jaminan",
            "NOMOR JAMINAN": "nomor_jaminan",
            "TANGGAL JAMINAN": "tanggal_jaminan",
            "NILAI JAMINAN": "nilai_jaminan",
            "PENJAMIN": "penjamin",
            "TANGGAL JATUH TEMPO": "tanggal_jatuh_tempo",
            "NOMOR BPJ": "nomor_bpj",
            "TANGGAL BPJ": "tanggal_bpj"
        })
        
        # BANKDEVISA
        create_child("bank_devisa", "parent", "BANKDEVISA", {
            "SERI": "seri",
            "KODE": "kode",
            "NAMA": "nama"
        })
        
        # KOMPONENBIAYA (optional, might not exist in all files)
        create_child("komponen_biaya", "parent", "KOMPONENBIAYA", {
           "JENIS NILAI" : "jenisNilai",
            "HARGA INVOICE": "hargaInvoice",
            "PEMBAYARAN TIDAK LANGSUNG": "pembayaranTidakLangsung",
            "DISKON": "diskon",
            "KOMISI PENJUALAN": "komisiPenjualan",
            "BIAYA PENGEMASAN": "biayaPengemasan",
            "BIAYA PENGEPAKAN": "biayaPengepakan",
            "ASSIST": "assist",
            "ROYALTI": "royalti",
            "PROCEEDS": "proceeds",
            "BIAYA TRANSPORTASI": "biayaTransportasi",
            "BIAYA PEMUATAN": "biayaPemuatan",
            "ASURANSI": "asuransi",
            "GARANSI": "garansi",
            "BIAYA KEPENTINGAN SENDIRI": "biayaKepentinganSendiri",
            "BIAYA PASCA IMPOR": "biayaPascaImpor",
            "BIAYA PAJAK INTERNAL": "biayaPajakInternal",
            "BUNGA": "bunga",
            "DEVIDEN": "deviden"
        }, optional=True)

        doc.save(ignore_permissions=True)
        
        # 3. Barang V1 - Import from BARANG sheet
        barang_rows = get_sheet_data("BARANG")
        
        # Barang mapping
        BARANG_MAPPING = {
            "SERI BARANG": "seri_barang",
            "HS": "hs",
            "KODE BARANG": "kode_barang",
            "URAIAN": "uraian",
            "MERK": "merek",
            "TIPE": "tipe",
            "UKURAN": "ukuran",
            "SPESIFIKASI LAIN": "spesifikasi_lain",
            "KODE SATUAN": "kode_satuan",
            "JUMLAH SATUAN": "jumlah_satuan",
            "KODE KEMASAN": "kode_kemasan",
            "JUMLAH KEMASAN": "jumlah_kemasan",
            "NETTO": "netto",
            "BRUTO": "bruto",
            "VOLUME": "volume",
            "CIF": "cif",
            "CIF RUPIAH": "cif_rupiah",
            "NDPBM": "ndpbm",
            "FOB": "fob",
            "ASURANSI": "asuransi",
            "FREIGHT": "freight",
            "NILAI TAMBAH": "nilai_tambah",
            "DISKON": "diskon",
            "HARGA PENYERAHAN": "harga_penyerahan",
            "HARGA PEROLEHAN": "harga_perolehan",
            "HARGA SATUAN": "harga_satuan",
            "HARGA EKSPOR": "harga_ekspor",
            "HARGA PATOKAN": "harga_patokan",
            "NILAI BARANG": "nilai_barang",
            "NILAI DANA SAWIT": "nilai_dana_sawit",
            "NILAI DEVISA": "nilai_devisa",
            "KODE NEGARA ASAL": "kode_negara_asal",
            "KODE KONDISI BARANG": "kode_kondisi_barang",
            "PERNYATAAN LARTAS": "pernyataan_lartas"
        }
        
        for b_row in barang_rows:
            seri_barang = cint(b_row.get("SERI BARANG"))
            if not seri_barang: 
                continue
            
            # Check if exists
            existing_barang = frappe.get_all("BARANG V1", filters={"nomoraju": nomor_aju, "seri_barang": seri_barang})
            
            if existing_barang:
                b_doc = frappe.get_doc("BARANG V1", existing_barang[0].name)
            else:
                b_doc = frappe.new_doc("BARANG V1")
                b_doc.nomoraju = nomor_aju
                b_doc.seri_barang = seri_barang
            
            # Map all Barang Fields dynamically
            for excel_col, doc_field in BARANG_MAPPING.items():
                value = b_row.get(excel_col)
                if value is not None and value != "":
                    if isinstance(value, (int, float)):
                        b_doc.set(doc_field, flt(value))
                    else:
                        b_doc.set(doc_field, str(value))
            
            b_doc.flags.ignore_links = True
            b_doc.save(ignore_permissions=True)
            
            # Barang Child Tables
            # BARANGTARIF
            tarif_rows = get_sheet_data("BARANGTARIF")
            b_tarifs = []
            for t_row in tarif_rows:
                if cint(t_row.get("SERI BARANG")) == seri_barang:
                    b_tarifs.append({
                        "seri_barang": cint(t_row.get("SERI BARANG")),
                        "kode_pungutan": t_row.get("KODE PUNGUTAN"),
                        "kode_tarif": t_row.get("KODE TARIF"),
                        "tarif": flt(t_row.get("TARIF")),
                        "nilai_bayar": flt(t_row.get("NILAI BAYAR")),
                        "nilai_fasilitas": flt(t_row.get("NILAI FASILITAS")),
                        "tarif_fasilitas": flt(t_row.get("TARIF FASILITAS")),
                        "kode_fasilitas": t_row.get("KODE FASILITAS"),
                        "jumlah_satuan": flt(t_row.get("JUMLAH SATUAN")),
                        "kode_satuan": t_row.get("KODE SATUAN")
                    })
            if b_tarifs:
                b_doc.set("barang_tarif", b_tarifs)
            
            # BARANGDOKUMEN
            bd_rows = get_sheet_data("BARANGDOKUMEN")
            b_docs = []
            for d_row in bd_rows:
                if cint(d_row.get("SERI BARANG")) == seri_barang:
                    b_docs.append({
                        "seri_dokumen": d_row.get("SERI DOKUMEN"),
                        "seri_izin": d_row.get("SERI IZIN")
                    })
            if b_docs:
                b_doc.set("barang_dokumen", b_docs)
            
            # BARANGENTITAS
            be_rows = get_sheet_data("BARANGENTITAS")
            b_entitas = []
            for e_row in be_rows:
                if cint(e_row.get("SERI BARANG")) == seri_barang:
                    b_entitas.append({
                        "seri_entitas": cint(e_row.get("SERI ENTITAS"))
                    })
            if b_entitas:
                b_doc.set("barang_pemilik", b_entitas)
            
            # BARANGSPEKKHUSUS
            bsk_rows = get_sheet_data("BARANGSPEKKHUSUS")
            b_spek = []
            for sk_row in bsk_rows:
                if cint(sk_row.get("SERI BARANG")) == seri_barang:
                    b_spek.append({
                        "kode_spek_khusus": sk_row.get("KODE"),
                        "uraian": sk_row.get("URAIAN")
                    })
            if b_spek:
                b_doc.set("barang_spek_khusus", b_spek)
            
            # BARANGVD
            bvd_rows = get_sheet_data("BARANGVD")
            b_vd = []
            for vd_row in bvd_rows:
                if cint(vd_row.get("SERI BARANG")) == seri_barang:
                    b_vd.append({
                        "kode_jenis_vd": vd_row.get("KODE VD"),
                        "nilai_barang": flt(vd_row.get("NILAI BARANG"))
                    })
            if b_vd:
                b_doc.set("barang_vd", b_vd)
            
            # Save BARANG V1 with all child tables
            b_doc.flags.ignore_links = True
            b_doc.save(ignore_permissions=True)
            
            # BAHANBAKU (Linked Standard DocType)
            # We import this AFTER saving BARANG V1 so we can link it
            bb_rows = get_sheet_data("BAHANBAKU", optional=True)
            for bb_row in bb_rows:
                if cint(bb_row.get("SERI BARANG")) == seri_barang:
                    seri_bahan_baku = cint(bb_row.get("SERI BAHAN BAKU"))
                    
                    # Check if exists
                    existing_bb = frappe.get_all("BAHAN BAKU", filters={
                        "nomoraju": nomor_aju, 
                        "seri_barang": seri_barang,
                        "seri_bahan_baku": seri_bahan_baku
                    })
                    
                    if existing_bb:
                        bb_doc = frappe.get_doc("BAHAN BAKU", existing_bb[0].name)
                    else:
                        bb_doc = frappe.new_doc("BAHAN BAKU")
                        bb_doc.nomoraju = nomor_aju
                        bb_doc.seri_barang = seri_barang
                        bb_doc.seri_bahan_baku = seri_bahan_baku

                    # Map Fields
                    bb_doc.parent_barang = b_doc.name # Link to specific BARANG V1
                    bb_doc.kode_asal_bahan_baku = bb_row.get("KODE ASAL BAHAN BAKU")
                    bb_doc.hs = bb_row.get("HS")
                    bb_doc.kode_barang = bb_row.get("KODE BARANG")
                    bb_doc.uraian = bb_row.get("URAIAN")
                    bb_doc.merek = bb_row.get("MEREK")
                    bb_doc.tipe = bb_row.get("TIPE")
                    bb_doc.ukuran = bb_row.get("UKURAN")
                    bb_doc.spesifikasi_lain = bb_row.get("SPESIFIKASI LAIN")
                    bb_doc.kode_satuan = bb_row.get("KODE SATUAN")
                    bb_doc.jumlah_satuan = flt(bb_row.get("JUMLAH SATUAN"))
                    bb_doc.netto = flt(bb_row.get("NETTO"))
                    
                    # BAHANBAKUTARIF (child of BAHANBAKU)
                    bbt_rows = get_sheet_data("BAHANBAKUTARIF", optional=True)
                    bb_tarifs = []
                    for bbt_row in bbt_rows:
                        if (cint(bbt_row.get("SERI BARANG")) == seri_barang and 
                            cint(bbt_row.get("SERI BAHAN BAKU")) == seri_bahan_baku):
                            bb_tarifs.append({
                                "kode_pungutan": bbt_row.get("KODE PUNGUTAN"),
                                "kode_tarif": bbt_row.get("KODE TARIF"),
                                "tarif": flt(bbt_row.get("TARIF")),
                                "kode_fasilitas": bbt_row.get("KODE FASILITAS"),
                                "tarif_fasilitas": flt(bbt_row.get("TARIF FASILITAS")),
                                "nilai_bayar": flt(bbt_row.get("NILAI BAYAR")),
                                "nilai_fasilitas": flt(bbt_row.get("NILAI FASILITAS"))
                            })
                    if bb_tarifs:
                        bb_doc.set("bahan_tarif", bb_tarifs) # Fieldname in JSON is 'bahan_tarif'
                    
                    # BAHANBAKUDOKUMEN (child of BAHANBAKU)
                    bbd_rows = get_sheet_data("BAHANBAKUDOKUMEN", optional=True)
                    bb_docs = []
                    for bbd_row in bbd_rows:
                        if (cint(bbd_row.get("SERI BARANG")) == seri_barang and 
                            cint(bbd_row.get("SERI BAHAN BAKU")) == seri_bahan_baku):
                            bb_docs.append({
                                "seri_dokumen": bbd_row.get("SERI DOKUMEN"),
                                "seri_izin": bbd_row.get("SERI IZIN")
                            })
                    if bb_docs:
                        bb_doc.set("bahan_baku_dokumen", bb_docs)
                    
                    bb_doc.flags.ignore_links = True
                    bb_doc.save(ignore_permissions=True)


        return {"status": "success", "message": f"Successfully imported Header {nomor_aju} with {len(barang_rows)} barang items"}


    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Import CEISA Excel Error")
        return {"status": "error", "message": str(e)}


BEACUKAI_BASE_URL = "https://apis-gw.beacukai.go.id"

def get_cached_token():
    return frappe.cache().hget("beacukai_token", frappe.session.user)

@frappe.whitelist()
def login_beacukai(username, password):
    url = f"{BEACUKAI_BASE_URL}/nle-oauth/v1/user/login"
    payload = {
        "username": username,
        "password": password
    }
    try:
        response = requests.post(url, json=payload)
        response.raise_for_status()
        data = response.json()
        
        token = None
        if "item" in data and "access_token" in data["item"]:
             token = data["item"]["access_token"]
        elif "access_token" in data:
             token = data["access_token"]
        
        if token:
            frappe.cache().hset("beacukai_token", frappe.session.user, token)
            return {"status": "success", "message": "Login successful"}
        else:
            return {"status": "error", "message": "Token not found", "response": data}
            
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Beacukai Login Error")
        return {"status": "error", "message": str(e)}

@frappe.whitelist()
def check_ceisa_status(nomor_aju):
    token = get_cached_token()
    if not token:
        return {"status": "error", "message": "Please login to Beacukai first."}
        
    url = f"{BEACUKAI_BASE_URL}/openapi/status/{nomor_aju}"
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json"
    }
    
    try:
        response = requests.get(url, headers=headers)
        return {
            "status": "success" if response.status_code == 200 else "error",
            "http_code": response.status_code,
            "response": response.json() if response.content else response.text
        }
    except Exception as e:
        return {"status": "error", "message": str(e)}

@frappe.whitelist()
def send_ceisa_document(docname):
    token = get_cached_token()
    if not token:
        return {"status": "error", "message": "Please login to Beacukai first."}
    
    try:
        # Use existing function to get payload
        # Assuming docname passed is the nomor_aju or name
        payload = get_ceisa_bc20_json(docname)
        
        if isinstance(payload, dict) and payload.get("status") == "error":
             return payload # Propagate error from get_ceisa_bc20_json
        
        url = f"{BEACUKAI_BASE_URL}/openapi/document?isFinal=false"
        headers = {
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json"
        }
        
        response = requests.post(url, json=payload, headers=headers)
        
        return {
            "status": "success" if response.status_code == 200 else "error",
            "http_code": response.status_code,
            "response": response.json() if response.content else response.text
        }
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Send CEISA Document Error")
        return {"status": "error", "message": str(e)}
