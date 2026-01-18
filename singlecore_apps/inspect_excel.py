import openpyxl
import os
import frappe

def main():
    files = [
        "/home/acer25/frappe-bench/sites/dens9.com/private/files/00026272017220251017000063.xlsx",
        "/home/acer25/frappe-bench/sites/dens9.com/private/files/00002021115720251119001216.xlsx"
    ]

    for file_path in files:
        print(f"\n--- Inspecting: {os.path.basename(file_path)} ---")
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            print(f"Sheets ({len(wb.sheetnames)}): {wb.sheetnames}")
            
            # Show first few rows of each sheet
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                print(f"\n  Sheet: {sheet_name}")
                headers = [cell.value for cell in ws[1]]
                print(f"  Headers: {headers[:10]}...")  # First 10 columns
                
        except Exception as e:
            print(f"Error: {e}")
            import traceback
            traceback.print_exc()
